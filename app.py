from flask import Flask, render_template, redirect, url_for, flash, request, send_file, jsonify
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
from config import Config
from models import db, User, Department, Publication, Notification, AuditLog, EditRequest, Feedback, IncentiveConfig, WalletTransaction, ApplicationForm, FAQ
from forms import RegistrationForm, LoginForm, PublicationForm, ChangePasswordForm
import os
from functools import wraps
from datetime import datetime
import pandas as pd
from io import BytesIO
from sqlalchemy import func, extract
import json
from fuzzywuzzy import fuzz
import re
import base64
import pytz
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.interval import IntervalTrigger
import atexit
from pathlib import Path

# Indian Standard Time (Asia/Kolkata) timezone
IST = pytz.timezone('Asia/Kolkata')

def now_ist():
    """Get current time in Asia/Kolkata timezone"""
    return datetime.now(IST)

app = Flask(__name__)
app.config.from_object(Config)

# Initialize extensions
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Add custom Jinja2 filter for base64 encoding
@app.template_filter('b64encode')
def b64encode_filter(s):
    if s is None:
        return ''
    # If already bytes, encode directly; if string, encode to bytes first
    if isinstance(s, bytes):
        return base64.b64encode(s).decode('utf-8')
    return base64.b64encode(s.encode('utf-8')).decode('utf-8')

# Add custom Jinja2 filter for Indian number formatting
@app.template_filter('indian_format')
def indian_format_filter(number):
    """Format number in Indian numbering system (lakhs, crores)"""
    if number is None:
        return '0'
    
    # Convert to integer if it's a float
    number = int(number)
    
    # Convert to string and handle negative numbers
    s = str(abs(number))
    
    if len(s) <= 3:
        formatted = s
    else:
        # Last 3 digits
        formatted = s[-3:]
        s = s[:-3]
        
        # Add commas every 2 digits for remaining
        while s:
            if len(s) <= 2:
                formatted = s + ',' + formatted
                s = ''
            else:
                formatted = s[-2:] + ',' + formatted
                s = s[:-2]
    
    # Add back negative sign if needed
    if number < 0:
        formatted = '-' + formatted
    
    return formatted

# Add custom Jinja2 filter to convert datetime to IST
@app.template_filter('to_ist')
def to_ist_filter(dt):
    """Convert datetime to IST timezone"""
    if dt is None:
        return None
    # If datetime is naive (no timezone), assume it's UTC
    if dt.tzinfo is None:
        dt = pytz.utc.localize(dt)
    # Convert to IST
    return dt.astimezone(IST)

# Add custom Jinja2 filter to parse JSON strings
@app.template_filter('from_json')
def from_json_filter(s):
    """Parse JSON string to Python object"""
    if s is None or s == '':
        return {}
    try:
        return json.loads(s)
    except:
        return {}

# Add custom Jinja2 filter for number formatting with commas
@app.template_filter('format_number')
def format_number_filter(number):
    """Format number with commas (e.g., 1,234,567)"""
    if number is None:
        return '0'
    try:
        return "{:,}".format(int(number))
    except (ValueError, TypeError):
        return str(number)

# Create upload folders if they don't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['PUBLICATIONS_FOLDER'], exist_ok=True)
os.makedirs(app.config['PROFILES_FOLDER'], exist_ok=True)
os.makedirs(app.config['SIGNATURES_FOLDER'], exist_ok=True)
os.makedirs(app.config['FEEDBACK_SCREENSHOTS_FOLDER'], exist_ok=True)
os.makedirs(app.config['CIRCULARS_FOLDER'], exist_ok=True)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Initialize Background Scheduler for auto-forwarding ACM applications
scheduler = BackgroundScheduler()

def auto_forward_acm_applications():
    """Auto-forward applications with 2 ACM approvals after 24 hours"""
    with app.app_context():
        try:
            from models import ApplicationForm, ACMReview, ACMPanel
            from notification_utils import create_notification
            from datetime import timedelta
            
            # Find applications in acm_review status with 2 approvals
            applications = ApplicationForm.query.filter_by(
                status='acm_review',
                acm_approvals_count=2
            ).all()
            
            forwarded_count = 0
            
            for application in applications:
                if application.acm_review_started_at:
                    time_in_review = now_ist() - application.acm_review_started_at.replace(tzinfo=IST)
                    
                    # If more than 1 day (24 hours) has passed
                    if time_in_review >= timedelta(days=1):
                        all_reviews = ACMReview.query.filter_by(
                            application_id=application.id
                        ).all()
                        
                        total_acm_members = ACMPanel.query.filter_by(is_active=True).count()
                        ineligible_members = ACMPanel.query.filter_by(
                            member_id=application.applicant_id,
                            is_active=True
                        ).count()
                        
                        eligible_acm_count = total_acm_members - ineligible_members
                        total_reviews = len(all_reviews)
                        
                        # If 3rd member hasn't reviewed yet after 24 hours
                        if total_reviews < eligible_acm_count:
                            application.status = 'acm_approved'
                            
                            # Notify applicant
                            create_notification(
                                user_id=application.applicant_id,
                                title='Application Forwarded to Dean',
                                message=f'Your application {application.tracking_id} has been approved by 2 ACM members and forwarded to the Dean for review.',
                                notification_type='success'
                            )
                            
                            # Notify Dean
                            dean_user = User.query.filter_by(role='dean').first()
                            if dean_user:
                                create_notification(
                                    user_id=dean_user.id,
                                    title='New Application for Review',
                                    message=f'Application {application.tracking_id} has been approved by 2 ACM members (auto-forwarded after 24 hours) and requires your review.',
                                    notification_type='info'
                                )
                            
                            forwarded_count += 1
            
            if forwarded_count > 0:
                db.session.commit()
                print(f"âœ… Auto-forwarded {forwarded_count} application(s) to Dean")
                
        except Exception as e:
            print(f"âŒ Error in auto-forward task: {e}")
            db.session.rollback()

# Schedule the auto-forward task to run every hour
scheduler.add_job(
    func=auto_forward_acm_applications,
    trigger=IntervalTrigger(hours=1),
    id='acm_auto_forward_job',
    name='Auto-forward ACM applications to Dean after 24 hours',
    replace_existing=True
)

# Citation Auto-Update Job - Runs on 1st of every month at 2 AM
def quarterly_citation_update():
    """Background job to auto-update citation counts quarterly"""
    try:
        with app.app_context():
            from citation_updater import update_publication_citations
            print("\n" + "="*60)
            print("STARTING QUARTERLY CITATION AUTO-UPDATE")
            print("="*60)
            
            stats = update_publication_citations(app, update_all=False)
            
            print("\n" + "="*60)
            print("CITATION UPDATE COMPLETE")
            print(f"Checked: {stats['total_checked']} | "
                  f"Updated: {stats['successfully_updated']} | "
                  f"No change: {stats['no_change']} | "
                  f"Failed: {stats['failed']}")
            print("="*60 + "\n")
            
    except Exception as e:
        print(f"âŒ Error in citation update task: {e}")

# Schedule citation update: 1st of every month at 2:00 AM
scheduler.add_job(
    func=quarterly_citation_update,
    trigger='cron',
    day=1,  # 1st day of month
    hour=2,  # 2 AM
    minute=0,
    id='citation_update_job',
    name='Quarterly citation auto-update from Google Scholar',
    replace_existing=True
)

print("âœ… Citation Auto-Update Scheduler configured - runs 1st of every month at 2 AM")

# Start the scheduler
scheduler.start()

# Shut down the scheduler when exiting the app
atexit.register(lambda: scheduler.shutdown())

# Context processor for global template variables
@app.context_processor
def inject_globals():
    """Inject global variables into all templates"""
    is_acm_member = False
    acm_pending_count = 0
    dean_pending_review = 0
    hod_pending_review = 0
    unread_circulars = 0
    
    if current_user.is_authenticated:
        from models import Circular, CircularRead, ACMPanel, ACMReview
        
        # Count unread circulars (active circulars not yet read by this user)
        read_circular_ids = [cr.circular_id for cr in CircularRead.query.filter_by(
            user_id=current_user.id
        ).all()]
        
        if read_circular_ids:
            unread_circulars = Circular.query.filter(
                Circular.is_active == True,
                ~Circular.id.in_(read_circular_ids)
            ).count()
        else:
            unread_circulars = Circular.query.filter_by(is_active=True).count()
        
        if current_user.role == 'faculty':
            # Check if user is an active ACM member
            acm_membership = ACMPanel.query.filter_by(
                member_id=current_user.id,
                is_active=True
            ).first()
            
            if acm_membership:
                is_acm_member = True
                
                # Count pending applications (not yet reviewed by this member)
                reviewed_app_ids = [r.application_id for r in ACMReview.query.filter_by(
                    reviewer_id=current_user.id
                ).all()]
                
                acm_pending_count = ApplicationForm.query.filter(
                    ApplicationForm.status.in_(['hod_approved', 'acm_review']),
                    ~ApplicationForm.id.in_(reviewed_app_ids) if reviewed_app_ids else True,
                    ApplicationForm.applicant_id != current_user.id  # Exclude own applications
                ).count()
        
        elif current_user.role == 'dean':
            # Count applications pending Dean review (ACM approved)
            dean_pending_review = ApplicationForm.query.filter_by(
                status='acm_approved'
            ).count()
        
        elif current_user.role == 'hod':
            # Get faculty in HOD's department
            dept_faculty_ids = [u.id for u in User.query.filter_by(
                department_id=current_user.department_id,
                role='faculty'
            ).all()]
            
            # Count applications from HOD's department pending review
            hod_pending_review = ApplicationForm.query.filter(
                ApplicationForm.status == 'submitted',
                ApplicationForm.applicant_id.in_(dept_faculty_ids)
            ).count()
    
    return dict(
        is_acm_member=is_acm_member,
        acm_pending_count=acm_pending_count,
        dean_pending_review=dean_pending_review,
        hod_pending_review=hod_pending_review,
        unread_circulars=unread_circulars
    )


# Helper functions for notifications and audit logging
def create_notification(user_id, title, message, notification_type='info', publication_id=None, send_email=True):
    """
    Create a notification for a user and optionally send email
    
    Args:
        user_id: User ID to notify
        title: Notification title
        message: Notification message
        notification_type: Type of notification (info, success, warning, danger)
        publication_id: Optional publication ID
        send_email: Whether to send email notification (default: True)
    """
    from models import User
    
    notification = Notification(
        user_id=user_id,
        title=title,
        message=message,
        type=notification_type,
        publication_id=publication_id
    )
    db.session.add(notification)
    db.session.commit()
    
    # Send email notification if enabled
    if send_email:
        user = User.query.get(user_id)
        if user and user.email:
            from email_utils import send_notification_email
            # Send email asynchronously, skip validation for existing users
            try:
                send_notification_email(user, title, message, skip_validation=True)
            except Exception as e:
                # Log error but don't fail the notification
                print(f"Failed to send email notification to {user.email}: {str(e)}")
    
    return notification


def notify_hod_of_application(application):
    """Notify HOD of new application from their department faculty"""
    from models import User
    
    # Get the applicant's department HODs
    hods = User.query.filter_by(
        department_id=application.applicant.department_id,
        role='hod'
    ).all()
    
    for hod in hods:
        create_notification(
            user_id=hod.id,
            title=f'New Application: {application.tracking_id}',
            message=f'{application.applicant.name} submitted a {application.application_type.replace("_", " ").title()} application. Please review.',
            notification_type='info'
        )


def create_audit_log(action, target_type, target_id=None, details=None, user_id=None):
    """Create an audit log entry"""
    if user_id is None and current_user.is_authenticated:
        user_id = current_user.id
    
    audit = AuditLog(
        user_id=user_id,
        action=action,
        target_type=target_type,
        target_id=target_id,
        details=json.dumps(details) if details else None,
        ip_address=request.remote_addr if request else None
    )
    db.session.add(audit)
    db.session.commit()
    return audit


def normalize_string(s):
    """Normalize string for duplicate detection"""
    if not s:
        return ""
    # Convert to lowercase, remove extra spaces and special characters
    s = s.lower().strip()
    s = re.sub(r'[^\w\s]', '', s)  # Remove punctuation
    s = re.sub(r'\s+', ' ', s)  # Normalize spaces
    return s


def allowed_file(filename, file_type='pdf'):
    """Check if file has allowed extension"""
    if file_type == 'pdf':
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']
    elif file_type == 'image':
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_IMAGE_EXTENSIONS']
    elif file_type == 'document':
        # For application documents (PDF, images, Word)
        allowed_extensions = {'pdf', 'png', 'jpg', 'jpeg', 'docx'}
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions
    return False


def save_application_documents(files, application_id, user_id):
    """Save multiple documents for an application"""
    from models import ApplicationDocument
    from werkzeug.utils import secure_filename
    import secrets
    
    saved_documents = []
    
    # Create application documents folder if it doesn't exist
    app_docs_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'application_documents')
    os.makedirs(app_docs_folder, exist_ok=True)
    
    for file in files:
        if file and file.filename and allowed_file(file.filename, 'document'):
            # Generate secure filename
            original_filename = secure_filename(file.filename)
            file_ext = original_filename.rsplit('.', 1)[1].lower()
            stored_filename = f"{secrets.token_hex(16)}_{application_id}.{file_ext}"
            
            # Save file
            file_path = os.path.join(app_docs_folder, stored_filename)
            file.save(file_path)
            
            # Get file size
            file_size = os.path.getsize(file_path)
            
            # Create database record
            document = ApplicationDocument(
                application_id=application_id,
                filename=original_filename,
                stored_filename=stored_filename,
                file_type=file_ext,
                file_size=file_size,
                uploaded_by=user_id
            )
            db.session.add(document)
            saved_documents.append(document)
    
    return saved_documents


def notify_third_member_deadline(application, eligible_acm_count):
    """Notify remaining ACM member about 48-hour deadline after 2nd approval"""
    from models import ACMPanel, ACMReview, User
    
    # Find ACM members who haven't reviewed yet
    reviewed_member_ids = [r.reviewer_id for r in application.acm_reviews]
    
    remaining_members = ACMPanel.query.filter(
        ACMPanel.is_active == True,
        ACMPanel.member_id != application.applicant_id,
        ~ACMPanel.member_id.in_(reviewed_member_ids)
    ).all()
    
    for member in remaining_members:
        create_notification(
            user_id=member.member_id,
            title=f'â° Urgent: Review Application {application.tracking_id}',
            message=f'Application {application.tracking_id} has received 2 approvals. You have 48 hours to submit your review before it automatically forwards to Dean. Your input is valuable!',
            notification_type='warning'
        )


def check_and_auto_forward_applications():
    """Background job to check and auto-forward applications after 48-hour timer expires"""
    from models import ApplicationForm, User, ACMReview
    
    with app.app_context():
        try:
            # Find applications waiting for auto-forward
            now = now_ist()
            
            # Applications that:
            # 1. Have 2+ approvals
            # 2. Timer started (acm_second_approval_at is set)
            # 3. Not already forwarded
            # 4. 48 hours have passed
            applications = ApplicationForm.query.filter(
                ApplicationForm.acm_second_approval_at.isnot(None),
                ApplicationForm.acm_auto_forward_scheduled == True,
                ApplicationForm.status == 'acm_review',
                ApplicationForm.acm_approvals_count >= 2
            ).all()
            
            for application in applications:
                # Calculate time elapsed since 2nd approval
                time_elapsed = now - application.acm_second_approval_at
                hours_elapsed = time_elapsed.total_seconds() / 3600
                
                # Send reminders at 24h and 47h marks
                if 23.5 <= hours_elapsed < 24.5:
                    # 24-hour reminder
                    send_timer_reminder(application, hours_remaining=24)
                elif 46.5 <= hours_elapsed < 47.5:
                    # 1-hour warning
                    send_timer_reminder(application, hours_remaining=1)
                elif hours_elapsed >= 48:
                    # 48 hours passed - auto-forward to Dean
                    auto_forward_to_dean(application)
                    
        except Exception as e:
            print(f"ERROR in check_and_auto_forward_applications: {str(e)}")
            import traceback
            traceback.print_exc()


def send_timer_reminder(application, hours_remaining):
    """Send reminder notification to remaining ACM members"""
    from models import ACMPanel, ACMReview, User
    
    # Find ACM members who haven't reviewed yet
    reviewed_member_ids = [r.reviewer_id for r in application.acm_reviews]
    
    remaining_members = ACMPanel.query.filter(
        ACMPanel.is_active == True,
        ACMPanel.member_id != application.applicant_id,
        ~ACMPanel.member_id.in_(reviewed_member_ids)
    ).all()
    
    if hours_remaining == 24:
        message = f'Reminder: You have 24 hours remaining to review application {application.tracking_id} before it auto-forwards to Dean.'
        title = f'â° 24 Hours Remaining: {application.tracking_id}'
    else:
        message = f'URGENT: Only {hours_remaining} hour(s) left to review application {application.tracking_id}!'
        title = f'ðŸš¨ {hours_remaining}h Left: {application.tracking_id}'
    
    for member in remaining_members:
        create_notification(
            user_id=member.member_id,
            title=title,
            message=message,
            notification_type='warning'
        )


def auto_forward_to_dean(application):
    """Auto-forward application to Dean after 48-hour timer expires"""
    from models import User, ACMReview
    
    try:
        # Mark as auto-forwarded
        application.acm_auto_forwarded = True
        application.acm_auto_forward_scheduled = False
        application.status = 'acm_approved'
        application.acm_review_date = now_ist()
        application.acm_approved_at = now_ist()  # Record when forwarded to Dean
        
        # Apply any field edits from approved ACM reviews
        approved_reviews = ACMReview.query.filter_by(
            application_id=application.id,
            decision='approved'
        ).all()
        
        for review in approved_reviews:
            if review.edited_fields:
                try:
                    edits = json.loads(review.edited_fields)
                    for field, change_data in edits.items():
                        if hasattr(application, field):
                            if isinstance(change_data, dict) and 'new' in change_data:
                                new_value = change_data['new']
                            else:
                                new_value = change_data
                            setattr(application, field, new_value)
                except:
                    pass
        
        db.session.commit()
        
        # Notify applicant
        create_notification(
            user_id=application.applicant_id,
            title='Application Auto-Forwarded to Dean',
            message=f'Your application {application.tracking_id} received 2 ACM approvals and has been automatically forwarded to Dean after 48-hour review period.',
            notification_type='success'
        )
        
        # Notify Dean
        dean_user = User.query.filter_by(role='dean').first()
        if dean_user:
            create_notification(
                user_id=dean_user.id,
                title='Application Auto-Forwarded from ACM',
                message=f'Application {application.tracking_id} (2/{application.acm_approvals_count} ACM approvals) has been auto-forwarded after 48-hour review period.',
                notification_type='info'
            )
        
        # Notify remaining ACM members (informational)
        reviewed_member_ids = [r.reviewer_id for r in application.acm_reviews]
        from models import ACMPanel
        
        remaining_members = ACMPanel.query.filter(
            ACMPanel.is_active == True,
            ACMPanel.member_id != application.applicant_id,
            ~ACMPanel.member_id.in_(reviewed_member_ids)
        ).all()
        
        for member in remaining_members:
            create_notification(
                user_id=member.member_id,
                title='Application Auto-Forwarded',
                message=f'Application {application.tracking_id} has been auto-forwarded to Dean after 48-hour deadline. Review period closed.',
                notification_type='info'
            )
        
        print(f"Auto-forwarded application {application.tracking_id} to Dean")
        
    except Exception as e:
        db.session.rollback()
        print(f"ERROR auto-forwarding application {application.id}: {str(e)}")
        import traceback
        traceback.print_exc()



def get_user_folder(user_id, folder_type='publications'):
    """Get user-specific folder path"""
    if folder_type == 'publications':
        folder_path = os.path.join(app.config['PUBLICATIONS_FOLDER'], f'user_{user_id}')
    elif folder_type == 'profiles':
        folder_path = app.config['PROFILES_FOLDER']
    elif folder_type == 'signatures':
        folder_path = app.config['SIGNATURES_FOLDER']
    elif folder_type == 'feedback_screenshots':
        folder_path = app.config['FEEDBACK_SCREENSHOTS_FOLDER']
    else:
        folder_path = app.config['UPLOAD_FOLDER']
    
    os.makedirs(folder_path, exist_ok=True)
    return folder_path


def find_duplicate_publications(title, year, doi=None, journal_name=None, threshold=85):
    """
    Find potential duplicate publications based on DOI or fuzzy title matching
    
    Args:
        title: Publication title
        year: Publication year
        doi: DOI (optional - if present, exact match only)
        journal_name: Journal/Conference name (optional)
        threshold: Fuzzy match threshold (0-100, default 85)
    
    Returns:
        List of (publication, similarity_score) tuples
    """
    duplicates = []
    
    # Priority 1: DOI matching (exact match)
    if doi and doi.strip():
        doi_matches = Publication.query.filter(
            func.lower(Publication.doi) == doi.lower().strip()
        ).all()
        if doi_matches:
            return [(pub, 100) for pub in doi_matches]
    
    # Priority 2: Fuzzy title matching
    normalized_title = normalize_string(title)
    
    # Get publications from same year and nearby years (Â±1 year)
    candidates = Publication.query.filter(
        Publication.year.between(year - 1, year + 1)
    ).all()
    
    for pub in candidates:
        pub_normalized_title = normalize_string(pub.title)
        
        # Calculate similarity score
        similarity = fuzz.ratio(normalized_title, pub_normalized_title)
        
        # Boost score if journal/conference names match
        if journal_name and pub.journal_conference_name:
            journal_similarity = fuzz.ratio(
                normalize_string(journal_name),
                normalize_string(pub.journal_conference_name)
            )
            if journal_similarity > 80:
                similarity = min(100, similarity + 10)
        
        if similarity >= threshold:
            duplicates.append((pub, similarity))
    
    # Sort by similarity score (highest first)
    duplicates.sort(key=lambda x: x[1], reverse=True)
    
    return duplicates


# Role-based access decorators
def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                flash('Please log in to access this page.', 'warning')
                return redirect(url_for('login'))
            if current_user.role not in roles:
                flash('You do not have permission to access this page.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator


# Routes
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


# ============================================================================
# EMAIL VALIDATION API
# ============================================================================

@app.route('/api/validate-email', methods=['POST'])
def validate_email_api():
    """API endpoint to validate if an email exists"""
    from email_utils import validate_email_exists
    
    data = request.get_json()
    email = data.get('email', '').strip()
    
    if not email:
        return jsonify({'valid': False, 'message': 'Email is required'})
    
    # Check if email domain is allowed (sjec.ac.in)
    if not email.endswith('@' + app.config['ALLOWED_EMAIL_DOMAIN']):
        return jsonify({
            'valid': False,
            'message': f'Only @{app.config["ALLOWED_EMAIL_DOMAIN"]} emails are allowed'
        })
    
    # Validate if email exists
    result = validate_email_exists(email)
    return jsonify(result)


# ============================================================================
# AUTHENTICATION ROUTES
# ============================================================================

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = RegistrationForm()
    form.department.choices = [(d.id, d.name) for d in Department.query.order_by(Department.name).all()]
    
    if form.validate_on_submit():
        # Validate email (must be @sjec.ac.in, no dummy emails allowed for self-registration)
        from email_utils import validate_email_exists
        email_validation = validate_email_exists(form.email.data, allow_sjec_dummy=False)
        
        if not email_validation['valid']:
            flash(f'Invalid email: {email_validation["message"]}', 'danger')
            return render_template('register.html', form=form)
        
        # Check if email already exists
        if User.query.filter_by(email=form.email.data).first():
            flash('Email already registered. Please login.', 'danger')
            return render_template('register.html', form=form)
        
        user = User(
            name=form.name.data,
            email=form.email.data,
            department_id=form.department.data,
            role='faculty'  # Default role for self-registration
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        
        # Send welcome email (don't skip validation for faculty registration)
        from email_utils import send_welcome_email
        from flask import current_app
        email_result = send_welcome_email(user, skip_validation=False)
        
        if email_result['success']:
            flash('Registration successful! A welcome email has been sent to your email address. Please log in.', 'success')
        else:
            flash(f'Registration successful! Please log in. (Email notification: {email_result["message"]})', 'warning')
        
        return redirect(url_for('login'))
    
    return render_template('register.html', form=form)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    
    # Get latest 3 circulars for display on login page
    from models import Circular
    recent_circulars = Circular.query.filter_by(is_active=True)\
        .order_by(Circular.created_at.desc()).limit(3).all()
    
    if form.validate_on_submit():
        user = User.query.filter_by(email=form.email.data).first()
        if user and user.check_password(form.password.data):
            login_user(user)
            next_page = request.args.get('next')
            flash(f'Welcome back, {user.name}!', 'success')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('Invalid email or password.', 'danger')
    
    return render_template('login.html', form=form, recent_circulars=recent_circulars)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    # Redirect admin to admin dashboard
    if current_user.role == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Dean Secretary uses Final Approval dashboard
    if current_user.role == 'dean_secretary':
        return redirect(url_for('final_approval_pending'))
    
    # Director uses same dashboard as Principal
    if current_user.role == 'director':
        current_user.role = 'principal'
    
    if current_user.role == 'faculty':
        # Pagination
        page = request.args.get('page', 1, type=int)
        per_page = 10
        
        pagination = Publication.query.filter_by(user_id=current_user.id)\
            .order_by(Publication.created_at.desc())\
            .paginate(page=page, per_page=per_page, error_out=False)
        
        publications = pagination.items
        
        # Get edit requests for each publication to check if faculty has already requested
        edit_requests = {}
        for pub in publications:
            # Check if there's any edit request (pending, approved, or denied)
            # For confirmed publications that can't be edited anymore
            if pub.status == 'confirmed' and not pub.can_edit:
                any_request = EditRequest.query.filter_by(
                    publication_id=pub.id,
                    faculty_id=current_user.id
                ).order_by(EditRequest.requested_at.desc()).first()
                
                if any_request:
                    # Map statuses for UI
                    if any_request.status in ['approved', 'admin_approved']:
                        # If approved but can_edit is False, they already used their edit
                        edit_requests[pub.id] = 'used'
                    elif any_request.status == 'pending':
                        edit_requests[pub.id] = 'pending'
                    elif any_request.status == 'pending_admin':
                        edit_requests[pub.id] = 'pending_admin'
                    elif any_request.status in ['denied', 'admin_denied']:
                        edit_requests[pub.id] = 'denied'
        
        # Get total count and stats for all publications (not paginated)
        all_pubs = Publication.query.filter_by(user_id=current_user.id).all()
        total_publications = len(all_pubs)
        scopus_count = len([p for p in all_pubs if p.indexing_status == 'Scopus'])
        total_citations = sum([p.citation_count or 0 for p in all_pubs])
        
        # Calculate average citations per publication
        avg_citations = (total_citations / total_publications) if total_publications > 0 else 0
        scopus_percentage = round((scopus_count / total_publications * 100), 1) if total_publications > 0 else 0
        
        # Get application statistics
        conference_apps = ApplicationForm.query.filter_by(
            applicant_id=current_user.id, 
            application_type='Conference Registration'
        ).count()
        
        # Count approved applications (principal_approved or director_approved)
        conference_approved = ApplicationForm.query.filter(
            ApplicationForm.applicant_id == current_user.id,
            ApplicationForm.application_type == 'Conference Registration',
            (ApplicationForm.principal_approved == True) | (ApplicationForm.director_approved == True)
        ).count()
        
        incentive_apps = ApplicationForm.query.filter_by(
            applicant_id=current_user.id,
            application_type='Publication Incentive'
        ).count()
        
        incentive_approved = ApplicationForm.query.filter(
            ApplicationForm.applicant_id == current_user.id,
            ApplicationForm.application_type == 'Publication Incentive',
            (ApplicationForm.principal_approved == True) | (ApplicationForm.director_approved == True)
        ).count()
        
        journal_apps = ApplicationForm.query.filter_by(
            applicant_id=current_user.id,
            application_type='Journal Support'
        ).count()
        
        journal_approved = ApplicationForm.query.filter(
            ApplicationForm.applicant_id == current_user.id,
            ApplicationForm.application_type == 'Journal Support',
            (ApplicationForm.principal_approved == True) | (ApplicationForm.director_approved == True)
        ).count()
        
        # Create performance statistics dictionary
        performance_stats = {
            'total_publications': total_publications,
            'scopus_count': scopus_count,
            'scopus_percentage': scopus_percentage,
            'total_citations': total_citations,
            'avg_citations': avg_citations,
            'conference_apps': conference_apps,
            'conference_approved': conference_approved,
            'incentive_apps': incentive_apps,
            'incentive_approved': incentive_approved,
            'journal_apps': journal_apps,
            'journal_approved': journal_approved
        }
        
        # Check if user is Head of Research for any department
        is_head_of_research = Department.query.filter_by(head_of_research_id=current_user.id).first()
        
        return render_template('faculty_dashboard.html', 
                             publications=publications,
                             total_publications=total_publications,
                             scopus_count=scopus_count,
                             total_citations=total_citations,
                             edit_requests=edit_requests,
                             pagination=pagination,
                             is_head_of_research=is_head_of_research,
                             performance_stats=performance_stats)
    
    elif current_user.role == 'hod':
        # Pagination
        page = request.args.get('page', 1, type=int)
        per_page = 10
        
        # Filter publications
        dept_id = request.args.get('department', current_user.department_id, type=int)
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        query = Publication.query.filter_by(department_id=dept_id)
        if year:
            query = query.filter_by(year=year)
        if month:
            query = query.filter_by(month=month)
        
        all_publications = query.order_by(Publication.year.desc(), Publication.month.desc()).all()
        
        # Remove duplicates by title (case-insensitive) - same title = 1 publication
        unique_pubs_dict = {}
        for pub in all_publications:
            title_key = pub.title.strip().lower()
            if title_key not in unique_pubs_dict:
                unique_pubs_dict[title_key] = pub
        
        publications_list = list(unique_pubs_dict.values())
        unique_pub_ids = [p.id for p in publications_list]
        
        # Manual pagination
        total = len(publications_list)
        start = (page - 1) * per_page
        end = start + per_page
        publications = publications_list[start:end]
        
        # Create pagination object
        class SimplePagination:
            def __init__(self, items, page, per_page, total):
                self.items = items
                self.page = page
                self.per_page = per_page
                self.total = total
                self.pages = (total + per_page - 1) // per_page if per_page > 0 else 0
                self.has_prev = page > 1
                self.has_next = page < self.pages
                self.prev_num = page - 1 if self.has_prev else None
                self.next_num = page + 1 if self.has_next else None
        
        pagination = SimplePagination(publications, page, per_page, total)
        
        departments = Department.query.order_by(Department.name).all()
        
        # Get unique years
        years = db.session.query(Publication.year).distinct().order_by(Publication.year.desc()).all()
        years = [y[0] for y in years]
        
        # Get top 5 faculty by publications (annually) - only count unique publications
        current_year = datetime.now().year
        selected_year = year if year else current_year
        top_faculty = db.session.query(
            User.name,
            User.email,
            func.count(Publication.id).label('pub_count')
        ).select_from(User).join(
            Publication, Publication.user_id == User.id
        ).filter(
            User.department_id == current_user.department_id,
            Publication.year == selected_year,
            Publication.id.in_(unique_pub_ids)
        ).group_by(User.id).order_by(func.count(Publication.id).desc()).limit(5).all()
        
        # Get pending edit requests count
        pending_edit_requests = EditRequest.query.join(Publication).filter(
            Publication.department_id == current_user.department_id,
            EditRequest.status == 'pending'
        ).count()
        
        return render_template('hod_dashboard.html', 
                             publications=publications,
                             all_publications=publications_list,
                             departments=departments,
                             years=years,
                             selected_dept=dept_id,
                             selected_year=year,
                             selected_month=month,
                             top_faculty=top_faculty,
                             pending_edit_requests=pending_edit_requests,
                             pagination=pagination)
    
    elif current_user.role == 'principal':
        # Pagination
        page = request.args.get('page', 1, type=int)
        per_page = 10
        
        # Filter publications
        dept_id = request.args.get('department', type=int)
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        query = Publication.query
        if dept_id:
            query = query.filter_by(department_id=dept_id)
        if year:
            query = query.filter_by(year=year)
        if month:
            query = query.filter_by(month=month)
        
        all_publications = query.order_by(Publication.year.desc(), Publication.month.desc()).all()
        
        # Remove duplicates by title (case-insensitive) - same title = 1 publication
        unique_pubs_dict = {}
        for pub in all_publications:
            title_key = pub.title.strip().lower()
            if title_key not in unique_pubs_dict:
                unique_pubs_dict[title_key] = pub
        
        publications_list = list(unique_pubs_dict.values())
        unique_pub_ids = [p.id for p in publications_list]
        
        # Manual pagination
        total = len(publications_list)
        start = (page - 1) * per_page
        end = start + per_page
        publications = publications_list[start:end]
        
        # Create pagination object
        class SimplePagination:
            def __init__(self, items, page, per_page, total):
                self.items = items
                self.page = page
                self.per_page = per_page
                self.total = total
                self.pages = (total + per_page - 1) // per_page if per_page > 0 else 0
                self.has_prev = page > 1
                self.has_next = page < self.pages
                self.prev_num = page - 1 if self.has_prev else None
                self.next_num = page + 1 if self.has_next else None
        
        pagination = SimplePagination(publications, page, per_page, total)
        
        departments = Department.query.order_by(Department.name).all()
        
        # Get unique years
        years = db.session.query(Publication.year).distinct().order_by(Publication.year.desc()).all()
        years = [y[0] for y in years]
        
        # Get top 5 faculty by publications (annually) - only count unique publications
        current_year = datetime.now().year
        top_faculty = db.session.query(
            User.name,
            User.email,
            Department.name.label('dept_name'),
            func.count(Publication.id).label('pub_count')
        ).select_from(User).join(
            Publication, Publication.user_id == User.id
        ).join(
            Department, User.department_id == Department.id
        ).filter(
            Publication.year == (year if year else current_year),
            Publication.id.in_(unique_pub_ids)
        ).group_by(User.id, Department.name).order_by(func.count(Publication.id).desc()).limit(5).all()
        
        return render_template('principal_dashboard.html', 
                             publications=publications,
                             all_publications=publications_list,
                             departments=departments,
                             years=years,
                             selected_dept=dept_id,
                             selected_year=year,
                             selected_month=month,
                             top_faculty=top_faculty,
                             pagination=pagination)
    
    elif current_user.role == 'vice_principal':
        # Pagination
        page = request.args.get('page', 1, type=int)
        per_page = 10
        
        # Vice Principal: View-only access (same view as Principal)
        dept_id = request.args.get('department', type=int)
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        query = Publication.query
        if dept_id:
            query = query.filter_by(department_id=dept_id)
        if year:
            query = query.filter_by(year=year)
        if month:
            query = query.filter_by(month=month)
        
        all_publications = query.order_by(Publication.year.desc(), Publication.month.desc()).all()
        
        # Remove duplicates by title (case-insensitive)
        unique_pubs_dict = {}
        for pub in all_publications:
            title_key = pub.title.strip().lower()
            if title_key not in unique_pubs_dict:
                unique_pubs_dict[title_key] = pub
        
        publications_list = list(unique_pubs_dict.values())
        unique_pub_ids = [p.id for p in publications_list]
        
        # Manual pagination
        total = len(publications_list)
        start = (page - 1) * per_page
        end = start + per_page
        publications = publications_list[start:end]
        
        # Create pagination object
        class SimplePagination:
            def __init__(self, items, page, per_page, total):
                self.items = items
                self.page = page
                self.per_page = per_page
                self.total = total
                self.pages = (total + per_page - 1) // per_page if per_page > 0 else 0
                self.has_prev = page > 1
                self.has_next = page < self.pages
                self.prev_num = page - 1 if self.has_prev else None
                self.next_num = page + 1 if self.has_next else None
        
        pagination = SimplePagination(publications, page, per_page, total)
        
        departments = Department.query.order_by(Department.name).all()
        
        # Get unique years
        years = db.session.query(Publication.year).distinct().order_by(Publication.year.desc()).all()
        years = [y[0] for y in years]
        
        # Get top 5 faculty by publications (annually)
        current_year = datetime.now().year
        top_faculty = db.session.query(
            User.name,
            User.email,
            Department.name.label('dept_name'),
            func.count(Publication.id).label('pub_count')
        ).select_from(User).join(
            Publication, Publication.user_id == User.id
        ).join(
            Department, User.department_id == Department.id
        ).filter(
            Publication.year == (year if year else current_year),
            Publication.id.in_(unique_pub_ids)
        ).group_by(User.id, Department.name).order_by(func.count(Publication.id).desc()).limit(5).all()
        
        return render_template('vice_principal_dashboard.html', 
                             publications=publications,
                             all_publications=publications_list,
                             departments=departments,
                             years=years,
                             selected_dept=dept_id,
                             selected_year=year,
                             selected_month=month,
                             top_faculty=top_faculty,
                             pagination=pagination)
    
    elif current_user.role == 'dean':
        # Pagination
        page = request.args.get('page', 1, type=int)
        per_page = 10
        
        # Dean R&D: Dashboard with quick actions and statistics
        from models import ACMPanel, Feedback
        
        # Quick Action Statistics
        total_users = User.query.count()
        pending_acm_review = ApplicationForm.query.filter(
            ApplicationForm.status.in_(['hod_approved', 'acm_review'])
        ).count()
        acm_approved_apps = ApplicationForm.query.filter_by(status='acm_approved').count()
        total_applications = ApplicationForm.query.count()
        active_acm_members = ACMPanel.query.filter_by(is_active=True).count()
        pending_feedbacks = Feedback.query.filter_by(status='pending').count()
        
        # Recent Applications (last 10)
        recent_applications = ApplicationForm.query.order_by(
            ApplicationForm.created_at.desc()
        ).limit(10).all()
        
        # Application Status Breakdown
        draft_apps = ApplicationForm.query.filter_by(status='draft').count()
        submitted_apps = ApplicationForm.query.filter_by(status='submitted').count()
        dean_approved_apps = ApplicationForm.query.filter_by(dean_approved=True).count()
        rejected_apps = ApplicationForm.query.filter_by(status='rejected').count()
        
        # Publications Overview
        dept_id = request.args.get('department', type=int)
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)
        
        query = Publication.query
        if dept_id:
            query = query.filter_by(department_id=dept_id)
        if year:
            query = query.filter_by(year=year)
        if month:
            query = query.filter_by(month=month)
        
        all_publications = query.order_by(Publication.year.desc(), Publication.month.desc()).all()
        
        # Remove duplicates by title (case-insensitive)
        unique_pubs_dict = {}
        for pub in all_publications:
            title_key = pub.title.strip().lower()
            if title_key not in unique_pubs_dict:
                unique_pubs_dict[title_key] = pub
        
        publications_list = list(unique_pubs_dict.values())
        unique_pub_ids = [p.id for p in publications_list]
        
        # Manual pagination
        total = len(publications_list)
        start = (page - 1) * per_page
        end = start + per_page
        publications = publications_list[start:end]
        
        # Create pagination object
        class SimplePagination:
            def __init__(self, items, page, per_page, total):
                self.items = items
                self.page = page
                self.per_page = per_page
                self.total = total
                self.pages = (total + per_page - 1) // per_page if per_page > 0 else 0
                self.has_prev = page > 1
                self.has_next = page < self.pages
                self.prev_num = page - 1 if self.has_prev else None
                self.next_num = page + 1 if self.has_next else None
        
        pagination = SimplePagination(publications, page, per_page, total)
        
        departments = Department.query.order_by(Department.name).all()
        
        # Get unique years
        years = db.session.query(Publication.year).distinct().order_by(Publication.year.desc()).all()
        years = [y[0] for y in years]
        
        # Get top 5 faculty by publications (annually)
        current_year = datetime.now().year
        top_faculty = db.session.query(
            User.name,
            User.email,
            Department.name.label('dept_name'),
            func.count(Publication.id).label('pub_count')
        ).select_from(User).join(
            Publication, Publication.user_id == User.id
        ).join(
            Department, User.department_id == Department.id
        ).filter(
            Publication.year == (year if year else current_year),
            Publication.id.in_(unique_pub_ids)
        ).group_by(User.id, Department.name).order_by(func.count(Publication.id).desc()).limit(5).all()
        
        return render_template('dean-dashboard.html', 
                             publications=publications,
                             all_publications=publications_list,
                             departments=departments,
                             years=years,
                             selected_dept=dept_id,
                             selected_year=year,
                             selected_month=month,
                             top_faculty=top_faculty,
                             total_users=total_users,
                             pending_acm_review=pending_acm_review,
                             acm_approved_apps=acm_approved_apps,
                             total_applications=total_applications,
                             active_acm_members=active_acm_members,
                             pending_feedbacks=pending_feedbacks,
                             recent_applications=recent_applications,
                             draft_apps=draft_apps,
                             submitted_apps=submitted_apps,
                             dean_approved_apps=dean_approved_apps,
                             rejected_apps=rejected_apps,
                             pagination=pagination)
    
    return redirect(url_for('index'))


@app.route('/publications/add', methods=['GET', 'POST'])
@role_required('faculty')
def add_publication():
    form = PublicationForm()
    
    if form.validate_on_submit():
        # Validate BibTeX format explicitly (in case validator didn't catch it)
        if form.bibtex_entry.data and form.bibtex_entry.data.strip():
            import re
            bibtex_text = form.bibtex_entry.data.strip()
            # Valid BibTeX entry types
            valid_types = ['article', 'book', 'inproceedings', 'conference', 'proceedings', 
                          'incollection', 'inbook', 'booklet', 'manual', 'techreport', 
                          'mastersthesis', 'phdthesis', 'misc', 'unpublished']
            bibtex_pattern = r'^@(' + '|'.join(valid_types) + r')\s*\{[^,]+,[\s\S]+\}\s*\}$'
            if not re.match(bibtex_pattern, bibtex_text, re.IGNORECASE | re.MULTILINE):
                flash('BibTeX entry is in incorrect format. Must start with @article, @book, @inproceedings, etc. and end with }}', 'danger')
                return render_template('add_publication.html', form=form)
        
        # Check for duplicates BEFORE saving
        duplicates = find_duplicate_publications(
            title=form.title.data,
            year=form.year.data,
            doi=form.doi.data,
            journal_name=form.journal_conference_name.data,
            threshold=85
        )
        
        # If duplicates found and user hasn't confirmed to proceed
        if duplicates and not request.form.get('ignore_duplicates'):
            # Store form data in session and show duplicates
            return render_template('add_publication.html', 
                                 form=form, 
                                 duplicates=duplicates,
                                 show_duplicate_warning=True)
        
        # Determine action: save or confirm
        action = request.form.get('action', 'save')  # 'save' or 'confirm'
        
        # Handle file upload
        pdf_filename = None
        if form.pdf_file.data:
            file = form.pdf_file.data
            if file and allowed_file(file.filename):
                filename = secure_filename(file.filename)
                # Add timestamp to avoid conflicts
                timestamp = now_ist().strftime('%Y%m%d_%H%M%S')
                pdf_filename = f"{timestamp}_{filename}"
                # Save to user-specific folder
                user_pub_folder = get_user_folder(current_user.id, 'publications')
                file.save(os.path.join(user_pub_folder, pdf_filename))
        
        publication = Publication(
            publication_type=form.publication_type.data or 'Journal',
            title=form.title.data,
            abstract=form.abstract.data,
            publisher_name=form.publisher_name.data,
            authors_names=form.authors_names.data,
            journal_conference_name=form.journal_conference_name.data,
            volume=form.volume.data,
            issue=form.issue.data,
            pages=form.pages.data,
            indexing_status=form.indexing_status.data if form.publication_type.data == 'Journal' else None,
            quartile=form.quartile.data if form.publication_type.data == 'Journal' else None,
            impact_factor=form.impact_factor.data if form.publication_type.data == 'Journal' else None,
            isbn=form.isbn.data if form.publication_type.data in ['Book', 'Book Chapter'] else None,
            edition=form.edition.data if form.publication_type.data in ['Book', 'Book Chapter'] else None,
            doi=form.doi.data,
            year=form.year.data,
            month=form.month.data if form.month.data != 0 else None,
            citation_count=form.citation_count.data or 0 if form.publication_type.data == 'Journal' else None,
            bibtex_entry=form.bibtex_entry.data,
            pdf_filename=pdf_filename,
            user_id=current_user.id,
            department_id=current_user.department_id,
            status='confirmed' if action == 'confirm' else 'saved',
            confirmed_at=now_ist() if action == 'confirm' else None,
            can_edit=True if action == 'saved' else False
        )
        
        db.session.add(publication)
        db.session.commit()
        
        # Create audit log
        create_audit_log(
            action='create_publication' if action == 'save' else 'confirm_publication',
            target_type='publication',
            target_id=publication.id,
            details={'title': publication.title, 'status': publication.status}
        )
        
        if action == 'confirm':
            flash('Publication confirmed successfully! It is now locked from editing.', 'success')
        else:
            flash('Publication saved as draft! You can edit it anytime until you confirm.', 'info')
        
        return redirect(url_for('dashboard'))
    
    return render_template('add_publication.html', form=form)



@app.route('/publications/edit/<int:id>', methods=['GET', 'POST'])
@role_required('faculty')
def edit_publication(id):
    publication = Publication.query.get_or_404(id)
    
    # Ensure user can only edit their own publications
    if publication.user_id != current_user.id:
        flash('You can only edit your own publications.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Check if publication can be edited
    if publication.status == 'confirmed' and not publication.can_edit:
        flash('This publication is confirmed and locked. Request edit permission from your HoD.', 'warning')
        return redirect(url_for('view_publication', id=id))
    
    form = PublicationForm(obj=publication)
    
    if form.validate_on_submit():
        # Validate BibTeX format explicitly (in case validator didn't catch it)
        if form.bibtex_entry.data and form.bibtex_entry.data.strip():
            import re
            bibtex_text = form.bibtex_entry.data.strip()
            # Valid BibTeX entry types
            valid_types = ['article', 'book', 'inproceedings', 'conference', 'proceedings', 
                          'incollection', 'inbook', 'booklet', 'manual', 'techreport', 
                          'mastersthesis', 'phdthesis', 'misc', 'unpublished']
            bibtex_pattern = r'^@(' + '|'.join(valid_types) + r')\s*\{[^,]+,[\s\S]+\}\s*\}$'
            if not re.match(bibtex_pattern, bibtex_text, re.IGNORECASE | re.MULTILINE):
                flash('BibTeX entry is in incorrect format. Must start with @article, @book, @inproceedings, etc. and end with }}', 'danger')
                return render_template('edit_publication.html', form=form, publication=publication)
        
        # Determine action: save or confirm
        action = request.form.get('action', 'save')
        
        # Handle file upload
        if form.pdf_file.data:
            file = form.pdf_file.data
            if file and allowed_file(file.filename):
                # Delete old file if exists
                if publication.pdf_filename:
                    user_pub_folder = get_user_folder(current_user.id, 'publications')
                    old_file_path = os.path.join(user_pub_folder, publication.pdf_filename)
                    if os.path.exists(old_file_path):
                        os.remove(old_file_path)
                
                filename = secure_filename(file.filename)
                timestamp = now_ist().strftime('%Y%m%d_%H%M%S')
                pdf_filename = f"{timestamp}_{filename}"
                # Save to user-specific folder
                user_pub_folder = get_user_folder(current_user.id, 'publications')
                file.save(os.path.join(user_pub_folder, pdf_filename))
                publication.pdf_filename = pdf_filename
        
        publication.title = form.title.data
        publication.abstract = form.abstract.data
        publication.publisher_name = form.publisher_name.data
        publication.authors_names = form.authors_names.data
        publication.journal_conference_name = form.journal_conference_name.data
        publication.volume = form.volume.data
        publication.issue = form.issue.data
        publication.pages = form.pages.data
        publication.indexing_status = form.indexing_status.data
        publication.quartile = form.quartile.data
        publication.impact_factor = form.impact_factor.data
        publication.doi = form.doi.data
        publication.year = form.year.data
        publication.month = form.month.data if form.month.data != 0 else None
        publication.citation_count = form.citation_count.data or 0
        publication.bibtex_entry = form.bibtex_entry.data
        
        # Update status if confirming
        if action == 'confirm':
            publication.status = 'confirmed'
            publication.confirmed_at = now_ist()
            publication.can_edit = False
            
            # Create audit log
            create_audit_log(
                action='confirm_publication',
                target_type='publication',
                target_id=publication.id,
                details={'title': publication.title}
            )
            
            flash('Publication confirmed and locked successfully!', 'success')
        else:
            # If this was a one-time edit permission, lock it again
            if publication.status == 'confirmed' and publication.can_edit:
                publication.can_edit = False
                publication.edit_count = (publication.edit_count or 0) + 1
                flash('Publication updated and locked again. You\'ll need to request edit permission for future changes.', 'success')
            else:
                flash('Publication updated successfully!', 'success')
            
            # Create audit log for edit
            create_audit_log(
                action='edit_publication',
                target_type='publication',
                target_id=publication.id,
                details={'title': publication.title, 'was_one_time_edit': publication.status == 'confirmed', 'edit_count': publication.edit_count}
            )
        
        db.session.commit()
        return redirect(url_for('dashboard'))
    
    return render_template('edit_publication.html', form=form, publication=publication)


@app.route('/publications/delete/<int:id>', methods=['POST'])
@login_required
def delete_publication(id):
    # Only HoD, Principal, and Admin can delete publications
    if current_user.role not in ['hod', 'principal', 'admin']:
        flash('Access denied. Only HoD, Principal, or Admin can delete publications.', 'danger')
        return redirect(url_for('dashboard'))
    
    publication = Publication.query.get_or_404(id)
    
    # HoD can only delete from their department
    if current_user.role == 'hod' and publication.department_id != current_user.department_id:
        flash('You can only delete publications from your department.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Store details for audit log
    pub_title = publication.title
    pub_owner_id = publication.user_id
    
    # Delete PDF file if exists
    if publication.pdf_filename:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], publication.pdf_filename)
        if os.path.exists(file_path):
            os.remove(file_path)
    
    # Notify the publication owner
    if pub_owner_id != current_user.id:
        create_notification(
            user_id=pub_owner_id,
            title='Publication Deleted',
            message=f'Your publication "{pub_title[:50]}..." has been deleted by {current_user.role.upper()}.',
            notification_type='warning'
        )
    
    # Create audit log
    create_audit_log(
        action='delete_publication',
        target_type='publication',
        target_id=id,
        details={'title': pub_title, 'owner_id': pub_owner_id, 'deleted_by_role': current_user.role}
    )
    
    db.session.delete(publication)
    db.session.commit()
    
    flash('Publication deleted successfully!', 'success')
    return redirect(url_for('dashboard'))


@app.route('/publications/view/<int:id>')
@login_required
def view_publication(id):
    publication = Publication.query.get_or_404(id)
    
    # Check access rights
    if current_user.role == 'faculty' and publication.user_id != current_user.id:
        flash('You can only view your own publications.', 'danger')
        return redirect(url_for('dashboard'))
    elif current_user.role == 'hod' and publication.department_id != current_user.department_id:
        flash('You can only view publications from your department.', 'danger')
        return redirect(url_for('dashboard'))
    
    return render_template('view_publication.html', publication=publication)


@app.route('/publications/download/<int:id>')
@login_required
def download_publication(id):
    publication = Publication.query.get_or_404(id)
    
    # Check access rights
    if current_user.role == 'faculty' and publication.user_id != current_user.id:
        flash('You can only download your own publications.', 'danger')
        return redirect(url_for('dashboard'))
    elif current_user.role == 'hod' and publication.department_id != current_user.department_id:
        flash('You can only download publications from your department.', 'danger')
        return redirect(url_for('dashboard'))
    
    if not publication.pdf_filename:
        flash('No PDF file available for this publication.', 'warning')
        return redirect(url_for('view_publication', id=id))
    
    # Get user-specific publication folder
    user_pub_folder = get_user_folder(publication.user_id, 'publications')
    file_path = os.path.join(user_pub_folder, publication.pdf_filename)
    if not os.path.exists(file_path):
        flash('PDF file not found.', 'danger')
        return redirect(url_for('view_publication', id=id))
    
    return send_file(file_path, as_attachment=True)


@app.route('/publications/update-citations/<int:id>', methods=['POST'])
@login_required
def update_publication_citations_manual(id):
    """Manually update citations for a specific publication"""
    publication = Publication.query.get_or_404(id)
    
    # Check access rights
    if current_user.role == 'faculty' and publication.user_id != current_user.id:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403
    
    if publication.publication_type != 'Journal':
        return jsonify({'success': False, 'message': 'Only journal publications have citations'}), 400
    
    try:
        from citation_updater import manual_citation_update
        success, message, new_count = manual_citation_update(app, id)
        
        if success:
            return jsonify({
                'success': True,
                'message': message,
                'new_count': new_count,
                'old_count': publication.citation_count
            })
        else:
            return jsonify({'success': False, 'message': message}), 400
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@app.route('/download-publications-xlsx')
@login_required
@role_required('principal', 'hod', 'admin', 'vice_principal', 'dean', 'director')
def download_publications_xlsx():
    """Download publications as Excel file with current filters"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    
    # Get filter parameters
    year_filter = request.args.get('year', '')
    department_filter = request.args.get('department', '')
    indexing_filter = request.args.get('indexing', '')
    quartile_filter = request.args.get('quartile', '')
    pub_type_filter = request.args.get('pub_type', '')
    
    # Build query based on role and filters
    query = Publication.query
    
    # Role-based filtering
    if current_user.role == 'hod':
        query = query.filter_by(department_id=current_user.department_id)
    
    # Apply filters
    if year_filter:
        query = query.filter_by(year=int(year_filter))
    if department_filter:
        query = query.filter_by(department_id=int(department_filter))
    if indexing_filter:
        query = query.filter_by(indexing_status=indexing_filter)
    if quartile_filter:
        query = query.filter_by(quartile=quartile_filter)
    if pub_type_filter:
        query = query.filter_by(publication_type=pub_type_filter)
    
    publications = query.order_by(Publication.year.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Publications"
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = ['S.No', 'Title', 'Authors', 'Type', 'Year', 'Journal/Book', 'Publisher',
               'Indexing', 'Quartile', 'Impact Factor', 'Citations', 'DOI', 'Department', 'Faculty']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    for row_num, pub in enumerate(publications, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=pub.title)
        ws.cell(row=row_num, column=3, value=pub.authors_names)
        ws.cell(row=row_num, column=4, value=pub.publication_type or 'Journal')
        ws.cell(row=row_num, column=5, value=pub.year)
        ws.cell(row=row_num, column=6, value=pub.journal_conference_name or '')
        ws.cell(row=row_num, column=7, value=pub.publisher_name or '')
        ws.cell(row=row_num, column=8, value=pub.indexing_status or '')
        ws.cell(row=row_num, column=9, value=pub.quartile or '')
        ws.cell(row=row_num, column=10, value=pub.impact_factor if pub.impact_factor else '')
        ws.cell(row=row_num, column=11, value=pub.citation_count if pub.citation_count else 0)
        ws.cell(row=row_num, column=12, value=pub.doi or '')
        ws.cell(row=row_num, column=13, value=pub.department.name if pub.department else '')
        ws.cell(row=row_num, column=14, value=pub.author.name if pub.author else '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename with filters
    filename_parts = ['publications']
    if year_filter:
        filename_parts.append(f'year{year_filter}')
    if department_filter:
        dept = Department.query.get(int(department_filter))
        if dept:
            filename_parts.append(dept.code)
    if pub_type_filter:
        filename_parts.append(pub_type_filter.replace(' ', ''))
    
    filename = '_'.join(filename_parts) + '.xlsx'
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/statistics')
@login_required
@role_required('principal', 'hod', 'admin', 'vice_principal', 'dean', 'director')
def statistics():
    # Get year filter for top faculty
    selected_year = request.args.get('year', type=int)
    
    if current_user.role == 'hod':
        dept_id = current_user.department_id
        publications = Publication.query.filter_by(department_id=dept_id).all()
        dept_name = current_user.department.name
    else:  # principal, vice_principal, dean, or admin
        publications = Publication.query.all()
        dept_name = "All Departments"
    
    # Remove duplicates by title (case-insensitive)
    unique_pubs = {}
    for pub in publications:
        title_key = pub.title.strip().lower()
        if title_key not in unique_pubs:
            unique_pubs[title_key] = pub
    
    unique_publications = list(unique_pubs.values())
    
    # Calculate statistics using unique publications
    total_pubs = len(unique_publications)
    
    # Get all unique publication IDs for filtering
    unique_pub_ids = [p.id for p in unique_publications]
    
    # Year-wise count (ordered by year ascending) - only unique publications
    year_counts = db.session.query(
        Publication.year, func.count(Publication.id)
    ).filter(Publication.id.in_(unique_pub_ids))\
     .group_by(Publication.year).order_by(Publication.year.asc()).all()
    
    # Get list of years for dropdown
    available_years = [year for year, count in year_counts]
    
    # Indexing status count - only unique publications
    indexing_counts = db.session.query(
        Publication.indexing_status, func.count(Publication.id)
    ).filter(Publication.id.in_(unique_pub_ids))\
     .filter(Publication.indexing_status != None)\
     .filter(Publication.indexing_status != '')\
     .group_by(Publication.indexing_status).all()
    
    # Quartile count - only unique publications
    quartile_counts = db.session.query(
        Publication.quartile, func.count(Publication.id)
    ).filter(Publication.id.in_(unique_pub_ids))\
     .filter(Publication.quartile != None)\
     .filter(Publication.quartile != '')\
     .group_by(Publication.quartile).all()
    
    # Department-wise count (for principal, admin, vice_principal, dean, director) - only unique publications
    dept_counts = []
    if current_user.role in ['principal', 'admin', 'vice_principal', 'dean', 'director']:
        # Get all departments dynamically
        all_departments = Department.query.order_by(Department.name).all()
        
        # Count publications for each department
        for dept in all_departments:
            pub_count = db.session.query(func.count(Publication.id))\
                .filter(Publication.department_id == dept.id)\
                .filter(Publication.id.in_(unique_pub_ids))\
                .scalar() or 0
            
            # Only include departments with publications or always show all
            dept_counts.append((dept.name, pub_count))
    
    # Top 5 Faculty by publication count - year-based
    if current_user.role == 'hod':
        faculty_query = db.session.query(
            User.id,
            User.name,
            User.email,
            func.count(Publication.id).label('pub_count')
        ).select_from(User).join(
            Publication, Publication.user_id == User.id
        ).filter(
            User.department_id == dept_id
        ).filter(Publication.id.in_(unique_pub_ids))
        
        if selected_year:
            faculty_query = faculty_query.filter(Publication.year == selected_year)
        
        top_faculty = faculty_query.group_by(User.id)\
                                  .order_by(func.count(Publication.id).desc())\
                                  .limit(5).all()
    else:  # principal or admin
        faculty_query = db.session.query(
            User.id,
            User.name,
            User.email,
            Department.name.label('dept_name'),
            func.count(Publication.id).label('pub_count')
        ).select_from(User).join(
            Publication, Publication.user_id == User.id
        ).join(
            Department, User.department_id == Department.id
        ).filter(Publication.id.in_(unique_pub_ids))
        
        if selected_year:
            faculty_query = faculty_query.filter(Publication.year == selected_year)
        
        top_faculty = faculty_query.group_by(User.id, Department.name)\
                                  .order_by(func.count(Publication.id).desc())\
                                  .limit(5).all()
    
    # Total citations
    total_citations = sum([p.citation_count or 0 for p in unique_publications])
    
    return render_template('statistics.html',
                         total_pubs=total_pubs,
                         year_counts=year_counts,
                         indexing_counts=indexing_counts,
                         quartile_counts=quartile_counts,
                         dept_counts=dept_counts,
                         total_citations=total_citations,
                         dept_name=dept_name,
                         top_faculty=top_faculty,
                         available_years=available_years,
                         selected_year=selected_year)


@app.route('/export/excel')
@login_required
def export_excel():
    # Get filter parameters
    dept_filter = request.args.get('department', type=int)
    year_filter = request.args.get('year', type=int)
    month_filter = request.args.get('month', type=int)
    indexing_filter = request.args.get('indexing')
    
    # Build base query based on role
    if current_user.role == 'faculty':
        query = Publication.query.filter_by(user_id=current_user.id)
    elif current_user.role == 'hod':
        query = Publication.query.filter_by(department_id=current_user.department_id)
    else:  # principal, dean, vice_principal
        query = Publication.query
    
    # Apply filters
    if dept_filter:
        query = query.filter_by(department_id=dept_filter)
    if year_filter:
        query = query.filter_by(year=year_filter)
    if month_filter:
        query = query.filter_by(month=month_filter)
    if indexing_filter:
        query = query.filter_by(indexing_status=indexing_filter)
    
    publications = query.all()
    
    # Prepare data for Excel
    data = []
    for pub in publications:
        data.append({
            'Title': pub.title,
            'Authors': pub.authors_names,
            'Journal/Conference': pub.journal_conference_name,
            'Publisher': pub.publisher_name,
            'Volume': pub.volume,
            'Issue': pub.issue,
            'Pages': pub.pages,
            'Year': pub.year,
            'Month': pub.month,
            'DOI': pub.doi,
            'Indexing Status': pub.indexing_status,
            'Quartile': pub.quartile,
            'Impact Factor': pub.impact_factor,
            'Citations': pub.citation_count,
            'Department': pub.department.name
        })
    
    df = pd.DataFrame(data)
    
    # Create Excel file in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Publications')
    output.seek(0)
    
    filename = f"SJEC_Publications_{now_ist().strftime('%Y%m%d')}.xlsx"
    return send_file(output, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=filename)


@app.route('/export/bibtex')
@login_required
def export_bibtex():
    # Get filter parameters
    dept_filter = request.args.get('department', type=int)
    year_filter = request.args.get('year', type=int)
    month_filter = request.args.get('month', type=int)
    indexing_filter = request.args.get('indexing')
    
    # Build base query based on role
    if current_user.role == 'faculty':
        query = Publication.query.filter_by(user_id=current_user.id)
    elif current_user.role == 'hod':
        query = Publication.query.filter_by(department_id=current_user.department_id)
    else:  # principal, dean, vice_principal
        query = Publication.query
    
    # Apply filters
    if dept_filter:
        query = query.filter_by(department_id=dept_filter)
    if year_filter:
        query = query.filter_by(year=year_filter)
    if month_filter:
        query = query.filter_by(month=month_filter)
    if indexing_filter:
        query = query.filter_by(indexing_status=indexing_filter)
    
    publications = query.all()
    
    # Generate BibTeX content
    bibtex_content = ""
    for pub in publications:
        if pub.bibtex_entry:
            bibtex_content += pub.bibtex_entry + "\n\n"
        else:
            # Generate basic BibTeX entry if not provided
            cite_key = f"{pub.authors_names.split(',')[0].strip().replace(' ', '')}_{pub.year}"
            bibtex_content += f"@article{{{cite_key},\n"
            bibtex_content += f"  title={{{pub.title}}},\n"
            bibtex_content += f"  author={{{pub.authors_names}}},\n"
            if pub.journal_conference_name:
                bibtex_content += f"  journal={{{pub.journal_conference_name}}},\n"
            if pub.volume:
                bibtex_content += f"  volume={{{pub.volume}}},\n"
            if pub.issue:
                bibtex_content += f"  number={{{pub.issue}}},\n"
            if pub.pages:
                bibtex_content += f"  pages={{{pub.pages}}},\n"
            bibtex_content += f"  year={{{pub.year}}},\n"
            if pub.publisher_name:
                bibtex_content += f"  publisher={{{pub.publisher_name}}},\n"
            if pub.doi:
                bibtex_content += f"  doi={{{pub.doi}}},\n"
            bibtex_content += "}\n\n"
    
    # Create file in memory
    output = BytesIO(bibtex_content.encode('utf-8'))
    output.seek(0)
    
    filename = f"SJEC_Publications_{now_ist().strftime('%Y%m%d')}.bib"
    return send_file(output,
                     mimetype='application/x-bibtex',
                     as_attachment=True,
                     download_name=filename)


@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    form = ChangePasswordForm()
    if form.validate_on_submit():
        if current_user.check_password(form.current_password.data):
            current_user.set_password(form.new_password.data)
            db.session.commit()
            flash('Your password has been updated successfully!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Current password is incorrect.', 'danger')
    return render_template('change_password.html', form=form)


@app.route('/delete_account', methods=['POST'])
@login_required
def delete_account():
    user_id = current_user.id
    user_name = current_user.name
    
    # Delete user's publications first (if faculty)
    if current_user.role == 'faculty':
        publications = Publication.query.filter_by(user_id=user_id).all()
        for pub in publications:
            # Delete PDF file if exists
            if pub.pdf_filename:
                pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], pub.pdf_filename)
                if os.path.exists(pdf_path):
                    os.remove(pdf_path)
            db.session.delete(pub)
    
    # Delete user account
    user = User.query.get(user_id)
    db.session.delete(user)
    db.session.commit()
    
    logout_user()
    flash(f'Account for {user_name} has been deleted successfully.', 'info')
    return redirect(url_for('login'))


@app.route('/publications/grant_edit/<int:id>', methods=['POST'])
@login_required
@role_required('hod', 'principal')
def grant_edit_permission(id):
    publication = Publication.query.get_or_404(id)
    
    # Check if HoD/Principal has access to this publication
    if current_user.role == 'hod' and publication.department_id != current_user.department_id:
        flash('You can only grant edit permission for publications in your department.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Check if edit permission was already granted before
    already_granted = EditRequest.query.filter_by(
        publication_id=publication.id,
        status='approved'
    ).first()
    
    if already_granted:
        flash(f'Edit permission has already been granted once for this publication on {(already_granted.responded_at|to_ist).strftime("%b %d, %Y at %I:%M %p")}. You can only grant edit permission ONCE per publication.', 'warning')
        return redirect(request.referrer or url_for('dashboard'))
    
    # Grant edit permission
    publication.can_edit = True
    publication.edit_granted_by = current_user.id
    publication.edit_granted_at = now_ist()
    
    db.session.commit()
    
    # Create notification for publication owner
    create_notification(
        user_id=publication.user_id,
        title='Edit Permission Granted',
        message=f'{current_user.name} has granted you permission to edit the publication "{publication.title[:50]}..."',
        notification_type='success',
        publication_id=publication.id
    )
    
    # Create audit log
    create_audit_log(
        action='grant_edit_permission',
        target_type='publication',
        target_id=publication.id,
        details={
            'title': publication.title,
            'granted_by': current_user.name,
            'owner': publication.author.name
        }
    )
    
    flash(f'Edit permission granted to {publication.author.name} for publication "{publication.title[:50]}..."', 'success')
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/publications/revoke_edit/<int:id>', methods=['POST'])
@login_required
@role_required('hod', 'principal')
def revoke_edit_permission(id):
    publication = Publication.query.get_or_404(id)
    
    # Check if HoD/Principal has access to this publication
    if current_user.role == 'hod' and publication.department_id != current_user.department_id:
        flash('You can only revoke edit permission for publications in your department.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Revoke edit permission
    publication.can_edit = False
    publication.edit_granted_by = None
    publication.edit_granted_at = None
    
    db.session.commit()
    
    # Create notification for publication owner
    create_notification(
        user_id=publication.user_id,
        title='Edit Permission Revoked',
        message=f'{current_user.name} has revoked edit permission for the publication "{publication.title[:50]}..."',
        notification_type='warning',
        publication_id=publication.id
    )
    
    # Create audit log
    create_audit_log(
        action='revoke_edit_permission',
        target_type='publication',
        target_id=publication.id,
        details={
            'title': publication.title,
            'revoked_by': current_user.name,
            'owner': publication.author.name
        }
    )
    
    flash(f'Edit permission revoked for "{publication.title[:50]}..."', 'info')
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/faq')
def faq():
    """FAQ page accessible to everyone (including non-logged-in users on login page)"""
    from models import FAQ
    
    # Get all active FAQs grouped by category
    faqs = FAQ.query.filter_by(is_active=True).order_by(FAQ.category, FAQ.display_order, FAQ.id).all()
    
    # Group FAQs by category
    faq_categories = {}
    for faq_item in faqs:
        if faq_item.category not in faq_categories:
            faq_categories[faq_item.category] = []
        faq_categories[faq_item.category].append(faq_item)
    
    return render_template('faq.html', faq_categories=faq_categories)


@app.route('/faq/view/<int:faq_id>')
def faq_view(faq_id):
    """Increment view count for FAQ"""
    from models import FAQ
    
    faq_item = FAQ.query.get_or_404(faq_id)
    faq_item.view_count += 1
    db.session.commit()
    
    return jsonify({'success': True, 'views': faq_item.view_count})


@app.route('/admin/faq_management')
@login_required
def admin_faq_management():
    """Admin FAQ management interface"""
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    from models import FAQ
    
    faqs = FAQ.query.order_by(FAQ.category, FAQ.display_order, FAQ.id).all()
    
    return render_template('admin_faq_management.html', faqs=faqs)


@app.route('/admin/faq/add', methods=['POST'])
@login_required
def admin_faq_add():
    """Add new FAQ"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    from models import FAQ
    
    try:
        category = request.form.get('category', '').strip()
        question = request.form.get('question', '').strip()
        answer = request.form.get('answer', '').strip()
        display_order = int(request.form.get('display_order', 0))
        
        if not all([category, question, answer]):
            flash('All fields are required!', 'danger')
            return redirect(url_for('admin_faq_management'))
        
        new_faq = FAQ(
            category=category,
            question=question,
            answer=answer,
            display_order=display_order,
            created_by=current_user.id
        )
        
        db.session.add(new_faq)
        db.session.commit()
        
        # Create audit log
        create_audit_log('create_faq', 'faq', new_faq.id, {
            'category': category,
            'question': question[:100]
        })
        
        flash('âœ… FAQ added successfully!', 'success')
        return redirect(url_for('admin_faq_management'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding FAQ: {str(e)}', 'danger')
        return redirect(url_for('admin_faq_management'))


@app.route('/admin/faq/edit/<int:faq_id>', methods=['POST'])
@login_required
def admin_faq_edit(faq_id):
    """Edit existing FAQ"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    from models import FAQ
    
    try:
        faq_item = FAQ.query.get_or_404(faq_id)
        
        faq_item.category = request.form.get('category', '').strip()
        faq_item.question = request.form.get('question', '').strip()
        faq_item.answer = request.form.get('answer', '').strip()
        faq_item.display_order = int(request.form.get('display_order', 0))
        
        db.session.commit()
        
        # Create audit log
        create_audit_log('update_faq', 'faq', faq_id, {
            'category': faq_item.category,
            'question': faq_item.question[:100]
        })
        
        flash('âœ… FAQ updated successfully!', 'success')
        return redirect(url_for('admin_faq_management'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating FAQ: {str(e)}', 'danger')
        return redirect(url_for('admin_faq_management'))


@app.route('/admin/faq/delete/<int:faq_id>', methods=['POST'])
@login_required
def admin_faq_delete(faq_id):
    """Delete FAQ (soft delete by setting is_active=False)"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    from models import FAQ
    
    try:
        faq_item = FAQ.query.get_or_404(faq_id)
        faq_item.is_active = False
        db.session.commit()
        
        # Create audit log
        create_audit_log('delete_faq', 'faq', faq_id, {
            'category': faq_item.category,
            'question': faq_item.question[:100]
        })
        
        flash('âœ… FAQ deleted successfully!', 'success')
        return redirect(url_for('admin_faq_management'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting FAQ: {str(e)}', 'danger')
        return redirect(url_for('admin_faq_management'))


@app.route('/user-guide')
def user_guide():
    """User Guide page with comprehensive portal documentation - accessible to all"""
    return render_template('user_guide.html')


@app.route('/contact-us')
def contact_us():
    """Contact Us page with developer and support information - accessible to all"""
    return render_template('contact_us.html')


@app.route('/notifications')
@login_required
def notifications():
    # Get all notifications for current user, ordered by newest first
    user_notifications = Notification.query.filter_by(user_id=current_user.id)\
        .order_by(Notification.created_at.desc()).all()
    
    unread_count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    
    return render_template('notifications.html', 
                         notifications=user_notifications,
                         unread_count=unread_count)


@app.route('/notifications/mark_read/<int:id>', methods=['POST'])
@login_required
def mark_notification_read(id):
    notification = Notification.query.get_or_404(id)
    
    # Ensure user can only mark their own notifications
    if notification.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    notification.is_read = True
    db.session.commit()
    
    return jsonify({'success': True})


@app.route('/notifications/mark_all_read', methods=['POST'])
@login_required
def mark_all_notifications_read():
    Notification.query.filter_by(user_id=current_user.id, is_read=False)\
        .update({'is_read': True})
    db.session.commit()
    
    return jsonify({'success': True})


@app.route('/notifications/delete/<int:id>', methods=['POST'])
@login_required
def delete_notification(id):
    notification = Notification.query.get_or_404(id)
    
    # Ensure user can only delete their own notifications
    if notification.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    db.session.delete(notification)
    db.session.commit()
    
    return jsonify({'success': True})


@app.route('/notifications/clear_all', methods=['POST'])
@login_required
def clear_all_notifications():
    # Delete all notifications for current user
    deleted_count = Notification.query.filter_by(user_id=current_user.id).delete()
    db.session.commit()
    
    return jsonify({'success': True, 'deleted': deleted_count})


@app.route('/api/notifications/unread_count')
@login_required
def get_unread_count():
    count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    
    # Add pending feedbacks count for HoD/Principal/Admin
    if current_user.role in ['hod', 'principal', 'admin']:
        feedback_count = 0
        if current_user.role == 'hod':
            feedback_count = Feedback.query.filter_by(recipient_role='hod', status='pending').count()
        elif current_user.role == 'principal':
            feedback_count = Feedback.query.filter_by(recipient_role='principal', status='pending').count()
        elif current_user.role == 'admin':
            feedback_count = Feedback.query.filter_by(recipient_role='admin', status='pending').count()
        count += feedback_count
    
    # Add pending edit requests for HoD
    if current_user.role == 'hod':
        edit_request_count = EditRequest.query.join(Publication).filter(
            Publication.department_id == current_user.department_id,
            EditRequest.status == 'pending'
        ).count()
        count += edit_request_count
    
    return jsonify({'count': count})


@app.route('/api/check_updates')
@login_required
def check_updates():
    """Check for new publications, notifications, feedback, and edit requests"""
    response_data = {}
    
    # For faculty - check their own publications
    if current_user.role == 'faculty':
        publication_count = Publication.query.filter_by(user_id=current_user.id).count()
        response_data['publication_count'] = publication_count
    
    # For HOD - check department publications and edit requests
    elif current_user.role == 'hod':
        publication_count = Publication.query.filter_by(department_id=current_user.department_id).count()
        edit_request_count = EditRequest.query.join(Publication).filter(
            Publication.department_id == current_user.department_id
        ).count()
        response_data['publication_count'] = publication_count
        response_data['edit_request_count'] = edit_request_count
    
    # For principal and admin - check all publications
    elif current_user.role in ['principal', 'admin']:
        publication_count = Publication.query.count()
        response_data['publication_count'] = publication_count
    
    # Check notifications for all roles
    notification_count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    response_data['notification_count'] = notification_count
    
    # Check feedback for HOD/Principal/Admin
    if current_user.role in ['hod', 'principal', 'admin']:
        if current_user.role == 'hod':
            feedback_count = Feedback.query.filter_by(
                recipient_role='hod',
                department_id=current_user.department_id,
                status='pending'
            ).count()
        elif current_user.role == 'principal':
            feedback_count = Feedback.query.filter_by(recipient_role='principal', status='pending').count()
        else:  # admin
            feedback_count = Feedback.query.filter_by(recipient_role='admin', status='pending').count()
        response_data['feedback_count'] = feedback_count
    
    return jsonify(response_data)


# ============================================================================
# PROFILE ROUTES
# ============================================================================

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'update_name':
            new_name = request.form.get('name', '').strip()
            if new_name:
                current_user.name = new_name
                db.session.commit()
                flash('Name updated successfully!', 'success')
                create_audit_log('update_profile', 'user', current_user.id, {'field': 'name'})
            else:
                flash('Name cannot be empty!', 'danger')
        
        elif action == 'upload_profile_picture':
            if 'profile_picture' not in request.files:
                flash('No file selected!', 'danger')
            else:
                file = request.files['profile_picture']
                if file.filename == '':
                    flash('No file selected!', 'danger')
                elif not allowed_file(file.filename, 'image'):
                    flash('Only image files (JPG, JPEG, PNG, GIF) are allowed!', 'danger')
                else:
                    # Delete old profile picture if exists
                    if current_user.profile_picture:
                        old_file_path = os.path.join(get_user_folder(current_user.id, 'profiles'), 
                                                     current_user.profile_picture)
                        if os.path.exists(old_file_path):
                            os.remove(old_file_path)
                    
                    # Save new profile picture
                    filename = secure_filename(file.filename)
                    file_ext = filename.rsplit('.', 1)[1].lower()
                    new_filename = f'user_{current_user.id}_profile.{file_ext}'
                    file_path = os.path.join(get_user_folder(current_user.id, 'profiles'), new_filename)
                    file.save(file_path)
                    
                    current_user.profile_picture = new_filename
                    db.session.commit()
                    flash('Profile picture uploaded successfully!', 'success')
                    create_audit_log('upload_profile_picture', 'user', current_user.id)
        
        elif action == 'upload_signature':
            if 'signature' not in request.files:
                flash('No file selected!', 'danger')
            else:
                file = request.files['signature']
                if file.filename == '':
                    flash('No file selected!', 'danger')
                elif not allowed_file(file.filename, 'image'):
                    flash('Only image files (JPG, JPEG, PNG, GIF) are allowed!', 'danger')
                else:
                    # Delete old signature if exists
                    if current_user.signature:
                        old_file_path = os.path.join(get_user_folder(current_user.id, 'signatures'), 
                                                     current_user.signature)
                        if os.path.exists(old_file_path):
                            os.remove(old_file_path)
                    
                    # Save new signature
                    filename = secure_filename(file.filename)
                    file_ext = filename.rsplit('.', 1)[1].lower()
                    new_filename = f'user_{current_user.id}_signature.png'  # Always save as PNG
                    file_path = os.path.join(get_user_folder(current_user.id, 'signatures'), new_filename)
                    file.save(file_path)
                    
                    # Process signature: remove background, enhance, and optimize
                    try:
                        from process_signature import remove_background_and_enhance
                        remove_background_and_enhance(file_path, file_path, target_width=400, target_height=150)
                    except Exception as e:
                        print(f"Signature processing error: {e}")
                        # Continue even if processing fails - signature is still uploaded
                    
                    current_user.signature = new_filename
                    db.session.commit()
                    flash('Signature uploaded and processed successfully!', 'success')
                    create_audit_log('upload_signature', 'user', current_user.id)
        
        elif action == 'change_password':
            old_password = request.form.get('old_password')
            new_password = request.form.get('new_password')
            confirm_password = request.form.get('confirm_password')
            
            if not current_user.check_password(old_password):
                flash('Current password is incorrect!', 'danger')
            elif new_password != confirm_password:
                flash('New passwords do not match!', 'danger')
            elif len(new_password) < 6:
                flash('Password must be at least 6 characters long!', 'danger')
            else:
                current_user.set_password(new_password)
                db.session.commit()
                flash('Password changed successfully!', 'success')
                create_audit_log('change_password', 'user', current_user.id)
        
        return redirect(url_for('profile'))
    
    return render_template('profile.html')


@app.route('/my-wallet')
@login_required
@role_required('faculty')
def my_wallet():
    """Faculty wallet dashboard showing RRF, EPP, and transaction history"""
    from models import WalletTransaction, IncentiveConfig, ApplicationForm
    
    # Get wallet balances
    rrf_balance = current_user.wallet_balance or 0
    epp_balance = current_user.epp_balance or 0
    
    # Get configuration for EPP to INR conversion
    config = IncentiveConfig.query.first()
    epp_to_inr = config.epp_to_inr_rate if config else 2000
    
    # Calculate totals
    epp_inr_value = epp_balance * epp_to_inr
    total_available = rrf_balance + epp_inr_value
    
    # Get transaction history
    transactions = WalletTransaction.query.filter_by(
        user_id=current_user.id
    ).order_by(WalletTransaction.created_at.desc()).all()
    
    # Get applications that resulted in wallet transactions
    approved_applications = ApplicationForm.query.filter(
        ApplicationForm.applicant_id == current_user.id,
        ApplicationForm.wallet_debited == True
    ).order_by(ApplicationForm.principal_approved_at.desc()).all()
    
    # Get EPP-earning publications
    epp_earning_apps = ApplicationForm.query.filter(
        ApplicationForm.applicant_id == current_user.id,
        ApplicationForm.epp_awarded.isnot(None),
        ApplicationForm.epp_awarded > 0
    ).order_by(ApplicationForm.principal_approved_at.desc()).all()
    
    # Calculate statistics
    total_rrf_received = 0
    total_rrf_spent = 0
    total_epp_earned = 0
    total_epp_spent = 0
    
    for txn in transactions:
        if txn.transaction_type == 'credit':
            total_rrf_received += txn.amount
        elif txn.transaction_type == 'debit':
            total_rrf_spent += txn.amount
    
    for app in epp_earning_apps:
        total_epp_earned += app.epp_awarded or 0
    
    # EPP spent (calculated from transactions where EPP was used)
    # This would need to be tracked separately, for now we can calculate it
    total_epp_spent = total_epp_earned - epp_balance
    
    return render_template('my_wallet.html',
                         rrf_balance=rrf_balance,
                         epp_balance=epp_balance,
                         epp_to_inr=epp_to_inr,
                         epp_inr_value=epp_inr_value,
                         total_available=total_available,
                         transactions=transactions,
                         approved_applications=approved_applications,
                         epp_earning_apps=epp_earning_apps,
                         total_rrf_received=total_rrf_received,
                         total_rrf_spent=total_rrf_spent,
                         total_epp_earned=total_epp_earned,
                         total_epp_spent=total_epp_spent,
                         config=config)


@app.route('/my-statistics')
@login_required
@role_required('faculty')
def my_statistics():
    """Faculty performance and evaluation metrics page"""
    # Get publication statistics
    all_pubs = Publication.query.filter_by(user_id=current_user.id).all()
    total_publications = len(all_pubs)
    scopus_count = len([p for p in all_pubs if p.indexing_status == 'Scopus'])
    total_citations = sum([p.citation_count or 0 for p in all_pubs])
    
    # Calculate average citations per publication
    avg_citations = (total_citations / total_publications) if total_publications > 0 else 0
    scopus_percentage = round((scopus_count / total_publications * 100), 1) if total_publications > 0 else 0
    
    # ============================================================
    # PUBLICATION ENGAGEMENT METRICS (PEM)
    # ============================================================
    
    # Calculate h-index
    # h-index: A researcher has index h if h of their N papers have at least h citations each
    citation_counts = sorted([p.citation_count or 0 for p in all_pubs], reverse=True)
    h_index = 0
    for i, citations in enumerate(citation_counts, start=1):
        if citations >= i:
            h_index = i
        else:
            break
    
    # Calculate i10-index
    # i10-index: Number of publications with at least 10 citations
    i10_index = len([p for p in all_pubs if (p.citation_count or 0) >= 10])
    
    # Calculate quartile distribution
    q1_count = len([p for p in all_pubs if p.quartile == 'Q1'])
    q2_count = len([p for p in all_pubs if p.quartile == 'Q2'])
    q3_count = len([p for p in all_pubs if p.quartile == 'Q3'])
    q4_count = len([p for p in all_pubs if p.quartile == 'Q4'])
    
    # Calculate impact factor metrics
    pubs_with_if = [p for p in all_pubs if p.impact_factor and p.impact_factor > 0]
    avg_impact_factor = (sum([p.impact_factor for p in pubs_with_if]) / len(pubs_with_if)) if pubs_with_if else 0
    max_impact_factor = max([p.impact_factor for p in pubs_with_if], default=0) if pubs_with_if else 0
    
    # Calculate citations per year (research productivity)
    current_year = 2025
    years_active = len(set([p.year for p in all_pubs])) if all_pubs else 1
    citations_per_year = total_citations / years_active if years_active > 0 else 0
    
    # Store PEM metrics
    pem_metrics = {
        'h_index': h_index,
        'i10_index': i10_index,
        'q1_count': q1_count,
        'q2_count': q2_count,
        'q3_count': q3_count,
        'q4_count': q4_count,
        'avg_impact_factor': round(avg_impact_factor, 3),
        'max_impact_factor': round(max_impact_factor, 3),
        'citations_per_year': round(citations_per_year, 1),
        'years_active': years_active
    }
    
    # Get application statistics
    conference_apps = ApplicationForm.query.filter_by(
        applicant_id=current_user.id, 
        application_type='Conference Registration'
    ).count()
    
    # Count approved applications (principal_approved or director_approved)
    conference_approved = ApplicationForm.query.filter(
        ApplicationForm.applicant_id == current_user.id,
        ApplicationForm.application_type == 'Conference Registration',
        (ApplicationForm.principal_approved == True) | (ApplicationForm.director_approved == True)
    ).count()
    
    incentive_apps = ApplicationForm.query.filter_by(
        applicant_id=current_user.id,
        application_type='Publication Incentive'
    ).count()
    
    incentive_approved = ApplicationForm.query.filter(
        ApplicationForm.applicant_id == current_user.id,
        ApplicationForm.application_type == 'Publication Incentive',
        (ApplicationForm.principal_approved == True) | (ApplicationForm.director_approved == True)
    ).count()
    
    journal_apps = ApplicationForm.query.filter_by(
        applicant_id=current_user.id,
        application_type='Journal Support'
    ).count()
    
    journal_approved = ApplicationForm.query.filter(
        ApplicationForm.applicant_id == current_user.id,
        ApplicationForm.application_type == 'Journal Support',
        (ApplicationForm.principal_approved == True) | (ApplicationForm.director_approved == True)
    ).count()
    
    # Create performance statistics dictionary
    performance_stats = {
        'total_publications': total_publications,
        'scopus_count': scopus_count,
        'scopus_percentage': scopus_percentage,
        'total_citations': total_citations,
        'avg_citations': avg_citations,
        'conference_apps': conference_apps,
        'conference_approved': conference_approved,
        'incentive_apps': incentive_apps,
        'incentive_approved': incentive_approved,
        'journal_apps': journal_apps,
        'journal_approved': journal_approved
    }
    
    return render_template('my_statistics.html', 
                         performance_stats=performance_stats,
                         pem_metrics=pem_metrics)


# ============================================================================
# USER MEDIA ROUTES
# ============================================================================

@app.route('/user/profile-picture/<int:user_id>')
@login_required
def user_profile_picture(user_id):
    """Serve user profile picture"""
    user = User.query.get_or_404(user_id)
    if user.profile_picture:
        file_path = os.path.join(get_user_folder(user_id, 'profiles'), user.profile_picture)
        if os.path.exists(file_path):
            return send_file(file_path)
    # Return default avatar if no profile picture
    return redirect(url_for('static', filename='images/default-avatar.png'))


@app.route('/user/signature/<int:user_id>')
@login_required
def user_signature(user_id):
    """Serve user signature"""
    user = User.query.get_or_404(user_id)
    if user.signature:
        file_path = os.path.join(get_user_folder(user_id, 'signatures'), user.signature)
        if os.path.exists(file_path):
            return send_file(file_path)
    return "No signature uploaded", 404


# ============================================================================
# VIEW USERS ROUTES
# ============================================================================

@app.route('/view_users')
@login_required
def view_users():
    if current_user.role not in ['hod', 'principal', 'vice_principal', 'dean', 'director']:
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))
    
    if current_user.role == 'hod':
        # HoD sees only department users
        users = User.query.filter_by(department_id=current_user.department_id)\
            .order_by(User.name).all()
    else:
        # Principal, Vice Principal, and Dean see all users
        users = User.query.order_by(User.department_id, User.name).all()
    
    # Get publication counts for each user
    user_stats = []
    for user in users:
        pub_count = Publication.query.filter_by(user_id=user.id).count()
        user_stats.append({
            'user': user,
            'publication_count': pub_count
        })
    
    return render_template('view_users.html', user_stats=user_stats)


# ============================================================================
# EDIT REQUEST ROUTES
# ============================================================================

@app.route('/publications/<int:pub_id>/request_edit', methods=['POST'])
@login_required
def request_edit_permission(pub_id):
    publication = Publication.query.get_or_404(pub_id)
    
    # Check if user owns this publication
    if publication.user_id != current_user.id:
        flash('You can only request edit for your own publications!', 'danger')
        return redirect(url_for('dashboard'))
    
    # Check if publication is confirmed
    if publication.status != 'confirmed':
        flash('You can only request edit for confirmed publications!', 'warning')
        return redirect(url_for('dashboard'))
    
    # Check if already has pending request
    existing_request = EditRequest.query.filter_by(
        publication_id=pub_id,
        faculty_id=current_user.id
    ).filter(EditRequest.status.in_(['pending', 'pending_admin'])).first()
    
    if existing_request:
        if existing_request.status == 'pending_admin':
            flash('Your second edit request is pending with Admin for approval!', 'warning')
        else:
            flash('You already have a pending edit request for this publication!', 'warning')
        return redirect(url_for('dashboard'))
    
    # Determine if this is first or second edit request
    previous_approved = EditRequest.query.filter_by(
        publication_id=pub_id,
        faculty_id=current_user.id,
        status='approved'
    ).first()
    
    # Count how many times edit was granted (first level only)
    first_level_approved = EditRequest.query.filter_by(
        publication_id=pub_id,
        faculty_id=current_user.id,
        edit_level=1,
        status='approved'
    ).count()
    
    # Count second level requests
    second_level_requests = EditRequest.query.filter_by(
        publication_id=pub_id,
        faculty_id=current_user.id,
        edit_level=2
    ).count()
    
    # Check if already requested twice (1st by HoD, 2nd by Admin)
    if second_level_requests > 0:
        flash('You have already requested a second edit. Maximum 2 edit requests allowed per publication.', 'warning')
        return redirect(url_for('dashboard'))
    
    # Check if first level was denied
    first_denied = EditRequest.query.filter_by(
        publication_id=pub_id,
        faculty_id=current_user.id,
        edit_level=1,
        status='denied'
    ).first()
    
    if first_denied:
        flash('Your first edit request was denied. You cannot request again.', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get HoD of the department
    hod = User.query.filter_by(
        department_id=current_user.department_id,
        role='hod'
    ).first()
    
    if not hod:
        flash('No HoD found for your department!', 'danger')
        return redirect(url_for('dashboard'))
    
    # Determine edit level
    edit_level = 2 if first_level_approved else 1
    
    # Create edit request
    reason = request.form.get('reason', '').strip()
    edit_request = EditRequest(
        publication_id=pub_id,
        faculty_id=current_user.id,
        hod_id=hod.id,
        reason=reason,
        edit_level=edit_level,
        status='pending'
    )
    db.session.add(edit_request)
    db.session.commit()
    
    # Send notification to HoD
    level_text = "SECOND" if edit_level == 2 else "first"
    create_notification(
        user_id=hod.id,
        title=f'{"âš ï¸ SECOND" if edit_level == 2 else ""} Edit Request',
        message=f'{current_user.name} has requested permission for {level_text} edit of publication: {publication.title[:50]}...' + 
                (f'\n\nâš ï¸ This is a SECOND EDIT request. If approved, it must be forwarded to Admin for final approval.' if edit_level == 2 else ''),
        notification_type='warning' if edit_level == 2 else 'info',
        publication_id=pub_id
    )
    
    # Create audit log
    create_audit_log('request_edit', 'publication', pub_id, {
        'faculty_id': current_user.id,
        'hod_id': hod.id,
        'edit_level': edit_level
    })
    
    if edit_level == 2:
        flash('Second edit request sent to HoD. If HoD approves, it will be forwarded to Admin for final approval.', 'info')
    else:
        flash('Edit request sent to HoD successfully!', 'success')
    return redirect(url_for('dashboard'))


@app.route('/edit_requests/pending')
@login_required
def pending_edit_requests():
    if current_user.role != 'hod':
        flash('Access denied. HoD privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get pending edit requests for this HoD's department
    requests = EditRequest.query.join(Publication).filter(
        Publication.department_id == current_user.department_id,
        EditRequest.status == 'pending'
    ).order_by(EditRequest.requested_at.desc()).all()
    
    return render_template('pending_edit_requests.html', requests=requests)


@app.route('/edit_requests/<int:request_id>/approve', methods=['POST'])
@login_required
def approve_edit_request(request_id):
    if current_user.role != 'hod':
        flash('Access denied. HoD privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    edit_request = EditRequest.query.get_or_404(request_id)
    
    # Verify HoD owns this request
    if edit_request.hod_id != current_user.id:
        flash('You can only approve requests assigned to you!', 'danger')
        return redirect(url_for('pending_edit_requests'))
    
    publication = edit_request.publication
    
    # Check if this is a SECOND edit request (edit_level = 2)
    if edit_request.edit_level == 2:
        # For second edit, HoD forwards to Admin instead of directly approving
        edit_request.status = 'pending_admin'
        edit_request.responded_at = now_ist()
        edit_request.response = request.form.get('response', '').strip()
        
        # Get admin user
        admin = User.query.filter_by(role='admin').first()
        if admin:
            edit_request.admin_id = admin.id
            
            # Notify admin
            hod_comment = edit_request.response if edit_request.response else 'No additional comment from HoD'
            create_notification(
                user_id=admin.id,
                title='âš ï¸ Second Edit Request - Admin Approval Required',
                message=f'HoD has forwarded a SECOND edit request for your approval.\n\nPublication: "{publication.title[:50]}..."\nFaculty: {edit_request.faculty.name}\nFaculty Reason: {edit_request.reason or "Not specified"}\n\nHoD Comment: {hod_comment}',
                notification_type='warning',
                publication_id=publication.id
            )
            
            # Notify faculty that request is forwarded to admin
            create_notification(
                user_id=edit_request.faculty_id,
                title='Edit Request Forwarded to Admin',
                message=f'Your second edit request for "{publication.title[:50]}..." has been approved by HoD and forwarded to Admin for final approval.\n\nHoD Response: {hod_comment}',
                notification_type='info',
                publication_id=publication.id
            )
            
            db.session.commit()
            flash('Second edit request approved and forwarded to Admin for final approval!', 'success')
        else:
            flash('No Admin found in the system!', 'danger')
            return redirect(url_for('pending_edit_requests'))
    else:
        # First edit - HoD can directly approve (existing logic)
        # Check if edit permission was already granted before for this publication (first level)
        already_granted = EditRequest.query.filter_by(
            publication_id=edit_request.publication_id,
            edit_level=1,
            status='approved'
        ).filter(EditRequest.id != request_id).first()
        
        if already_granted:
            flash(f'First edit permission has already been granted for this publication.', 'warning')
            return redirect(url_for('pending_edit_requests'))
        
        # Approve the request
        edit_request.status = 'approved'
        edit_request.responded_at = now_ist()
        edit_request.response = request.form.get('response', '').strip()
        
        # Grant one-time edit permission
        publication.can_edit = True
        publication.edit_granted_by = current_user.id
        publication.edit_granted_at = now_ist()
        
        db.session.commit()
        
        # Notify faculty with response comment
        response_text = edit_request.response if edit_request.response else 'No additional comment provided.'
        create_notification(
            user_id=edit_request.faculty_id,
            title='Edit Request Approved',
            message=f'Your edit request for "{publication.title[:50]}..." has been approved by HoD. You can now edit this publication once.\n\nHoD Response: {response_text}',
            notification_type='success',
            publication_id=publication.id
        )
        
        flash('Edit request approved successfully!', 'success')
    
    # Create audit log
    create_audit_log('approve_edit_request', 'publication', publication.id, {
        'request_id': request_id,
        'faculty_id': edit_request.faculty_id,
        'edit_level': edit_request.edit_level,
        'status': edit_request.status
    })
    
    return redirect(url_for('pending_edit_requests'))


@app.route('/edit_requests/<int:request_id>/deny', methods=['POST'])
@login_required
def deny_edit_request(request_id):
    if current_user.role != 'hod':
        flash('Access denied. HoD privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    edit_request = EditRequest.query.get_or_404(request_id)
    
    # Verify HoD owns this request
    if edit_request.hod_id != current_user.id:
        flash('You can only deny requests assigned to you!', 'danger')
        return redirect(url_for('pending_edit_requests'))
    
    # Deny the request
    edit_request.status = 'denied'
    edit_request.responded_at = now_ist()
    edit_request.response = request.form.get('response', '').strip()
    
    db.session.commit()
    
    # Notify faculty with response comment
    response_text = edit_request.response if edit_request.response else 'No reason provided.'
    create_notification(
        user_id=edit_request.faculty_id,
        title='Edit Request Denied',
        message=f'Your edit request for "{edit_request.publication.title[:50]}..." has been denied by HoD.\n\nHoD Response: {response_text}',
        notification_type='warning',
        publication_id=edit_request.publication_id
    )
    
    # Create audit log
    create_audit_log('deny_edit_request', 'publication', edit_request.publication_id, {
        'request_id': request_id,
        'faculty_id': edit_request.faculty_id
    })
    
    flash('Edit request denied.', 'info')
    return redirect(url_for('pending_edit_requests'))


# ============================================================================
# FEEDBACK ROUTES
# ============================================================================

@app.route('/feedback/submit', methods=['POST'])
@login_required
def submit_feedback():
    category = request.form.get('category')
    subject = request.form.get('subject', '').strip()
    message = request.form.get('message', '').strip()
    recipient_role = request.form.get('recipient_role')
    
    if not subject or not message:
        return jsonify({'success': False, 'message': 'Subject and message are required!'})
    
    if recipient_role not in ['hod', 'dean', 'principal', 'admin']:
        return jsonify({'success': False, 'message': 'Invalid recipient!'})
    
    # Handle screenshot upload if provided
    screenshot_filename = None
    if 'screenshot' in request.files:
        screenshot = request.files['screenshot']
        if screenshot and screenshot.filename:
            filename = secure_filename(f"feedback_{current_user.id}_{now_ist().strftime('%Y%m%d_%H%M%S')}_{screenshot.filename}")
            feedback_folder = get_user_folder(current_user.id, 'feedback_screenshots')
            screenshot_path = os.path.join(feedback_folder, filename)
            screenshot.save(screenshot_path)
            screenshot_filename = filename
    
    # Create feedback
    feedback = Feedback(
        user_id=current_user.id,
        department_id=current_user.department_id,  # Store sender's department
        recipient_role=recipient_role,
        category=category,
        subject=subject,
        message=message,
        screenshot_filename=screenshot_filename
    )
    db.session.add(feedback)
    db.session.commit()
    
    # Notify recipients
    if recipient_role == 'hod':
        # Notify HoD of user's department
        hod = User.query.filter_by(department_id=current_user.department_id, role='hod').first()
        if hod:
            create_notification(
                user_id=hod.id,
                title=f'New {category.title()} Report',
                message=f'{current_user.name} reported: {subject}',
                notification_type='info'
            )
    elif recipient_role == 'dean':
        # Notify all deans
        deans = User.query.filter_by(role='dean').all()
        for dean in deans:
            create_notification(
                user_id=dean.id,
                title=f'New {category.title()} Report',
                message=f'{current_user.name} reported: {subject}',
                notification_type='info'
            )
    elif recipient_role == 'principal':
        # Notify all principals
        principals = User.query.filter_by(role='principal').all()
        for principal in principals:
            create_notification(
                user_id=principal.id,
                title=f'New {category.title()} Report',
                message=f'{current_user.name} reported: {subject}',
                notification_type='info'
            )
    elif recipient_role == 'admin':
        # Notify all admins
        admins = User.query.filter_by(role='admin').all()
        for admin in admins:
            create_notification(
                user_id=admin.id,
                title=f'New {category.title()} Report',
                message=f'{current_user.name} reported: {subject}',
                notification_type='info'
            )
    
    # Create audit log
    create_audit_log('submit_feedback', 'feedback', feedback.id, {
        'category': category,
        'recipient_role': recipient_role
    })
    
    return jsonify({'success': True, 'message': 'Feedback submitted successfully!'})


@app.route('/feedback/my_feedback')
@login_required
def my_feedback():
    """Faculty can view their own submitted feedback and responses"""
    if current_user.role not in ['faculty', 'hod', 'principal', 'admin']:
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get user's own submitted feedback
    my_feedbacks = Feedback.query.filter_by(user_id=current_user.id)\
        .order_by(Feedback.created_at.desc()).all()
    
    return render_template('my-feedback.html', feedbacks=my_feedbacks)


@app.route('/feedback/list')
@login_required
def list_feedbacks():
    if current_user.role not in ['hod', 'principal', 'director', 'admin', 'dean']:
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get feedbacks based on role
    if current_user.role == 'hod':
        # HoD sees only feedback from their department
        feedbacks = Feedback.query.filter_by(
            recipient_role='hod',
            department_id=current_user.department_id  # Filter by HoD's department
        ).order_by(Feedback.created_at.desc()).all()
    elif current_user.role == 'principal':
        feedbacks = Feedback.query.filter_by(recipient_role='principal')\
            .order_by(Feedback.created_at.desc()).all()
    elif current_user.role == 'director':
        # Director sees same feedbacks as principal
        feedbacks = Feedback.query.filter_by(recipient_role='principal')\
            .order_by(Feedback.created_at.desc()).all()
    elif current_user.role == 'dean':
        # Dean sees all feedbacks
        feedbacks = Feedback.query.order_by(Feedback.created_at.desc()).all()
    else:  # admin
        feedbacks = Feedback.query.filter_by(recipient_role='admin')\
            .order_by(Feedback.created_at.desc()).all()
    
    return render_template('feedbacks.html', feedbacks=feedbacks)


@app.route('/feedback/<int:feedback_id>/resolve', methods=['POST'])
@login_required
def resolve_feedback(feedback_id):
    if current_user.role not in ['hod', 'principal', 'director', 'admin', 'dean']:
        return jsonify({'success': False, 'message': 'Access denied!'})
    
    feedback = Feedback.query.get_or_404(feedback_id)
    
    # Verify user can resolve this feedback (Dean can resolve any feedback, Director same as Principal)
    if current_user.role not in ['dean', 'director'] and feedback.recipient_role != current_user.role:
        return jsonify({'success': False, 'message': 'You can only resolve feedback sent to your role!'})
    
    # Director can resolve principal feedbacks
    if current_user.role == 'director' and feedback.recipient_role != 'principal':
        return jsonify({'success': False, 'message': 'You can only resolve feedback sent to principal!'})
    
    # Resolve the feedback
    feedback.status = 'resolved'
    feedback.resolved_by = current_user.id
    feedback.resolved_at = now_ist()
    feedback.resolution_comment = request.form.get('comment', '').strip()
    
    db.session.commit()
    
    # Notify the reporter with the response comment
    response_text = feedback.resolution_comment if feedback.resolution_comment else 'No additional comment provided.'
    create_notification(
        user_id=feedback.user_id,
        title='Your Report has been Resolved',
        message=f'Your report "{feedback.subject}" has been resolved by {current_user.role.upper()}.\n\nResponse: {response_text}',
        notification_type='success'
    )
    
    # Create audit log
    create_audit_log('resolve_feedback', 'feedback', feedback_id, {
        'resolver_id': current_user.id,
        'comment': feedback.resolution_comment
    })
    
    # Flash success message
    flash('Feedback resolved successfully! The user has been notified.', 'success')
    
    return jsonify({'success': True, 'message': 'Feedback resolved successfully!'})


@app.route('/feedback/<int:feedback_id>/delete', methods=['POST'])
@login_required
def delete_feedback(feedback_id):
    if current_user.role not in ['hod', 'principal', 'director', 'admin']:
        return jsonify({'success': False, 'message': 'Access denied!'})
    
    feedback = Feedback.query.get_or_404(feedback_id)
    
    # Verify user can delete this feedback (Director can delete principal feedbacks)
    if current_user.role == 'director':
        if feedback.recipient_role != 'principal':
            return jsonify({'success': False, 'message': 'You can only delete feedback sent to principal!'})
    elif feedback.recipient_role != current_user.role:
        return jsonify({'success': False, 'message': 'You can only delete feedback sent to your role!'})
    
    # Delete screenshot if exists
    if feedback.screenshot_filename:
        screenshot_path = os.path.join(app.config['UPLOAD_FOLDER'], feedback.screenshot_filename)
        if os.path.exists(screenshot_path):
            os.remove(screenshot_path)
    
    # Create audit log before deleting
    create_audit_log('delete_feedback', 'feedback', feedback_id, {
        'subject': feedback.subject,
        'category': feedback.category
    })
    
    db.session.delete(feedback)
    db.session.commit()
    
    return jsonify({'success': True, 'message': 'Feedback deleted successfully!'})


# ============================================================================
# ADMIN ROUTES
# ============================================================================

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get statistics (only count faculty members as users)
    total_users = User.query.filter_by(role='faculty').count()
    total_publications = Publication.query.count()
    total_departments = Department.query.count()
    pending_notifications = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    
    # Get pending second edit requests count
    pending_second_edits = EditRequest.query.filter_by(
        status='pending_admin',
        edit_level=2
    ).count()
    
    # Recent users
    recent_users = User.query.order_by(User.created_at.desc()).limit(10).all()
    
    # Recent audit logs
    recent_logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(20).all()
    
    return render_template('admin_dashboard.html',
                         total_users=total_users,
                         total_publications=total_publications,
                         total_departments=total_departments,
                         pending_notifications=pending_notifications,
                         pending_second_edits=pending_second_edits,
                         recent_users=recent_users,
                         recent_logs=recent_logs)


@app.route('/admin/users')
@login_required
def admin_users():
    if current_user.role not in ['admin', 'dean']:
        flash('Access denied. Admin or Dean privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    users = User.query.order_by(User.created_at.desc()).all()
    departments = Department.query.all()
    
    return render_template('admin_users.html', users=users, departments=departments)


@app.route('/admin/users/bulk-import', methods=['GET', 'POST'])
@login_required
def admin_bulk_import_users():
    """Bulk import users from Excel file"""
    if current_user.role not in ['admin']:
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('No file uploaded!', 'danger')
            return redirect(url_for('admin_bulk_import_users'))
        
        file = request.files['excel_file']
        
        if file.filename == '':
            flash('No file selected!', 'danger')
            return redirect(url_for('admin_bulk_import_users'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'danger')
            return redirect(url_for('admin_bulk_import_users'))
        
        try:
            import pandas as pd
            from email_utils import send_welcome_email
            
            # Read Excel file
            df = pd.read_excel(file)
            
            success_count = 0
            error_count = 0
            email_sent_count = 0
            errors = []
            
            required_columns = ['name', 'email', 'password', 'role']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                flash(f'Missing required columns: {", ".join(missing_columns)}', 'danger')
                return redirect(url_for('admin_bulk_import_users'))
            
            for index, row in df.iterrows():
                try:
                    name = str(row['name']).strip()
                    email = str(row['email']).strip().lower()
                    password = str(row['password']).strip()
                    role = str(row['role']).strip().lower()
                    department_name = str(row.get('department', '')).strip()
                    
                    # Validate required fields
                    if not all([name, email, password, role]):
                        errors.append(f"Row {index + 2}: Missing required fields")
                        error_count += 1
                        continue
                    
                    # Validate role
                    valid_roles = ['faculty', 'hod', 'dean', 'principal', 'vice_principal', 'admin', 'director', 'dean_secretary']
                    if role not in valid_roles:
                        errors.append(f"Row {index + 2}: Invalid role '{role}'. Valid roles: {', '.join(valid_roles)}")
                        error_count += 1
                        continue
                    
                    # Check if user already exists
                    if User.query.filter_by(email=email).first():
                        errors.append(f"Row {index + 2}: User with email {email} already exists")
                        error_count += 1
                        continue
                    
                    # Get or find department
                    department_id = None
                    if department_name:
                        department = Department.query.filter_by(name=department_name).first()
                        if not department:
                            # Try case-insensitive search
                            department = Department.query.filter(
                                Department.name.ilike(f'%{department_name}%')
                            ).first()
                        
                        if department:
                            department_id = department.id
                        else:
                            errors.append(f"Row {index + 2}: Department '{department_name}' not found")
                            error_count += 1
                            continue
                    
                    # Validate department requirement for roles
                    if role in ['faculty', 'hod'] and not department_id:
                        errors.append(f"Row {index + 2}: Department is required for {role.upper()} role")
                        error_count += 1
                        continue
                    
                    # Force no department for admin roles
                    if role in ['principal', 'vice_principal', 'dean', 'admin', 'director', 'dean_secretary']:
                        department_id = None
                    
                    # Create user
                    user = User(
                        name=name,
                        email=email,
                        role=role,
                        department_id=department_id
                    )
                    user.set_password(password)
                    
                    db.session.add(user)
                    db.session.flush()  # Get user ID
                    
                    # Send welcome email (skip validation)
                    email_result = send_welcome_email(user, skip_validation=True)
                    if email_result['success']:
                        email_sent_count += 1
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")
                    error_count += 1
                    continue
            
            # Commit all successful imports
            if success_count > 0:
                db.session.commit()
                
                # Create audit log
                create_audit_log(
                    action='bulk_user_import',
                    target_type='user',
                    target_id=None,
                    details=json.dumps({
                        'success_count': success_count,
                        'error_count': error_count,
                        'email_sent_count': email_sent_count
                    })
                )
            
            # Show results
            if success_count > 0:
                flash(f'âœ… Successfully imported {success_count} users. ðŸ“§ {email_sent_count} welcome emails sent.', 'success')
            
            if error_count > 0:
                flash(f'âŒ {error_count} errors occurred during import.', 'warning')
                for error in errors[:10]:  # Show first 10 errors
                    flash(error, 'danger')
                if len(errors) > 10:
                    flash(f'... and {len(errors) - 10} more errors', 'danger')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error processing file: {str(e)}', 'danger')
        
        return redirect(url_for('admin_bulk_import_users'))
    
    # GET request - show template
    departments = Department.query.all()
    return render_template('admin_bulk_import_users.html', departments=departments)


@app.route('/admin/users/create', methods=['POST'])
@login_required
def admin_create_user():
    if current_user.role not in ['admin', 'dean']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        name = request.form.get('name')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role')
        department_id = request.form.get('department_id')
        
        # Validation
        if not all([name, email, password, role]):
            flash('All fields are required.', 'danger')
            return redirect(url_for('admin_users'))
        
        # Validate email domain (must be @sjec.ac.in, allow dummy for admin)
        from email_utils import validate_email_exists
        email_validation = validate_email_exists(email, allow_sjec_dummy=True)
        if not email_validation['valid']:
            flash(f'Invalid email: {email_validation["message"]}', 'danger')
            return redirect(url_for('admin_users'))
        
        is_dummy_email = email_validation.get('is_dummy', False)
        
        # Department validation based on role
        if role in ['faculty', 'hod']:
            if not department_id:
                flash(f'Department is required for {role.upper()} role.', 'danger')
                return redirect(url_for('admin_users'))
        
        # Principal, VP, Dean, and Admin should not have department
        if role in ['principal', 'vice_principal', 'dean', 'admin']:
            department_id = None  # Force N/A for these roles
        
        # Check if email already exists
        if User.query.filter_by(email=email).first():
            flash('Email already exists.', 'danger')
            return redirect(url_for('admin_users'))
        
        # Create user
        user = User(
            name=name,
            email=email,
            role=role,
            department_id=int(department_id) if department_id else None
        )
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        # Send welcome email (skip validation for admin-created users)
        # But don't send if it's a dummy email
        from email_utils import send_welcome_email
        email_result = send_welcome_email(user, skip_validation=True)
        
        email_status_msg = ''
        if email_result.get('skipped'):
            email_status_msg = '(Dummy email - notifications skipped)'
        elif not email_result['success']:
            email_status_msg = '(Email notification could not be sent)'
            print(f"Warning: Could not send welcome email to {user.email}: {email_result['message']}")
        else:
            email_status_msg = f'Welcome email sent to {email}'
        
        # Create audit log
        create_audit_log(
            action='user_created',
            target_type='user',
            target_id=user.id,
            details=json.dumps({
                'created_user_email': email,
                'created_user_role': role,
                'is_dummy_email': is_dummy_email,
                'email_sent': email_result['success'] and not email_result.get('skipped', False)
            })
        )
        
        flash(f'User {name} created successfully! {email_status_msg}', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating user: {str(e)}', 'danger')
    
    return redirect(url_for('admin_users'))


@app.route('/admin/users/reset_password/<int:user_id>', methods=['POST'])
@login_required
def admin_reset_password(user_id):
    if current_user.role not in ['admin', 'dean']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        user = User.query.get_or_404(user_id)
        new_password = request.form.get('new_password')
        
        if not new_password:
            flash('Password cannot be empty.', 'danger')
            return redirect(url_for('admin_users'))
        
        user.set_password(new_password)
        db.session.commit()
        
        # Create audit log
        create_audit_log(
            action='password_reset',
            target_type='user',
            target_id=user.id,
            details=json.dumps({
                'reset_for_user': user.email
            })
        )
        
        # Create notification for the user
        create_notification(
            user_id=user.id,
            title='Password Reset',
            message=f'Your password has been reset by the admin.',
            notification_type='warning'
        )
        
        flash(f'Password reset successfully for {user.name}.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error resetting password: {str(e)}', 'danger')
    
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/disable', methods=['POST'])
@login_required
def admin_disable_user(user_id):
    if current_user.role not in ['admin', 'dean']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        user = User.query.get_or_404(user_id)
        
        # Prevent disabling yourself
        if user.id == current_user.id:
            return jsonify({'success': False, 'error': 'Cannot disable your own account'}), 400
        
        user.is_active = False
        db.session.commit()
        
        # Send email notification to user
        create_notification(
            user_id=user.id,
            title='Account Disabled',
            message=f'Your account has been disabled by {current_user.role}. Please contact administration for more information.',
            notification_type='warning'
        )
        
        create_audit_log(
            action='user_disabled',
            target_type='user',
            target_id=user.id,
            details=json.dumps({'disabled_user': user.email, 'disabled_by': current_user.email})
        )
        
        return jsonify({'success': True, 'message': f'User {user.name} disabled successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/admin/users/<int:user_id>/enable', methods=['POST'])
@login_required
def admin_enable_user(user_id):
    if current_user.role not in ['admin', 'dean']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        user = User.query.get_or_404(user_id)
        user.is_active = True
        db.session.commit()
        
        # Send email notification to user
        create_notification(
            user_id=user.id,
            title='Account Enabled',
            message=f'Your account has been enabled by {current_user.role}. You can now access the portal.',
            notification_type='success'
        )
        
        create_audit_log(
            action='user_enabled',
            target_type='user',
            target_id=user.id,
            details=json.dumps({'enabled_user': user.email, 'enabled_by': current_user.email})
        )
        
        return jsonify({'success': True, 'message': f'User {user.name} enabled successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/admin/users/delete/<int:user_id>', methods=['POST'])
@login_required
def admin_delete_user(user_id):
    if current_user.role not in ['admin', 'dean']:
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    import secrets
    
    try:
        user = User.query.get_or_404(user_id)
        
        # Prevent deleting yourself
        if user.id == current_user.id:
            flash('You cannot delete your own account.', 'danger')
            return redirect(url_for('admin_users'))
        
        user_email = user.email
        user_name = user.name
        
        # Check if publications should be deleted
        delete_publications = request.form.get('delete_publications', 'true') == 'true'
        
        # Get count of user's publications
        publications = Publication.query.filter_by(user_id=user.id).all()
        pub_count = len(publications)
        
        # Delete user's profile picture and signature files
        if user.profile_picture:
            profile_path = os.path.join(app.config['PROFILES_FOLDER'], user.profile_picture)
            if os.path.exists(profile_path):
                try:
                    os.remove(profile_path)
                except Exception as e:
                    print(f"Error deleting profile picture: {str(e)}")
        
        if user.signature:
            signature_path = os.path.join(app.config['SIGNATURES_FOLDER'], user.signature)
            if os.path.exists(signature_path):
                try:
                    os.remove(signature_path)
                except Exception as e:
                    print(f"Error deleting signature: {str(e)}")
        
        # Delete publications based on option
        deleted_pdfs = 0
        if delete_publications:
            # Option 1: Delete all publications and their PDF files
            for pub in publications:
                # Delete PDF file if exists
                if pub.pdf_filename:
                    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], 'publications', pub.pdf_filename)
                    if os.path.exists(pdf_path):
                        try:
                            os.remove(pdf_path)
                            deleted_pdfs += 1
                        except Exception as e:
                            print(f"Error deleting PDF {pub.pdf_filename}: {str(e)}")
                
                # Delete publication from database
                db.session.delete(pub)
        else:
            # Option 2: Keep publications - Create/get a system user for this specific deleted user
            # Use the deleted user's name to create a unique system account
            system_email = f'deleted.{user_email.split("@")[0]}@sjec.ac.in'
            deleted_user_account = User.query.filter_by(email=system_email).first()
            
            if not deleted_user_account:
                # Create system user with the original user's name for tracking
                deleted_user_account = User(
                    name=f'[Deleted] {user_name}',
                    email=system_email,
                    role='faculty'
                )
                deleted_user_account.set_password('deleted_user_account_' + secrets.token_hex(16))
                db.session.add(deleted_user_account)
                db.session.flush()  # Get the user ID
            
            # Transfer all publications to the system user (preserving original author attribution)
            for pub in publications:
                pub.user_id = deleted_user_account.id
        
        # Delete related records before deleting user
        # 1. Delete notifications sent to this user
        notif_count = Notification.query.filter_by(user_id=user.id).delete()
        
        # 2. Delete edit requests made by this user
        edit_req_count = EditRequest.query.filter_by(faculty_id=user.id).delete()
        
        # 3. Delete edit requests assigned to this user (if HoD)
        EditRequest.query.filter_by(hod_id=user.id).delete()
        
        # 4. Delete feedback sent by this user
        feedback_count = Feedback.query.filter_by(user_id=user.id).delete()
        
        # 5. Delete feedback resolved by this user (set resolver to NULL)
        Feedback.query.filter_by(resolved_by=user.id).update({'resolved_by': None})
        
        # 6. Delete audit logs (optional - or keep them with user_id as None)
        audit_count = AuditLog.query.filter_by(user_id=user.id).delete()
        
        # Create audit log before deletion
        create_audit_log(
            action='user_deleted_cascade',
            target_type='user',
            target_id=user.id,
            details=json.dumps({
                'deleted_user_email': user_email,
                'deleted_user_name': user_name,
                'deleted_user_role': user.role,
                'deleted_publications': pub_count,
                'deleted_pdfs': deleted_pdfs,
                'deleted_notifications': notif_count,
                'deleted_edit_requests': edit_req_count,
                'deleted_feedback': feedback_count,
                'deleted_audit_logs': audit_count
            })
        )
        
        # Now safe to delete the user
        db.session.delete(user)
        db.session.commit()
        
        # Create detailed success message
        message_parts = [f'User {user_name} ({user_email}) deleted successfully.']
        
        if delete_publications:
            if pub_count > 0:
                message_parts.append(f'âš ï¸ Deleted {pub_count} publication(s) and {deleted_pdfs} PDF file(s).')
        else:
            if pub_count > 0:
                message_parts.append(f'âœ“ Preserved {pub_count} publication(s) (attributed to "[Deleted] {user_name}").')
        
        if notif_count > 0:
            message_parts.append(f'Deleted {notif_count} notification(s).')
        if edit_req_count > 0:
            message_parts.append(f'Deleted {edit_req_count} edit request(s).')
        if feedback_count > 0:
            message_parts.append(f'Deleted {feedback_count} feedback(s).')
        
        flash(' '.join(message_parts), 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting user: {str(e)}', 'danger')
    
    return redirect(url_for('admin_users'))


@app.route('/admin/audit_logs')
@login_required
def admin_audit_logs():
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get filter parameters
    action_filter = request.args.get('action', '')
    user_filter = request.args.get('user_id', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query
    query = AuditLog.query
    
    if action_filter:
        query = query.filter(AuditLog.action == action_filter)
    
    if user_filter:
        query = query.filter(AuditLog.user_id == int(user_filter))
    
    if date_from:
        query = query.filter(AuditLog.created_at >= datetime.strptime(date_from, '%Y-%m-%d'))
    
    if date_to:
        query = query.filter(AuditLog.created_at <= datetime.strptime(date_to + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
    
    logs = query.order_by(AuditLog.created_at.desc()).limit(500).all()
    
    # Get unique actions for filter dropdown
    actions = db.session.query(AuditLog.action.distinct()).all()
    actions = [a[0] for a in actions]
    
    # Get all users for filter dropdown
    users = User.query.order_by(User.name).all()
    
    return render_template('admin_audit_logs.html',
                         logs=logs,
                         actions=actions,
                         users=users,
                         action_filter=action_filter,
                         user_filter=user_filter,
                         date_from=date_from,
                         date_to=date_to)


@app.route('/admin/audit_logs/clear', methods=['POST'])
@login_required
def admin_clear_audit_logs():
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    clear_from_date = request.form.get('clear_from_date')
    clear_to_date = request.form.get('clear_to_date')
    
    try:
        query = AuditLog.query
        
        if clear_from_date:
            from_date = datetime.strptime(clear_from_date, '%Y-%m-%d')
            query = query.filter(AuditLog.created_at >= from_date)
        
        if clear_to_date:
            to_date = datetime.strptime(clear_to_date, '%Y-%m-%d')
            to_date = to_date.replace(hour=23, minute=59, second=59)
            query = query.filter(AuditLog.created_at <= to_date)
        
        deleted_count = query.delete()
        db.session.commit()
        
        # Log the clearing action
        create_audit_log(
            action='clear_audit_logs',
            target_type='audit_log',
            target_id=None,
            details={
                'deleted_count': deleted_count,
                'from_date': clear_from_date or 'beginning',
                'to_date': clear_to_date or 'today'
            }
        )
        
        flash(f'Successfully cleared {deleted_count} audit log(s).', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error clearing audit logs: {str(e)}', 'danger')
    
    return redirect(url_for('admin_audit_logs'))


@app.route('/admin/departments')
@login_required
def admin_departments():
    """View and manage departments"""
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    departments = Department.query.order_by(Department.name).all()
    
    # Get stats for each department
    dept_stats = []
    for dept in departments:
        # Count faculty and HOD, excluding leadership roles (principal, dean, director, dean_secretary)
        faculty_count = User.query.filter(
            User.department_id == dept.id,
            User.role.in_(['faculty', 'hod'])
        ).count()
        hod = User.query.filter_by(department_id=dept.id, role='hod').first()
        publication_count = Publication.query.filter_by(department_id=dept.id).count()
        
        dept_stats.append({
            'department': dept,
            'faculty_count': faculty_count,
            'hod': hod,
            'publication_count': publication_count
        })
    
    return render_template('admin_departments.html', dept_stats=dept_stats)


@app.route('/admin/departments/create', methods=['POST'])
@login_required
def admin_create_department():
    """Create a new department"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        name = request.form.get('name', '').strip()
        code = request.form.get('code', '').strip().upper()
        
        if not name or not code:
            flash('Department name and code are required.', 'danger')
            return redirect(url_for('admin_departments'))
        
        # Check if department already exists
        if Department.query.filter_by(name=name).first():
            flash(f'Department "{name}" already exists.', 'danger')
            return redirect(url_for('admin_departments'))
        
        if Department.query.filter_by(code=code).first():
            flash(f'Department code "{code}" already exists.', 'danger')
            return redirect(url_for('admin_departments'))
        
        # Create department
        department = Department(name=name, code=code)
        db.session.add(department)
        db.session.commit()
        
        # Create audit log
        create_audit_log(
            action='department_created',
            target_type='department',
            target_id=department.id,
            details={
                'name': name,
                'code': code
            }
        )
        
        flash(f'Department "{name}" ({code}) created successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating department: {str(e)}', 'danger')
    
    return redirect(url_for('admin_departments'))


@app.route('/admin/departments/<int:dept_id>/edit', methods=['POST'])
@login_required
def admin_edit_department(dept_id):
    """Edit department details"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        department = Department.query.get_or_404(dept_id)
        
        name = request.form.get('name', '').strip()
        code = request.form.get('code', '').strip().upper()
        
        if not name or not code:
            flash('Department name and code are required.', 'danger')
            return redirect(url_for('admin_departments'))
        
        # Check for duplicates (excluding current department)
        existing_name = Department.query.filter(Department.name == name, Department.id != dept_id).first()
        if existing_name:
            flash(f'Department name "{name}" already exists.', 'danger')
            return redirect(url_for('admin_departments'))
        
        existing_code = Department.query.filter(Department.code == code, Department.id != dept_id).first()
        if existing_code:
            flash(f'Department code "{code}" already exists.', 'danger')
            return redirect(url_for('admin_departments'))
        
        old_name = department.name
        old_code = department.code
        
        department.name = name
        department.code = code
        db.session.commit()
        
        # Create audit log
        create_audit_log(
            action='department_updated',
            target_type='department',
            target_id=department.id,
            details={
                'old_name': old_name,
                'new_name': name,
                'old_code': old_code,
                'new_code': code
            }
        )
        
        flash(f'Department updated successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating department: {str(e)}', 'danger')
    
    return redirect(url_for('admin_departments'))


@app.route('/admin/departments/<int:dept_id>/delete', methods=['POST'])
@login_required
def admin_delete_department(dept_id):
    """Delete a department (only if no users or publications)"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'error': 'Unauthorized'}), 403
    
    try:
        department = Department.query.get_or_404(dept_id)
        
        # Check if department has users
        user_count = User.query.filter_by(department_id=dept_id).count()
        if user_count > 0:
            flash(f'Cannot delete department. It has {user_count} user(s) assigned.', 'danger')
            return redirect(url_for('admin_departments'))
        
        # Check if department has publications
        pub_count = Publication.query.filter_by(department_id=dept_id).count()
        if pub_count > 0:
            flash(f'Cannot delete department. It has {pub_count} publication(s).', 'danger')
            return redirect(url_for('admin_departments'))
        
        dept_name = department.name
        dept_code = department.code
        
        # Create audit log before deletion
        create_audit_log(
            action='department_deleted',
            target_type='department',
            target_id=dept_id,
            details={
                'name': dept_name,
                'code': dept_code
            }
        )
        
        db.session.delete(department)
        db.session.commit()
        
        flash(f'Department "{dept_name}" deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting department: {str(e)}', 'danger')
    
    return redirect(url_for('admin_departments'))


# Admin: View all publications with delete capability
@app.route('/admin/publications')
@login_required
def admin_publications():
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get filter parameters
    dept_id = request.args.get('department', type=int)
    year = request.args.get('year', type=int)
    status = request.args.get('status', '')
    
    # Build query
    query = Publication.query
    
    if dept_id:
        query = query.filter_by(department_id=dept_id)
    if year:
        query = query.filter_by(year=year)
    if status:
        query = query.filter_by(status=status)
    
    publications = query.order_by(Publication.created_at.desc()).all()
    departments = Department.query.order_by(Department.name).all()
    
    # Get unique years
    years = db.session.query(Publication.year).distinct().order_by(Publication.year.desc()).all()
    years = [y[0] for y in years]
    
    return render_template('admin_publications.html',
                         publications=publications,
                         departments=departments,
                         years=years,
                         selected_dept=dept_id,
                         selected_year=year,
                         selected_status=status)


# Admin: Delete publication with reason
@app.route('/admin/publications/<int:pub_id>/delete', methods=['POST'])
@login_required
def admin_delete_publication(pub_id):
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    publication = Publication.query.get_or_404(pub_id)
    reason = request.form.get('reason', '').strip()
    
    if not reason:
        flash('Please provide a reason for deleting this publication.', 'warning')
        return redirect(url_for('admin_publications'))
    
    # Store details for notification
    pub_title = publication.title
    pub_owner_id = publication.user_id
    faculty_name = publication.author.name
    
    # Delete PDF file if exists
    if publication.pdf_filename:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], publication.pdf_filename)
        if os.path.exists(file_path):
            os.remove(file_path)
    
    # Notify the publication owner
    create_notification(
        user_id=pub_owner_id,
        title='âš ï¸ Publication Deleted by Admin',
        message=f'Your publication "{pub_title[:100]}" has been deleted by Admin.\n\nReason: {reason}',
        notification_type='danger'
    )
    
    # Create audit log
    create_audit_log('admin_delete_publication', 'publication', pub_id, {
        'title': pub_title,
        'faculty_id': pub_owner_id,
        'faculty_name': faculty_name,
        'reason': reason
    })
    
    # Delete the publication
    db.session.delete(publication)
    db.session.commit()
    
    flash(f'Publication "{pub_title[:50]}..." has been deleted. Faculty has been notified.', 'success')
    return redirect(url_for('admin_publications'))


# Admin: View pending second edit requests
@app.route('/admin/edit_requests')
@login_required
def admin_edit_requests():
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get pending admin approval requests (second edit)
    pending_requests = EditRequest.query.filter_by(
        status='pending_admin',
        edit_level=2
    ).order_by(EditRequest.requested_at.desc()).all()
    
    # Get all second level edit requests for history
    all_second_requests = EditRequest.query.filter_by(
        edit_level=2
    ).order_by(EditRequest.requested_at.desc()).all()
    
    return render_template('admin_edit_requests.html',
                         pending_requests=pending_requests,
                         all_requests=all_second_requests)


# Admin: Approve second edit request
@app.route('/admin/edit_requests/<int:request_id>/approve', methods=['POST'])
@login_required
def admin_approve_edit_request(request_id):
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    edit_request = EditRequest.query.get_or_404(request_id)
    
    # Verify this is a second edit request pending admin approval
    if edit_request.status != 'pending_admin' or edit_request.edit_level != 2:
        flash('This request is not pending admin approval!', 'warning')
        return redirect(url_for('admin_edit_requests'))
    
    # Approve the second edit request
    edit_request.status = 'admin_approved'
    edit_request.admin_responded_at = now_ist()
    edit_request.admin_response = request.form.get('response', '').strip()
    
    # Grant edit permission
    publication = edit_request.publication
    publication.can_edit = True
    publication.edit_granted_by = current_user.id
    publication.edit_granted_at = now_ist()
    
    db.session.commit()
    
    # Notify faculty
    admin_comment = edit_request.admin_response if edit_request.admin_response else 'No additional comment provided.'
    create_notification(
        user_id=edit_request.faculty_id,
        title='âœ… Second Edit Request Approved by Admin',
        message=f'Your second edit request for "{publication.title[:50]}..." has been approved by Admin. You can now edit this publication once.\n\nAdmin Response: {admin_comment}',
        notification_type='success',
        publication_id=publication.id
    )
    
    # Notify HoD as well
    create_notification(
        user_id=edit_request.hod_id,
        title='Second Edit Request Approved by Admin',
        message=f'Admin has approved the second edit request for "{publication.title[:50]}..." by {edit_request.faculty.name}.',
        notification_type='info',
        publication_id=publication.id
    )
    
    # Create audit log
    create_audit_log('admin_approve_second_edit', 'publication', publication.id, {
        'request_id': request_id,
        'faculty_id': edit_request.faculty_id,
        'edit_level': 2
    })
    
    flash('Second edit request approved successfully!', 'success')
    return redirect(url_for('admin_edit_requests'))


# Admin: Deny second edit request
@app.route('/admin/edit_requests/<int:request_id>/deny', methods=['POST'])
@login_required
def admin_deny_edit_request(request_id):
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    edit_request = EditRequest.query.get_or_404(request_id)
    
    # Verify this is a second edit request pending admin approval
    if edit_request.status != 'pending_admin' or edit_request.edit_level != 2:
        flash('This request is not pending admin approval!', 'warning')
        return redirect(url_for('admin_edit_requests'))
    
    # Deny the second edit request
    edit_request.status = 'admin_denied'
    edit_request.admin_responded_at = now_ist()
    edit_request.admin_response = request.form.get('response', '').strip()
    
    db.session.commit()
    
    # Notify faculty
    admin_comment = edit_request.admin_response if edit_request.admin_response else 'No reason provided.'
    create_notification(
        user_id=edit_request.faculty_id,
        title='âŒ Second Edit Request Denied by Admin',
        message=f'Your second edit request for "{edit_request.publication.title[:50]}..." has been denied by Admin.\n\nAdmin Response: {admin_comment}',
        notification_type='danger',
        publication_id=edit_request.publication_id
    )
    
    # Notify HoD as well
    create_notification(
        user_id=edit_request.hod_id,
        title='Second Edit Request Denied by Admin',
        message=f'Admin has denied the second edit request for "{edit_request.publication.title[:50]}..." by {edit_request.faculty.name}.',
        notification_type='warning',
        publication_id=edit_request.publication_id
    )
    
    # Create audit log
    create_audit_log('admin_deny_second_edit', 'publication', edit_request.publication_id, {
        'request_id': request_id,
        'faculty_id': edit_request.faculty_id,
        'edit_level': 2
    })
    
    flash('Second edit request denied.', 'info')
    return redirect(url_for('admin_edit_requests'))


@app.route('/admin/database_management')
@login_required
def admin_database_management():
    """Database management and backup interface for admins"""
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    from db_optimizer import DatabaseOptimizer, DatabaseBackup
    
    # Get database statistics
    optimizer = DatabaseOptimizer()
    stats = optimizer.get_database_stats()
    
    # Get backup list
    backup_manager = DatabaseBackup()
    backups = backup_manager.list_backups()
    
    # Get database file size
    db_path = Path('instance/faculty_publications.db')
    db_size_mb = 0
    if db_path.exists():
        db_size_mb = db_path.stat().st_size / (1024 * 1024)
    
    return render_template('admin_database_management.html', 
                         stats=stats, 
                         backups=backups,
                         db_size_mb=db_size_mb)


@app.route('/admin/database_optimize', methods=['POST'])
@login_required
def admin_database_optimize():
    """Run database optimization"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    try:
        from db_optimizer import DatabaseOptimizer
        
        optimizer = DatabaseOptimizer()
        
        # Create indexes
        optimizer.create_indexes()
        
        # Analyze database
        optimizer.analyze_database()
        
        # Create audit log
        create_audit_log('database_optimization', 'system', None, {
            'performed_by': current_user.email
        })
        
        flash('âœ… Database optimized successfully! Indexes created and statistics updated.', 'success')
        return redirect(url_for('admin_database_management'))
        
    except Exception as e:
        flash(f'Error optimizing database: {str(e)}', 'danger')
        return redirect(url_for('admin_database_management'))


@app.route('/admin/database_backup', methods=['POST'])
@login_required
def admin_database_backup():
    """Create database backup"""
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    try:
        from db_optimizer import DatabaseBackup
        
        backup_manager = DatabaseBackup()
        backup_path = backup_manager.create_backup()
        
        if backup_path:
            # Create audit log
            create_audit_log('database_backup', 'system', None, {
                'backup_file': backup_path.name,
                'performed_by': current_user.email
            })
            
            flash(f'âœ… Database backup created successfully: {backup_path.name}', 'success')
        else:
            flash('âŒ Failed to create database backup.', 'danger')
        
        return redirect(url_for('admin_database_management'))
        
    except Exception as e:
        flash(f'Error creating backup: {str(e)}', 'danger')
        return redirect(url_for('admin_database_management'))


@app.route('/admin/database_download_backup/<filename>')
@login_required
def admin_download_backup(filename):
    """Download a backup file"""
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    try:
        from flask import send_from_directory
        backup_dir = Path('backups')
        
        # Security check - ensure filename doesn't contain path traversal
        if '..' in filename or '/' in filename or '\\' in filename:
            flash('Invalid filename.', 'danger')
            return redirect(url_for('admin_database_management'))
        
        return send_from_directory(backup_dir, filename, as_attachment=True)
        
    except Exception as e:
        flash(f'Error downloading backup: {str(e)}', 'danger')
        return redirect(url_for('admin_database_management'))


@app.route('/admin/email_settings', methods=['GET', 'POST'])
@login_required
def admin_email_settings():
    if current_user.role != 'admin':
        flash('Access denied. Admin privileges required.', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        # Get form data
        mail_server = request.form.get('mail_server', '').strip()
        mail_port = request.form.get('mail_port', '').strip()
        mail_username = request.form.get('mail_username', '').strip()
        mail_password = request.form.get('mail_password', '').strip()
        mail_default_sender = request.form.get('mail_default_sender', '').strip()
        mail_use_tls = request.form.get('mail_use_tls') == 'on'
        
        # Validate required fields
        if not all([mail_server, mail_port, mail_username, mail_password, mail_default_sender]):
            flash('All fields are required!', 'danger')
            return redirect(url_for('admin_email_settings'))
        
        try:
            # Read existing .env file
            env_path = os.path.join(os.path.dirname(__file__), '.env')
            env_lines = []
            
            if os.path.exists(env_path):
                with open(env_path, 'r') as f:
                    env_lines = f.readlines()
            
            # Update or add email settings
            email_keys = {
                'MAIL_SERVER': mail_server,
                'MAIL_PORT': mail_port,
                'MAIL_USE_TLS': str(mail_use_tls),
                'MAIL_USERNAME': mail_username,
                'MAIL_PASSWORD': mail_password,
                'MAIL_DEFAULT_SENDER': mail_default_sender
            }
            
            new_env_lines = []
            updated_keys = set()
            
            # Update existing keys
            for line in env_lines:
                updated = False
                for key, value in email_keys.items():
                    if line.startswith(f'{key}='):
                        new_env_lines.append(f'{key}={value}\n')
                        updated_keys.add(key)
                        updated = True
                        break
                if not updated:
                    new_env_lines.append(line)
            
            # Add missing keys
            for key, value in email_keys.items():
                if key not in updated_keys:
                    new_env_lines.append(f'{key}={value}\n')
            
            # Write back to .env file
            with open(env_path, 'w') as f:
                f.writelines(new_env_lines)
            
            # Update current app config
            app.config['MAIL_SERVER'] = mail_server
            app.config['MAIL_PORT'] = int(mail_port)
            app.config['MAIL_USE_TLS'] = mail_use_tls
            app.config['MAIL_USERNAME'] = mail_username
            app.config['MAIL_PASSWORD'] = mail_password
            app.config['MAIL_DEFAULT_SENDER'] = mail_default_sender
            
            # Create audit log
            create_audit_log('update_email_settings', 'system', None, {
                'mail_server': mail_server,
                'mail_username': mail_username
            })
            
            flash('âœ… Email settings updated successfully! Changes will take effect on next restart.', 'success')
            return redirect(url_for('admin_email_settings'))
            
        except Exception as e:
            flash(f'Error updating email settings: {str(e)}', 'danger')
            return redirect(url_for('admin_email_settings'))
    
    # GET request - display current settings
    current_settings = {
        'mail_server': app.config.get('MAIL_SERVER', 'smtp.gmail.com'),
        'mail_port': app.config.get('MAIL_PORT', 587),
        'mail_use_tls': app.config.get('MAIL_USE_TLS', True),
        'mail_username': app.config.get('MAIL_USERNAME', ''),
        'mail_password': app.config.get('MAIL_PASSWORD', ''),
        'mail_default_sender': app.config.get('MAIL_DEFAULT_SENDER', 'noreply@sjec.ac.in')
    }
    
    return render_template('admin-email-settings.html', settings=current_settings)


@app.route('/admin/test_email', methods=['POST'])
@login_required
def admin_test_email():
    if current_user.role != 'admin':
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    test_email = request.form.get('test_email', '').strip()
    
    if not test_email:
        return jsonify({'success': False, 'message': 'Email address required'}), 400
    
    try:
        from email_utils import send_email
        
        subject = "Test Email from SJEC Publication Portal"
        body_html = """
        <!DOCTYPE html>
        <html>
        <head>
            <style>
                body { font-family: Arial, sans-serif; line-height: 1.6; color: #333; }
                .container { max-width: 600px; margin: 0 auto; padding: 20px; }
                .header { background-color: #0d47a1; color: white; padding: 20px; text-align: center; }
                .content { background-color: white; padding: 30px; margin-top: 20px; border: 1px solid #ddd; }
            </style>
        </head>
        <body>
            <div class="container">
                <div class="header">
                    <h2>âœ… Test Email</h2>
                </div>
                <div class="content">
                    <h3>Email Configuration Test</h3>
                    <p>This is a test email from SJEC Faculty Publication Portal.</p>
                    <p>If you received this email, your email settings are configured correctly!</p>
                    <p><strong>Time:</strong> """ + datetime.now(IST).strftime('%Y-%m-%d %H:%M:%S IST') + """</p>
                </div>
            </div>
        </body>
        </html>
        """
        
        result = send_email(test_email, subject, body_html, skip_validation=True)
        
        if result['success']:
            return jsonify({'success': True, 'message': f'Test email sent to {test_email}!'})
        else:
            return jsonify({'success': False, 'message': result['message']}), 500
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@app.route('/admin/bulk_import', methods=['GET', 'POST'])
@login_required
def admin_bulk_import():
    # Allow faculty, admin, and hod
    if current_user.role not in ['admin', 'hod', 'faculty']:
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            flash('No file uploaded!', 'danger')
            return redirect(url_for('admin_bulk_import'))
        
        file = request.files['excel_file']
        
        if file.filename == '':
            flash('No file selected!', 'danger')
            return redirect(url_for('admin_bulk_import'))
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload an Excel file (.xlsx or .xls)', 'danger')
            return redirect(url_for('admin_bulk_import'))
        
        try:
            # Read Excel file
            df = pd.read_excel(file)
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Get default department (first one) for publications without department mapping
            default_dept = Department.query.first()
            
            for index, row in df.iterrows():
                try:
                    # Find user by email (required column)
                    user_email = row.get('faculty_email', '').strip()
                    if not user_email:
                        errors.append(f"Row {index + 2}: Missing faculty_email")
                        error_count += 1
                        continue
                    
                    user = User.query.filter_by(email=user_email).first()
                    if not user:
                        errors.append(f"Row {index + 2}: User with email {user_email} not found")
                        error_count += 1
                        continue
                    
                    # If faculty, only allow imports for themselves
                    if current_user.role == 'faculty' and user.id != current_user.id:
                        errors.append(f"Row {index + 2}: Faculty can only import publications for their own email ({current_user.email})")
                        error_count += 1
                        continue
                    
                    # If HoD, only allow imports for their own department
                    if current_user.role == 'hod' and user.department_id != current_user.department_id:
                        errors.append(f"Row {index + 2}: HoD can only import for their own department ({current_user.department.name})")
                        error_count += 1
                        continue
                    
                    # Validate and get year
                    year_value = row.get('year')
                    month_value = row.get('month')
                    
                    # Get current date in IST
                    now_ist = datetime.now(IST)
                    current_year = now_ist.year
                    current_month = now_ist.month
                    
                    if pd.notna(year_value):
                        year_value = int(year_value)
                        # Check if year is exactly 4 digits
                        if len(str(year_value)) != 4:
                            errors.append(f"Row {index + 2}: Year must be exactly 4 digits (e.g., 2024, not 24)")
                            error_count += 1
                            continue
                        if year_value < 2002:
                            errors.append(f"Row {index + 2}: Year must be 2002 or later")
                            error_count += 1
                            continue
                        if year_value > current_year:
                            errors.append(f"Row {index + 2}: Year cannot be in the future. Current year is {current_year}")
                            error_count += 1
                            continue
                        
                        # Check if month is in future for current year
                        if year_value == current_year and pd.notna(month_value):
                            month_int = int(month_value)
                            if month_int > current_month:
                                errors.append(f"Row {index + 2}: Cannot select a future month. Current month is {current_month}")
                                error_count += 1
                                continue
                    else:
                        year_value = current_year
                    
                    # Create publication
                    pub_type = row.get('publication_type', 'Journal')  # Default to Journal
                    if pub_type not in ['Journal', 'Book', 'Book Chapter']:
                        pub_type = 'Journal'  # Fallback to Journal if invalid
                    
                    pub = Publication(
                        publication_type=pub_type,
                        title=row.get('title', ''),
                        abstract=row.get('abstract', ''),
                        publisher_name=row.get('publisher_name', ''),
                        authors_names=row.get('authors_names', ''),
                        journal_conference_name=row.get('journal_conference_name', ''),
                        volume=row.get('volume', ''),
                        issue=row.get('issue', ''),
                        pages=row.get('pages', ''),
                        indexing_status=row.get('indexing_status', '') if pub_type == 'Journal' else None,
                        quartile=row.get('quartile', '') if pub_type == 'Journal' else None,
                        impact_factor=float(row.get('impact_factor', 0)) if (pd.notna(row.get('impact_factor')) and pub_type == 'Journal') else None,
                        isbn=row.get('isbn', '') if pub_type in ['Book', 'Book Chapter'] else None,
                        edition=row.get('edition', '') if pub_type in ['Book', 'Book Chapter'] else None,
                        doi=row.get('doi', ''),
                        year=year_value,
                        month=int(row.get('month', 0)) if pd.notna(row.get('month')) else None,
                        citation_count=int(row.get('citation_count', 0)) if (pd.notna(row.get('citation_count')) and pub_type == 'Journal') else None,
                        bibtex_entry=row.get('bibtex_entry', ''),
                        user_id=user.id,
                        department_id=user.department_id or default_dept.id,
                        status='confirmed',
                        confirmed_at=now_ist()
                    )
                    
                    db.session.add(pub)
                    success_count += 1
                    
                except ValueError as e:
                    # Handle data type conversion errors
                    error_msg = str(e).lower()
                    if 'year' in error_msg or 'int' in error_msg:
                        errors.append(f"Row {index + 2}: Invalid year format - must be a number (e.g., 2024)")
                    elif 'float' in error_msg:
                        errors.append(f"Row {index + 2}: Invalid impact factor - must be a decimal number")
                    else:
                        errors.append(f"Row {index + 2}: Invalid data format - {str(e)}")
                    error_count += 1
                except Exception as e:
                    # Handle database constraint errors with user-friendly messages
                    error_msg = str(e).lower()
                    if 'not null constraint' in error_msg or 'cannot be null' in error_msg:
                        # Extract field name from error
                        if 'title' in error_msg:
                            errors.append(f"Row {index + 2}: Missing required field 'title'")
                        elif 'authors_names' in error_msg:
                            errors.append(f"Row {index + 2}: Missing required field 'authors_names'")
                        elif 'year' in error_msg:
                            errors.append(f"Row {index + 2}: Missing required field 'year'")
                        else:
                            errors.append(f"Row {index + 2}: Missing required field(s). Title, authors, and year are mandatory.")
                    elif 'unique constraint' in error_msg or 'duplicate' in error_msg:
                        errors.append(f"Row {index + 2}: This publication may already exist (duplicate entry)")
                    elif 'foreign key' in error_msg:
                        errors.append(f"Row {index + 2}: Invalid reference - user or department not found")
                    else:
                        # Generic error with simplified message
                        errors.append(f"Row {index + 2}: Error - {str(e)[:100]}")
                    error_count += 1
            
            # Commit all publications
            db.session.commit()
            
            # Create audit log
            create_audit_log(
                action='bulk_import',
                target_type='publication',
                details={'success': success_count, 'errors': error_count}
            )
            
            flash(f'Bulk import complete! {success_count} publications added, {error_count} errors.', 
                  'success' if error_count == 0 else 'warning')
            
            if errors:
                # Show first 10 errors
                for error in errors[:10]:
                    flash(error, 'danger')
                if len(errors) > 10:
                    flash(f'... and {len(errors) - 10} more errors', 'danger')
            
        except pd.errors.EmptyDataError:
            flash('Error: The Excel file is empty. Please add data to the file.', 'danger')
        except pd.errors.ParserError:
            flash('Error: Unable to parse Excel file. Please check the file format.', 'danger')
        except KeyError as e:
            missing_column = str(e).strip("'")
            flash(f'Error: Missing required column "{missing_column}" in Excel file. Please check the template.', 'danger')
        except Exception as e:
            error_msg = str(e).lower()
            if 'no such file' in error_msg or 'file not found' in error_msg:
                flash('Error: File not found. Please select a valid Excel file.', 'danger')
            elif 'permission denied' in error_msg:
                flash('Error: Cannot access file. Please close it if it\'s open in Excel and try again.', 'danger')
            elif 'not a valid' in error_msg or 'corrupt' in error_msg:
                flash('Error: Invalid or corrupted Excel file. Please use a valid .xlsx or .xls file.', 'danger')
            else:
                flash(f'Error reading Excel file: Please check your file format and try again.', 'danger')
        
        return redirect(url_for('admin_bulk_import'))
    
    # GET request - show upload form
    departments = Department.query.all()
    total_users = User.query.count()
    total_publications = Publication.query.count()
    
    return render_template('admin-bulk-import.html', 
                         departments=departments,
                         total_users=total_users,
                         total_publications=total_publications)


# API Endpoints for auto-refresh functionality
@app.route('/api/check_edit_requests')
@login_required
def check_edit_requests():
    """Check if there are any updates to edit requests for current user"""
    try:
        if current_user.role == 'faculty':
            # Check for any status changes in edit requests
            has_updates = EditRequest.query.filter_by(
                faculty_id=current_user.id
            ).filter(
                EditRequest.status.in_(['approved', 'denied', 'admin_approved', 'admin_denied', 'pending_admin'])
            ).filter(
                EditRequest.responded_at != None
            ).count() > 0
            
            return jsonify({'has_updates': has_updates})
        
        elif current_user.role == 'hod':
            # Check for pending edit requests
            pending_count = EditRequest.query.join(Publication).filter(
                Publication.department_id == current_user.department_id,
                EditRequest.status == 'pending'
            ).count()
            
            return jsonify({'pending_count': pending_count})
        
        elif current_user.role == 'admin':
            # Check for pending second edit requests
            pending_count = EditRequest.query.filter_by(
                status='pending_admin',
                edit_level=2
            ).count()
            
            return jsonify({'pending_count': pending_count})
        
        return jsonify({'has_updates': False})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/check_notifications')
@login_required
def check_notifications():
    """Check for new unread notifications"""
    try:
        unread_count = Notification.query.filter_by(
            user_id=current_user.id,
            is_read=False
        ).count()
        
        return jsonify({'unread_count': unread_count})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# APPLICATION SYSTEM ROUTES (Conference, Publication Incentive, Journal Support)
# ============================================================================

@app.route('/applications/conference-registration', methods=['GET', 'POST'])
@login_required
@role_required('faculty')
def apply_conference_registration():
    """Conference Registration Application Form"""
    from models import ApplicationForm
    import secrets
    from datetime import datetime as dt
    
    if request.method == 'POST':
        action = request.form.get('action')  # 'draft' or 'submit'
        
        try:
            # Generate tracking ID with uniqueness check
            year = dt.now().year
            count = ApplicationForm.query.filter(
                ApplicationForm.application_type == 'conference_registration',
                ApplicationForm.tracking_id.like(f'CR{year}%')
            ).count() + 1
            
            # Ensure tracking_id is unique
            tracking_id = f"CR{year}{count:04d}"
            while ApplicationForm.query.filter_by(tracking_id=tracking_id).first():
                count += 1
                tracking_id = f"CR{year}{count:04d}"
            
            # Combine place and date for backward compatibility
            conference_place = request.form.get('conference_place', '')
            conference_date_str = request.form.get('conference_date', '')
            
            # Parse conference date string to date object
            conference_date_obj = None
            if conference_date_str:
                try:
                    from datetime import datetime as dt_parser
                    conference_date_obj = dt_parser.strptime(conference_date_str, '%Y-%m-%d').date()
                except:
                    conference_date_obj = None
            
            conference_place_date = f"{conference_place} - {conference_date_str}" if conference_place and conference_date_str else request.form.get('conference_place_date', '')
            
            # Create application
            application = ApplicationForm(
                tracking_id=tracking_id,
                application_type='conference_registration',
                applicant_id=current_user.id,
                applicant_name=current_user.name,
                title_of_paper=request.form.get('title_of_paper'),
                type_of_conference=request.form.get('type_of_conference'),
                conference_name=request.form.get('conference_name'),
                conference_organizer=request.form.get('conference_organizer'),
                conference_place_date=conference_place_date,
                conference_place=conference_place,
                conference_date=conference_date_obj,
                conference_fee=float(request.form.get('conference_fee', 0)),
                status='draft' if action == 'draft' else 'submitted',
                submitted_at=now_ist() if action == 'submit' else None
            )
            
            db.session.add(application)
            db.session.flush()  # Get application.id before handling files
            
            # Handle multiple file uploads
            if 'documents' in request.files:
                files = request.files.getlist('documents')
                if files and files[0].filename:  # Check if files were actually selected
                    save_application_documents(files, application.id, current_user.id)
            
            db.session.commit()
            
            if action == 'submit':
                # Notify HOD of applicant's department
                notify_hod_of_application(application)
                
                flash(f'Application submitted successfully! Tracking ID: {tracking_id}', 'success')
            else:
                flash(f'Application saved as draft. Tracking ID: {tracking_id}', 'info')
            
            return redirect(url_for('my_applications'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('apply_conference_registration.html', now=dt.now())


@app.route('/applications/publication-incentive', methods=['GET', 'POST'])
@login_required
@role_required('faculty')
def apply_publication_incentive():
    """Publication Incentive Application Form"""
    from models import ApplicationForm, ApplicationDocument
    from datetime import datetime as dt
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        try:
            # Handle multi-file document uploads
            uploaded_files = request.files.getlist('documents')
            
            # Only require documents for submit action (not for draft)
            if action == 'submit':
                if not uploaded_files or all(f.filename == '' for f in uploaded_files):
                    flash('At least one supporting document (preferably publication PDF) is required for submission!', 'danger')
                    return render_template('apply_publication_incentive.html', 
                                         now=dt.now(), 
                                         form_data=request.form)
                
                # Check if at least one PDF exists
                has_pdf = any(f.filename.lower().endswith('.pdf') for f in uploaded_files if f.filename != '')
                if not has_pdf:
                    flash('At least one PDF document is required for submission!', 'danger')
                    return render_template('apply_publication_incentive.html', 
                                         now=dt.now(), 
                                         form_data=request.form)
            
            year = dt.now().year
            count = ApplicationForm.query.filter(
                ApplicationForm.application_type == 'publication_incentive',
                ApplicationForm.tracking_id.like(f'PI{year}%')
            ).count() + 1
            
            # Ensure tracking_id is unique
            tracking_id = f"PI{year}{count:04d}"
            while ApplicationForm.query.filter_by(tracking_id=tracking_id).first():
                count += 1
                tracking_id = f"PI{year}{count:04d}"
            
            # Parse publication date - handle both journal and book date fields
            pub_date_str = request.form.get('publication_date') or request.form.get('publication_date_book')
            pub_date = dt.strptime(pub_date_str, '%Y-%m-%d').date() if pub_date_str else None
            
            # Validate publication date is not in future
            if pub_date and pub_date > dt.now().date():
                flash('Publication date cannot be in the future!', 'danger')
                return render_template('apply_publication_incentive.html', 
                                     now=dt.now(), 
                                     form_data=request.form)
            
            # Get new fields
            publication_category = request.form.get('publication_category')
            is_national_international = request.form.get('is_national_international')
            num_first_authors_sjec = int(request.form.get('num_first_authors_sjec', 0))
            num_corresponding_authors_sjec = int(request.form.get('num_corresponding_authors_sjec', 0))
            num_coauthors_sjec = int(request.form.get('num_coauthors_sjec', 0))
            
            application = ApplicationForm(
                tracking_id=tracking_id,
                application_type='publication_incentive',
                applicant_id=current_user.id,
                applicant_name=current_user.name,
                title_of_paper=request.form.get('title_of_paper'),
                
                # New fields
                publication_category=publication_category,
                is_national_international=is_national_international,
                num_first_authors_sjec=num_first_authors_sjec,
                num_corresponding_authors_sjec=num_corresponding_authors_sjec,
                num_coauthors_sjec=num_coauthors_sjec,
                
                # Journal-specific fields
                type_of_publication=request.form.get('type_of_publication'),
                type_of_journal=request.form.get('type_of_journal'),
                journal_name=request.form.get('journal_name'),
                journal_quartile=request.form.get('journal_quartile'),
                journal_impact_factor=float(request.form.get('journal_impact_factor', 0)) if request.form.get('journal_impact_factor') else None,
                indexing_type=request.form.get('indexing_type'),
                
                # Common fields - handle both journal and book publisher fields
                publisher_details=request.form.get('publisher_details') or request.form.get('publisher_details_book'),
                publication_date=pub_date,
                author_type=request.form.get('author_type'),
                status='draft' if action == 'draft' else 'submitted',
                submitted_at=now_ist() if action == 'submit' else None
            )
            
            db.session.add(application)
            db.session.flush()  # Get application.id before saving documents
            
            # Save uploaded documents
            timestamp = dt.now().strftime('%Y%m%d_%H%M%S')
            app_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'applications')
            os.makedirs(app_folder, exist_ok=True)
            
            for file in uploaded_files:
                if file and file.filename != '':
                    # Validate file type
                    file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                    allowed_extensions = {'pdf', 'png', 'jpg', 'jpeg', 'docx'}
                    
                    if file_ext not in allowed_extensions:
                        flash(f'File {file.filename} has invalid type. Only PDF, PNG, JPEG, DOCX allowed.', 'warning')
                        continue
                    
                    # Save file
                    original_filename = secure_filename(file.filename)
                    stored_filename = f'doc_{application.id}_{timestamp}_{original_filename}'
                    file_path = os.path.join(app_folder, stored_filename)
                    file.save(file_path)
                    
                    # Create document record
                    doc = ApplicationDocument(
                        application_id=application.id,
                        filename=file.filename,
                        stored_filename=stored_filename,
                        file_type=file_ext,
                        file_size=os.path.getsize(file_path),
                        document_type='publication' if file_ext == 'pdf' else 'supporting',
                        uploaded_by=current_user.id
                    )
                    db.session.add(doc)
            
            db.session.commit()
            
            if action == 'submit':
                # Auto-create publication from incentive application
                try:
                    from models import Publication
                    import shutil
                    
                    # Find first PDF document and copy to publications folder
                    publications_pdf = None
                    pdf_doc = next((doc for doc in application.documents if doc.file_type == 'pdf'), None)
                    
                    if pdf_doc:
                        app_pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], 'applications', pdf_doc.stored_filename)
                        
                        # Create user-specific publications folder
                        pub_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'publications', f'user_{current_user.id}')
                        os.makedirs(pub_folder, exist_ok=True)
                        
                        # Copy file to publications folder with new name
                        publications_pdf = f'pub_{current_user.id}_{timestamp}_{secure_filename(pdf_doc.filename)}'
                        pub_pdf_path = os.path.join(pub_folder, publications_pdf)
                        
                        shutil.copy2(app_pdf_path, pub_pdf_path)
                    
                    # Create publication record
                    publication = Publication(
                        publication_type=publication_category or 'Journal',  # Use application's publication category
                        title=application.title_of_paper,
                        publisher_name=application.publisher_details,
                        authors_names=current_user.name,  # Can be updated later by user
                        publication_details=f"{application.journal_name}, {pub_date.year if pub_date else 'N/A'}",
                        journal_conference_name=application.journal_name,
                        indexing_status=application.indexing_type if publication_category == 'Journal' else None,
                        quartile=application.journal_quartile if publication_category == 'Journal' else None,
                        impact_factor=application.journal_impact_factor if publication_category == 'Journal' else None,
                        year=pub_date.year if pub_date else dt.now().year,
                        month=pub_date.month if pub_date else None,
                        citation_count=0 if publication_category == 'Journal' else None,
                        pdf_filename=publications_pdf,
                        user_id=current_user.id,
                        department_id=current_user.department_id,
                        status='confirmed',  # Auto-confirmed from incentive
                        confirmed_at=now_ist()
                    )
                    
                    db.session.add(publication)
                    db.session.commit()
                    
                    flash(f'Publication automatically added to your publications list!', 'success')
                    
                except Exception as e:
                    # Don't fail the application submission if publication creation fails
                    print(f"Warning: Failed to auto-create publication: {str(e)}")
                
                # Notify HOD of applicant's department
                notify_hod_of_application(application)
                
                flash(f'Application submitted successfully! Tracking ID: {tracking_id}', 'success')
            else:
                flash(f'Application saved as draft. Tracking ID: {tracking_id}', 'info')
            
            return redirect(url_for('my_applications'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('apply_publication_incentive.html', now=dt.now())


@app.route('/applications/journal-support', methods=['GET', 'POST'])
@login_required
@role_required('faculty')
def apply_journal_support():
    """Journal Publication Support Application Form"""
    from models import ApplicationForm
    from datetime import datetime as dt
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        try:
            year = dt.now().year
            count = ApplicationForm.query.filter(
                ApplicationForm.application_type == 'journal_support',
                ApplicationForm.tracking_id.like(f'JS{year}%')
            ).count() + 1
            
            # Ensure tracking_id is unique
            tracking_id = f"JS{year}{count:04d}"
            while ApplicationForm.query.filter_by(tracking_id=tracking_id).first():
                count += 1
                tracking_id = f"JS{year}{count:04d}"
            
            pub_date_str = request.form.get('publication_date')
            pub_date = dt.strptime(pub_date_str, '%Y-%m-%d').date() if pub_date_str else None
            
            application = ApplicationForm(
                tracking_id=tracking_id,
                application_type='journal_support',
                applicant_id=current_user.id,
                applicant_name=current_user.name,
                title_of_paper=request.form.get('title_of_paper'),
                type_of_publication=request.form.get('type_of_publication'),
                type_of_journal=request.form.get('type_of_journal'),
                journal_name=request.form.get('journal_name'),
                journal_quartile=request.form.get('journal_quartile'),
                journal_impact_factor=float(request.form.get('journal_impact_factor', 0)) if request.form.get('journal_impact_factor') else None,
                indexing_type=request.form.get('indexing_type'),
                publisher_details=request.form.get('publisher_details'),
                apc_fees=float(request.form.get('apc_fees', 0)) if request.form.get('apc_fees') else None,
                expected_charges=float(request.form.get('expected_charges', 0)) if request.form.get('expected_charges') else None,
                publication_date=pub_date,
                status='draft' if action == 'draft' else 'submitted',
                submitted_at=now_ist() if action == 'submit' else None
            )
            
            db.session.add(application)
            db.session.commit()
            
            if action == 'submit':
                # Auto-create publication from journal support application
                try:
                    from models import Publication
                    import shutil
                    
                    # Find first PDF document and copy to publications folder
                    publications_pdf = None
                    # For journal support, look for uploaded documents if any
                    # (Note: journal support doesn't have multi-file upload yet, will be added in next task)
                    
                    # Create publication record
                    publication = Publication(
                        publication_type='Journal',
                        title=application.title_of_paper,
                        publisher_name=application.publisher_details,
                        authors_names=current_user.name,  # Can be updated later by user
                        publication_details=f"{application.journal_name}, {pub_date.year if pub_date else 'N/A'}",
                        journal_conference_name=application.journal_name,
                        indexing_status=application.indexing_type,
                        quartile=application.journal_quartile,
                        impact_factor=application.journal_impact_factor,
                        year=pub_date.year if pub_date else dt.now().year,
                        month=pub_date.month if pub_date else None,
                        pdf_filename=publications_pdf,
                        user_id=current_user.id,
                        department_id=current_user.department_id,
                        status='saved',  # Saved as draft - user can confirm later
                        confirmed_at=None,
                        can_edit=True
                    )
                    
                    db.session.add(publication)
                    db.session.commit()
                    
                    flash(f'Publication automatically added to your publications list as draft!', 'success')
                    
                except Exception as e:
                    # Don't fail the application submission if publication creation fails
                    print(f"Warning: Failed to auto-create publication from journal support: {str(e)}")
                
                # Notify HOD of applicant's department
                notify_hod_of_application(application)
                
                flash(f'Application submitted successfully! Tracking ID: {tracking_id}', 'success')
            else:
                flash(f'Application saved as draft. Tracking ID: {tracking_id}', 'info')
            
            return redirect(url_for('my_applications'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    return render_template('apply_journal_support.html', now=dt.now())


@app.route('/applications/my-applications')
@login_required
def my_applications():
    """Faculty view of their own applications"""
    from models import ApplicationForm
    
    applications = ApplicationForm.query.filter_by(applicant_id=current_user.id)\
        .order_by(ApplicationForm.created_at.desc()).all()
    
    return render_template('my_applications.html', applications=applications)


@app.route('/applications/view/<int:app_id>')
@login_required
def view_application(app_id):
    """View application details with full form preview"""
    from models import ApplicationForm, ACMReview, ACMPanel
    
    application = ApplicationForm.query.get_or_404(app_id)
    
    # Check permission
    has_access = False
    
    # Faculty can view their own applications
    if application.applicant_id == current_user.id:
        has_access = True
    
    # ACM members can view applications in ACM review stages and beyond
    # Also allow viewing if they have already reviewed the application
    elif current_user.role == 'faculty':
        acm_membership = ACMPanel.query.filter_by(
            member_id=current_user.id,
            is_active=True
        ).first()
        if acm_membership:
            # Check if application is in ACM review stage or beyond
            if application.status in ['hod_approved', 'acm_review', 'acm_approved', 'dean_review', 
                                      'dean_approved', 'principal_approved', 'director_approved', 'rejected']:
                has_access = True
            # Also allow if ACM member has already reviewed this application
            else:
                my_review = ACMReview.query.filter_by(
                    application_id=app_id,
                    reviewer_id=current_user.id
                ).first()
                if my_review:
                    has_access = True
    
    # HOD can view applications from their department
    elif current_user.role == 'hod':
        if application.applicant.department_id == current_user.department_id:
            has_access = True
    
    # Dean, Principal, VP, Admin can view all
    elif current_user.role in ['dean', 'principal', 'vice_principal', 'admin']:
        has_access = True
    
    if not has_access:
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get ACM reviews if any
    acm_reviews = ACMReview.query.filter_by(application_id=app_id).all()
    
    return render_template('view_application.html', 
                         application=application, 
                         acm_reviews=acm_reviews)


@app.route('/applications/print/<int:app_id>')
@login_required
def print_application(app_id):
    """Print-friendly view of application"""
    from models import ApplicationForm, ACMReview, ACMPanel, User
    from datetime import datetime
    
    application = ApplicationForm.query.get_or_404(app_id)
    
    # Check permission
    has_access = False
    
    # Faculty can view their own applications
    if application.applicant_id == current_user.id:
        has_access = True
    
    # ACM members can view applications in ACM review stages and beyond
    elif current_user.role == 'faculty':
        acm_membership = ACMPanel.query.filter_by(
            member_id=current_user.id,
            is_active=True
        ).first()
        if acm_membership:
            if application.status in ['hod_approved', 'acm_review', 'acm_approved', 'dean_review', 
                                      'dean_approved', 'principal_approved', 'director_approved', 'rejected']:
                has_access = True
            else:
                my_review = ACMReview.query.filter_by(
                    application_id=app_id,
                    reviewer_id=current_user.id
                ).first()
                if my_review:
                    has_access = True
    
    # HOD can view applications from their department
    elif current_user.role == 'hod':
        if application.applicant.department_id == current_user.department_id:
            has_access = True
    
    # Dean, Principal, VP, Admin can view all
    elif current_user.role in ['dean', 'principal', 'vice_principal', 'admin']:
        has_access = True
    
    if not has_access:
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get ACM reviews if any
    acm_reviews = ACMReview.query.filter_by(application_id=app_id).all()
    
    # Get all users for signature display
    all_users = User.query.all()
    
    # Current timestamp for document generation
    now = datetime.now()
    
    # Render appropriate template based on application type
    if application.application_type == 'conference_registration':
        template = 'print_conference_registration.html'
    elif application.application_type == 'publication_incentive':
        template = 'print_publication_incentive.html'
    elif application.application_type == 'journal_support':
        template = 'print_journal_support.html'
    else:
        # Fallback to conference registration template
        template = 'print_conference_registration.html'
    
    return render_template(template,
                         application=application,
                         acm_reviews=acm_reviews,
                         all_users=all_users,
                         now=now)


@app.route('/applications/download-pdf/<int:application_id>')
@login_required
def download_application_pdf(application_id):
    """Download application PDF file"""
    from models import ApplicationForm, ACMPanel, ACMReview
    from flask import send_file
    
    application = ApplicationForm.query.get_or_404(application_id)
    
    # Check permission
    has_access = False
    
    if application.applicant_id == current_user.id:
        has_access = True
    elif current_user.role == 'hod' and application.applicant.department_id == current_user.department_id:
        has_access = True
    elif current_user.role in ['dean', 'principal', 'vice_principal', 'admin']:
        has_access = True
    elif current_user.role == 'faculty':
        acm_membership = ACMPanel.query.filter_by(member_id=current_user.id, is_active=True).first()
        if acm_membership:
            # Check if application is in ACM review stage or beyond
            if application.status in ['hod_approved', 'acm_review', 'acm_approved', 'dean_review', 
                                      'dean_approved', 'principal_approved', 'director_approved', 'rejected']:
                has_access = True
            # Also allow if ACM member has already reviewed this application
            else:
                my_review = ACMReview.query.filter_by(
                    application_id=application_id,
                    reviewer_id=current_user.id
                ).first()
                if my_review:
                    has_access = True
    
    if not has_access:
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))
    
    if not application.publication_pdf:
        flash('No PDF file uploaded for this application.', 'warning')
        return redirect(url_for('view_application', app_id=application_id))
    
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], 'applications', application.publication_pdf)
    
    if not os.path.exists(pdf_path):
        flash('PDF file not found.', 'danger')
        return redirect(url_for('view_application', app_id=application_id))
    
    return send_file(pdf_path, as_attachment=True, download_name=f"{application.tracking_id}_publication.pdf")


@app.route('/application/document/<int:doc_id>/download')
@login_required
def download_application_document(doc_id):
    """Download application supporting document"""
    from models import ApplicationDocument, ApplicationForm, ACMPanel, ACMReview
    from flask import send_file
    
    document = ApplicationDocument.query.get_or_404(doc_id)
    application = document.application
    
    # Check permission (same as PDF download)
    has_access = False
    
    if application.applicant_id == current_user.id:
        has_access = True
    elif current_user.role == 'hod' and application.applicant.department_id == current_user.department_id:
        has_access = True
    elif current_user.role in ['dean', 'principal', 'vice_principal', 'admin', 'director', 'dean_secretary']:
        has_access = True
    elif current_user.role == 'faculty':
        acm_membership = ACMPanel.query.filter_by(member_id=current_user.id, is_active=True).first()
        if acm_membership and application.status in ['hod_approved', 'acm_review', 'acm_approved', 'dean_review', 
                                                      'dean_approved', 'principal_approved', 'director_approved', 'rejected']:
            has_access = True
    
    if not has_access:
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    
    # Check both possible locations (for backward compatibility)
    doc_path = os.path.join(app.config['UPLOAD_FOLDER'], 'applications', document.stored_filename)
    if not os.path.exists(doc_path):
        doc_path = os.path.join(app.config['UPLOAD_FOLDER'], 'application_documents', document.stored_filename)
    
    if not os.path.exists(doc_path):
        return jsonify({'success': False, 'error': 'File not found'}), 404
    
    # Check if view mode is requested (for PDFs)
    view_mode = request.args.get('view', '0') == '1'
    
    return send_file(doc_path, as_attachment=not view_mode, download_name=document.filename)


@app.route('/application/document/<int:doc_id>/delete', methods=['POST'])
@login_required
def delete_application_document(doc_id):
    """Delete application supporting document"""
    from models import ApplicationDocument
    
    document = ApplicationDocument.query.get_or_404(doc_id)
    application = document.application
    
    # Only applicant can delete documents and only if application is draft or not yet submitted
    if application.applicant_id != current_user.id:
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    
    if application.status not in ['draft', 'submitted']:
        return jsonify({'success': False, 'error': 'Cannot delete documents after application review has started'}), 403
    
    try:
        # Delete file from filesystem
        doc_path = os.path.join(app.config['UPLOAD_FOLDER'], 'application_documents', document.stored_filename)
        if os.path.exists(doc_path):
            os.remove(doc_path)
        
        # Delete database record
        db.session.delete(document)
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/applications/<int:application_id>/view-pdf')
@login_required
def view_application_pdf(application_id):
    """View application PDF file inline in browser"""
    from models import ApplicationForm, ACMPanel, ACMReview
    from flask import send_file
    
    application = ApplicationForm.query.get_or_404(application_id)
    
    # Check permission (same as download)
    has_access = False
    
    if application.applicant_id == current_user.id:
        has_access = True
    elif current_user.role == 'hod' and application.applicant.department_id == current_user.department_id:
        has_access = True
    elif current_user.role in ['dean', 'principal', 'vice_principal', 'admin']:
        has_access = True
    elif current_user.role == 'faculty':
        acm_membership = ACMPanel.query.filter_by(member_id=current_user.id, is_active=True).first()
        if acm_membership:
            # Check if application is in ACM review stage or beyond
            if application.status in ['hod_approved', 'acm_review', 'acm_approved', 'dean_review', 
                                      'dean_approved', 'principal_approved', 'director_approved', 'rejected']:
                has_access = True
            # Also allow if ACM member has already reviewed this application
            else:
                my_review = ACMReview.query.filter_by(
                    application_id=application_id,
                    reviewer_id=current_user.id
                ).first()
                if my_review:
                    has_access = True
    
    if not has_access:
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))
    
    if not application.publication_pdf:
        flash('No PDF file uploaded for this application.', 'warning')
        return redirect(url_for('view_application', app_id=application_id))
    
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], 'applications', application.publication_pdf)
    
    if not os.path.exists(pdf_path):
        flash('PDF file not found.', 'danger')
        return redirect(url_for('view_application', app_id=application_id))
    
    # Send file inline for viewing in browser
    return send_file(pdf_path, as_attachment=False, mimetype='application/pdf')


@app.route('/applications/edit/<int:app_id>', methods=['GET', 'POST'])
@login_required
def edit_application(app_id):
    """Edit draft application"""
    from models import ApplicationForm
    
    application = ApplicationForm.query.get_or_404(app_id)
    
    # Only owner can edit and only if it's still a draft
    if application.applicant_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('my_applications'))
    
    if application.status != 'draft':
        flash('Only draft applications can be edited.', 'warning')
        return redirect(url_for('view_application', app_id=app_id))
    
    if request.method == 'POST':
        try:
            action = request.form.get('action')
            
            # Update common fields
            application.title_of_paper = request.form.get('title_of_paper')
            
            if application.application_type == 'conference_registration':
                from datetime import datetime as dt_parser
                application.type_of_conference = request.form.get('type_of_conference')
                application.conference_name = request.form.get('conference_name')
                application.conference_organizer = request.form.get('conference_organizer')
                
                # Handle separate place and date fields
                conference_place = request.form.get('conference_place', '')
                conference_date_str = request.form.get('conference_date', '')
                
                # Parse date string to date object
                conference_date_obj = None
                if conference_date_str:
                    try:
                        conference_date_obj = dt_parser.strptime(conference_date_str, '%Y-%m-%d').date()
                    except:
                        conference_date_obj = None
                
                # Backward compatibility - try old field if new fields are empty
                if not conference_place and not conference_date_str:
                    application.conference_place_date = request.form.get('conference_place_date', '')
                else:
                    application.conference_place = conference_place
                    application.conference_date = conference_date_obj
                    application.conference_place_date = f"{conference_place} - {conference_date_str}" if conference_place and conference_date_str else ''
                
                application.conference_fee = float(request.form.get('conference_fee', 0))
            else:
                # Publication incentive or journal support
                from datetime import datetime as dt
                
                # Get new fields for publication incentive
                if application.application_type == 'publication_incentive':
                    application.publication_category = request.form.get('publication_category')
                    application.is_national_international = request.form.get('is_national_international')
                    application.num_first_authors_sjec = int(request.form.get('num_first_authors_sjec', 0))
                    application.num_corresponding_authors_sjec = int(request.form.get('num_corresponding_authors_sjec', 0))
                    application.num_coauthors_sjec = int(request.form.get('num_coauthors_sjec', 0))
                
                application.type_of_publication = request.form.get('type_of_publication')
                application.type_of_journal = request.form.get('type_of_journal')
                application.journal_name = request.form.get('journal_name')
                application.journal_quartile = request.form.get('journal_quartile')
                impact_factor = request.form.get('journal_impact_factor')
                application.journal_impact_factor = float(impact_factor) if impact_factor else None
                application.indexing_type = request.form.get('indexing_type')
                
                # Handle both journal and book publisher fields
                application.publisher_details = request.form.get('publisher_details') or request.form.get('publisher_details_book')
                
                # Handle both journal and book date fields
                pub_date_str = request.form.get('publication_date') or request.form.get('publication_date_book')
                application.publication_date = dt.strptime(pub_date_str, '%Y-%m-%d').date() if pub_date_str else None
                
                # Publication incentive specific - author type
                if application.application_type == 'publication_incentive':
                    application.author_type = request.form.get('author_type')
                
                # Journal support specific fields
                if application.application_type == 'journal_support':
                    apc_fees = request.form.get('apc_fees')
                    application.apc_fees = float(apc_fees) if apc_fees else None
                    expected_charges = request.form.get('expected_charges')
                    application.expected_charges = float(expected_charges) if expected_charges else None
                
                # Handle document uploads for publication incentive
                if application.application_type == 'publication_incentive':
                    from models import ApplicationDocument
                    
                    uploaded_files = request.files.getlist('documents')
                    
                    # Check if new documents uploaded
                    if uploaded_files and any(f.filename != '' for f in uploaded_files):
                        # Save new documents
                        timestamp = dt.now().strftime('%Y%m%d_%H%M%S')
                        app_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'applications')
                        os.makedirs(app_folder, exist_ok=True)
                        
                        for file in uploaded_files:
                            if file and file.filename != '':
                                # Validate file type
                                file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
                                allowed_extensions = {'pdf', 'png', 'jpg', 'jpeg', 'docx'}
                                
                                if file_ext not in allowed_extensions:
                                    flash(f'File {file.filename} has invalid type. Only PDF, PNG, JPEG, DOCX allowed.', 'warning')
                                    continue
                                
                                # Save file
                                original_filename = secure_filename(file.filename)
                                stored_filename = f'doc_{application.id}_{timestamp}_{original_filename}'
                                file_path = os.path.join(app_folder, stored_filename)
                                file.save(file_path)
                                
                                # Create document record
                                doc = ApplicationDocument(
                                    application_id=application.id,
                                    filename=file.filename,
                                    stored_filename=stored_filename,
                                    file_type=file_ext,
                                    file_size=os.path.getsize(file_path),
                                    document_type='publication' if file_ext == 'pdf' else 'supporting',
                                    uploaded_by=current_user.id
                                )
                                db.session.add(doc)
                    
                    # Validate at least one document exists if submitting
                    if action == 'submit':
                        existing_docs = ApplicationDocument.query.filter_by(application_id=application.id).count()
                        if existing_docs == 0:
                            flash('At least one supporting document is required for submission!', 'danger')
                            return render_template('apply_publication_incentive.html', 
                                                 form_data=application, 
                                                 now=dt.now(), 
                                                 edit_mode=True)
                        
                        # Check if at least one PDF exists
                        has_pdf = ApplicationDocument.query.filter_by(
                            application_id=application.id,
                            file_type='pdf'
                        ).first() is not None
                        
                        if not has_pdf:
                            flash('At least one PDF document is required for submission!', 'danger')
                            return render_template('apply_publication_incentive.html', 
                                                 form_data=application, 
                                                 now=dt.now(), 
                                                 edit_mode=True)
            
            # Update status if submitting
            if action == 'submit':
                application.status = 'submitted'
                application.submitted_at = now_ist()
                
                # Notify HOD of applicant's department
                notify_hod_of_application(application)
            
            db.session.commit()
            
            flash(f'Application {"submitted" if action == "submit" else "updated"} successfully!', 'success')
            return redirect(url_for('my_applications'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    # Render appropriate edit form based on type
    if application.application_type == 'conference_registration':
        template = 'apply_conference_registration.html'
    elif application.application_type == 'publication_incentive':
        template = 'apply_publication_incentive.html'
    else:
        template = 'apply_journal_support.html'
    
    from datetime import datetime as dt
    return render_template(template, form_data=application, now=dt.now(), edit_mode=True)


@app.route('/applications/delete/<int:app_id>', methods=['POST'])
@login_required
def delete_application(app_id):
    """Delete draft application"""
    from models import ApplicationForm
    
    application = ApplicationForm.query.get_or_404(app_id)
    
    # Only owner can delete and only if it's still a draft
    if application.applicant_id != current_user.id:
        return jsonify({'success': False, 'error': 'Access denied'}), 403
    
    if application.status != 'draft':
        return jsonify({'success': False, 'error': 'Only draft applications can be deleted'}), 400
    
    try:
        tracking_id = application.tracking_id
        db.session.delete(application)
        db.session.commit()
        
        return jsonify({'success': True, 'message': f'Application {tracking_id} deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# ACM PANEL MANAGEMENT ROUTES (Dean only)
# ============================================================================

@app.route('/acm-panel/manage', methods=['GET', 'POST'])
@login_required
@role_required('dean')
def manage_acm_panel():
    """Dean manages ACM panel members"""
    from models import ACMPanel, ApplicationForm
    
    if request.method == 'POST':
        member_id = request.form.get('member_id')
        
        try:
            # Check if already 3 active members
            active_count = ACMPanel.query.filter_by(is_active=True).count()
            if active_count >= 3:
                flash('Maximum 3 ACM members allowed. Please deactivate a member first.', 'warning')
                return redirect(url_for('manage_acm_panel'))
            
            # Check if member already exists
            existing = ACMPanel.query.filter_by(member_id=member_id).first()
            if existing:
                if existing.is_active:
                    flash('This faculty member is already in the ACM panel.', 'warning')
                else:
                    # Reactivate
                    existing.is_active = True
                    db.session.commit()
                    flash('ACM member reactivated successfully!', 'success')
            else:
                # Add new member
                acm_member = ACMPanel(
                    member_id=member_id,
                    appointed_by=current_user.id,
                    is_active=True
                )
                db.session.add(acm_member)
                db.session.commit()
                
                # Notify the member
                member = User.query.get(member_id)
                create_notification(
                    user_id=member_id,
                    title='Appointed to ACM Panel',
                    message=f'You have been appointed as an Assessment Committee Member by {current_user.name}. You will now review faculty applications.',
                    notification_type='success'
                )
                
                flash(f'{member.name} added to ACM panel successfully!', 'success')
            
            return redirect(url_for('manage_acm_panel'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    # GET request
    acm_members = ACMPanel.query.all()
    
    # Get faculty not in panel
    acm_member_ids = [m.member_id for m in acm_members if m.is_active]
    available_faculty = User.query.filter(
        User.role == 'faculty',
        ~User.id.in_(acm_member_ids)
    ).order_by(User.name).all()
    
    # Get statistics
    pending_applications = ApplicationForm.query.filter_by(status='submitted').count()
    acm_approved = ApplicationForm.query.filter_by(status='acm_approved').count()
    awaiting_dean = ApplicationForm.query.filter_by(status='dean_review').count()
    
    return render_template('manage_acm_panel.html',
                         acm_members=acm_members,
                         available_faculty=available_faculty,
                         pending_applications=pending_applications,
                         acm_approved=acm_approved,
                         awaiting_dean=awaiting_dean)


@app.route('/acm-panel/deactivate/<int:member_id>', methods=['POST'])
@login_required
@role_required('dean')
def deactivate_acm_member(member_id):
    """Deactivate an ACM member"""
    from models import ACMPanel
    
    try:
        acm_member = ACMPanel.query.get_or_404(member_id)
        acm_member.is_active = False
        db.session.commit()
        
        # Notify the member
        create_notification(
            user_id=acm_member.member_id,
            title='ACM Panel Status Changed',
            message='You have been deactivated from the Assessment Committee panel.',
            notification_type='warning'
        )
        
        return jsonify({'success': True, 'message': 'Member deactivated successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/acm-panel/activate/<int:member_id>', methods=['POST'])
@login_required
@role_required('dean')
def activate_acm_member(member_id):
    """Activate an ACM member"""
    from models import ACMPanel
    
    try:
        # Check if already 3 active members
        active_count = ACMPanel.query.filter_by(is_active=True).count()
        if active_count >= 3:
            return jsonify({'success': False, 'error': 'Maximum 3 active members allowed'}), 400
        
        acm_member = ACMPanel.query.get_or_404(member_id)
        acm_member.is_active = True
        db.session.commit()
        
        # Notify the member
        create_notification(
            user_id=acm_member.member_id,
            title='ACM Panel Status Changed',
            message='You have been activated in the Assessment Committee panel.',
            notification_type='success'
        )
        
        return jsonify({'success': True, 'message': 'Member activated successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/acm-panel/remove/<int:member_id>', methods=['POST'])
@login_required
@role_required('dean')
def remove_acm_member(member_id):
    """Remove an ACM member permanently"""
    from models import ACMPanel
    
    try:
        acm_member = ACMPanel.query.get_or_404(member_id)
        member_user_id = acm_member.member_id
        
        db.session.delete(acm_member)
        db.session.commit()
        
        # Notify the member
        create_notification(
            user_id=member_user_id,
            title='Removed from ACM Panel',
            message='You have been removed from the Assessment Committee panel.',
            notification_type='info'
        )
        
        return jsonify({'success': True, 'message': 'Member removed successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# HOD REVIEW ROUTES
# ============================================================================

@app.route('/hod/review-applications')
@login_required
@role_required('hod')
def hod_review_applications():
    """HOD review interface for department applications"""
    from models import ApplicationForm, User
    
    # Ensure HOD has a department assigned
    if not current_user.department_id:
        flash('Your account is not assigned to a department. Please contact the administrator.', 'danger')
        return redirect(url_for('dashboard'))
    
    # Get faculty in HOD's department only
    dept_faculty_ids = [u.id for u in User.query.filter_by(
        department_id=current_user.department_id,
        role='faculty'
    ).all()]
    
    # If no faculty in department, return empty lists
    if not dept_faculty_ids:
        dept_faculty_ids = [0]  # Dummy ID to avoid SQL errors
    
    # Pending applications from HOD's department only
    pending_applications = ApplicationForm.query.filter(
        ApplicationForm.status == 'submitted',
        ApplicationForm.applicant_id.in_(dept_faculty_ids)
    ).order_by(ApplicationForm.created_at.asc()).all()
    
    # Reviewed applications (approved or rejected by this HOD)
    reviewed_applications = ApplicationForm.query.filter(
        ApplicationForm.hod_reviewer_id == current_user.id
    ).order_by(ApplicationForm.hod_reviewed_at.desc()).all()
    
    # All applications from department only
    all_applications = ApplicationForm.query.filter(
        ApplicationForm.applicant_id.in_(dept_faculty_ids)
    ).order_by(ApplicationForm.created_at.desc()).all()
    
    # Statistics
    approved_count = ApplicationForm.query.filter_by(
        hod_reviewer_id=current_user.id,
        hod_approved=True
    ).count()
    rejected_count = ApplicationForm.query.filter_by(
        hod_reviewer_id=current_user.id,
        hod_approved=False
    ).count()
    
    return render_template('hod_review_applications.html',
                         pending_applications=pending_applications,
                         reviewed_applications=reviewed_applications,
                         all_applications=all_applications,
                         pending_count=len(pending_applications),
                         approved_count=approved_count,
                         rejected_count=rejected_count,
                         total_reviewed=approved_count + rejected_count)


@app.route('/hod/review/<int:app_id>', methods=['POST'])
@login_required
@role_required('hod')
def submit_hod_review(app_id):
    """Submit HOD review for an application"""
    from models import ApplicationForm, User, ACMPanel
    
    try:
        application = ApplicationForm.query.get_or_404(app_id)
        
        # Verify application is from HOD's department
        if application.applicant.department_id != current_user.department_id:
            return jsonify({'success': False, 'error': 'Access denied'}), 403
        
        # Verify application is in correct status
        if application.status != 'submitted':
            return jsonify({'success': False, 'error': 'Application not in correct status for HOD review'}), 400
        
        # Get review data
        data = request.get_json()
        decision = data.get('decision')
        comments = data.get('comments', '').strip()
        
        if not decision or decision not in ['approve', 'reject']:
            return jsonify({'success': False, 'error': 'Invalid decision'}), 400
        
        # Comments are mandatory only for rejection
        if decision == 'reject' and not comments:
            return jsonify({'success': False, 'error': 'Comments are required when rejecting an application'}), 400
        
        # Update application with HOD review
        application.hod_reviewer_id = current_user.id
        application.hod_comments = comments
        application.hod_reviewed_at = now_ist()
        
        if decision == 'approve':
            application.hod_approved = True
            application.status = 'hod_approved'
            
            # Notify applicant
            create_notification(
                user_id=application.applicant_id,
                title='Application Approved by HOD',
                message=f'Your application {application.tracking_id} has been approved by your HOD and forwarded to the ACM panel for review.',
                notification_type='success'
            )
            
            # Notify ACM panel members
            acm_members = ACMPanel.query.filter_by(is_active=True).all()
            for member in acm_members:
                create_notification(
                    user_id=member.member_id,
                    title=f'New Application for Review: {application.tracking_id}',
                    message=f'Application {application.tracking_id} from {application.applicant.name} has been approved by HOD and requires ACM review.',
                    notification_type='info'
                )
        else:
            application.hod_approved = False
            application.status = 'rejected'
            
            # Notify applicant
            create_notification(
                user_id=application.applicant_id,
                title='Application Rejected by HOD',
                message=f'Your application {application.tracking_id} has been rejected by your HOD. Please review the HOD comments.',
                notification_type='danger'
            )
        
        db.session.commit()
        
        # Create audit log
        create_audit_log(
            action='hod_review',
            target_type='application',
            target_id=application.id,
            details=f'Reviewed application {application.tracking_id}: {decision}',
            user_id=current_user.id
        )
        
        return jsonify({
            'success': True,
            'message': f'Review submitted successfully. Application {decision}d.',
            'status': application.status
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# ACM REVIEW ROUTES
# ============================================================================

# ACM Review Routes
@app.route('/acm/review-applications')
@login_required
@role_required('faculty')
def acm_review_applications():
    """ACM member review interface"""
    from models import ACMPanel, ApplicationForm, ACMReview
    
    # Verify user is an active ACM member
    acm_membership = ACMPanel.query.filter_by(
        member_id=current_user.id,
        is_active=True
    ).first()
    
    if not acm_membership:
        flash('Access denied. You are not an active ACM panel member.', 'danger')
        return redirect(url_for('faculty_dashboard'))
    
    # Get applications that user has already reviewed
    reviewed_app_ids = [r.application_id for r in ACMReview.query.filter_by(
        reviewer_id=current_user.id
    ).all()]
    
    # Pending applications: hod_approved or acm_review status, not reviewed by this user
    # EXCLUDE own applications from pending (can't review own)
    pending_applications = ApplicationForm.query.filter(
        ApplicationForm.status.in_(['hod_approved', 'acm_review']),
        ~ApplicationForm.id.in_(reviewed_app_ids) if reviewed_app_ids else True,
        ApplicationForm.applicant_id != current_user.id  # Can't review own applications
    ).order_by(ApplicationForm.created_at.asc()).all()
    
    # Applications already reviewed by this user
    if reviewed_app_ids:
        reviewed_applications = ApplicationForm.query.filter(
            ApplicationForm.id.in_(reviewed_app_ids)
        ).join(ACMReview, ApplicationForm.id == ACMReview.application_id).filter(
            ACMReview.reviewer_id == current_user.id
        ).order_by(ACMReview.reviewed_at.desc()).all()
    else:
        reviewed_applications = []
    
    # All applications in ACM review stages and beyond (INCLUDE own applications for viewing)
    # Show all applications that have passed HOD approval (so ACM members can see outcomes)
    all_applications = ApplicationForm.query.filter(
        ApplicationForm.status.in_(['hod_approved', 'acm_review', 'acm_approved', 'dean_review', 
                                     'dean_approved', 'principal_approved', 'director_approved', 'rejected'])
    ).order_by(ApplicationForm.created_at.desc()).all()
    
    # Statistics
    my_reviews = ACMReview.query.filter_by(reviewer_id=current_user.id).all()
    approved_count = len([r for r in my_reviews if r.decision == 'approved'])
    rejected_count = len([r for r in my_reviews if r.decision == 'rejected'])
    total_reviewed = len(my_reviews)
    pending_count = len(pending_applications)
    
    return render_template('acm_review_applications.html',
                         pending_applications=pending_applications,
                         reviewed_applications=reviewed_applications,
                         all_applications=all_applications,
                         pending_count=pending_count,
                         approved_count=approved_count,
                         rejected_count=rejected_count,
                         total_reviewed=total_reviewed)


@app.route('/acm/get-application-data/<int:app_id>')
@login_required
@role_required('faculty')
def get_acm_application_data(app_id):
    """Get application data for ACM review modal"""
    from models import ACMPanel, ApplicationForm
    
    # Verify user is an active ACM member
    acm_membership = ACMPanel.query.filter_by(
        member_id=current_user.id,
        is_active=True
    ).first()
    
    if not acm_membership:
        return jsonify({'success': False, 'error': 'You are not an active ACM panel member'}), 403
    
    application = ApplicationForm.query.get_or_404(app_id)
    
    # Check for ACM-updated values from previous reviews
    from models import ACMReview
    acm_reviews = ACMReview.query.filter_by(
        application_id=app_id,
        decision='approved'
    ).all()
    
    # Start with current application values (which may already be updated)
    impact_factor = application.journal_impact_factor
    indexing_type = application.indexing_type
    quartile = application.journal_quartile
    
    # Apply any edits from approved ACM reviews to show the latest values
    for review in acm_reviews:
        if review.edited_fields:
            try:
                edits = json.loads(review.edited_fields)
                for field, change_data in edits.items():
                    if isinstance(change_data, dict) and 'new' in change_data:
                        new_value = change_data['new']
                    else:
                        new_value = change_data
                    
                    if field == 'journal_impact_factor':
                        impact_factor = new_value
                    elif field == 'indexing_type':
                        indexing_type = new_value
                    elif field == 'journal_quartile':
                        quartile = new_value
            except:
                pass
    
    return jsonify({
        'success': True,
        'journal_impact_factor': impact_factor,
        'indexing_type': indexing_type,
        'journal_quartile': quartile
    })


@app.route('/acm/review/<int:app_id>', methods=['POST'])
@login_required
@role_required('faculty')
def submit_acm_review(app_id):
    """Submit ACM review for an application"""
    from models import ACMPanel, ApplicationForm, ACMReview
    
    # Verify user is an active ACM member
    acm_membership = ACMPanel.query.filter_by(
        member_id=current_user.id,
        is_active=True
    ).first()
    
    if not acm_membership:
        return jsonify({'success': False, 'error': 'You are not an active ACM panel member'}), 403
    
    try:
        application = ApplicationForm.query.get_or_404(app_id)
        
        # Block reviews if application is past ACM stage (dean_approved or later)
        if application.status not in ['submitted', 'hod_approved', 'acm_review']:
            return jsonify({'success': False, 'error': 'This application is no longer in ACM review stage'}), 403
        
        # Conflict of interest check
        if application.applicant_id == current_user.id:
            return jsonify({'success': False, 'error': 'Cannot review your own application'}), 403
        
        # Check if already reviewed
        existing_review = ACMReview.query.filter_by(
            application_id=app_id,
            reviewer_id=current_user.id
        ).first()
        
        if existing_review:
            return jsonify({'success': False, 'error': 'You have already reviewed this application'}), 400
        
        # Get review data
        data = request.get_json()
        decision = data.get('decision')
        comments = data.get('comments', '').strip()
        edited_fields = data.get('edited_fields')
        
        if not decision or decision not in ['approve', 'reject']:
            return jsonify({'success': False, 'error': 'Invalid decision'}), 400
        
        # Comments mandatory only for rejection, optional for approval
        if decision == 'reject' and not comments:
            return jsonify({'success': False, 'error': 'Comments are required for rejection'}), 400
        
        # Parse edited fields if provided
        edited_fields_json = None
        edited_fields_with_history = None
        if edited_fields:
            try:
                edited_fields_json = json.loads(edited_fields)
                # Store both old and new values for change tracking
                edited_fields_with_history = {}
                for field, new_value in edited_fields_json.items():
                    if hasattr(application, field):
                        old_value = getattr(application, field)
                        edited_fields_with_history[field] = {
                            'old': old_value,
                            'new': new_value
                        }
            except:
                return jsonify({'success': False, 'error': 'Invalid JSON in edited fields'}), 400
        
        # Read signature file as binary data for snapshot
        signature_binary = None
        if current_user.signature:
            try:
                signature_path = os.path.join(get_user_folder(current_user.id, 'signatures'), current_user.signature)
                if os.path.exists(signature_path):
                    with open(signature_path, 'rb') as f:
                        signature_binary = f.read()
            except Exception as e:
                print(f"Warning: Could not read signature file: {e}")
        
        # Create ACM review record
        review = ACMReview(
            application_id=app_id,
            reviewer_id=current_user.id,
            decision='approved' if decision == 'approve' else 'rejected',
            comments=comments,
            edited_fields=json.dumps(edited_fields_with_history) if edited_fields_with_history else None,
            reviewed_at=now_ist(),
            reviewer_signature_snapshot=signature_binary,
            reviewer_name_snapshot=current_user.name
        )
        db.session.add(review)
        db.session.flush()  # Flush to get the review in the session
        
        # Update application status
        if application.status == 'hod_approved':
            application.status = 'acm_review'
            # Set ACM review start time
            if not application.acm_review_started_at:
                application.acm_review_started_at = now_ist()
        
        # Count total ACM approvals (including the current one we just added)
        all_reviews = ACMReview.query.filter_by(application_id=app_id).all()
        approval_count = len([r for r in all_reviews if r.decision == 'approved'])
        application.acm_approvals_count = approval_count
        
        # Get total active ACM members (excluding ineligible ones)
        total_acm_members = ACMPanel.query.filter_by(is_active=True).count()
        
        # Check for ineligible members (conflict of interest)
        ineligible_members = ACMPanel.query.filter_by(
            member_id=application.applicant_id,
            is_active=True
        ).count()
        
        # Calculate eligible ACM members for this application
        eligible_acm_count = total_acm_members - ineligible_members
        total_reviews = len(all_reviews)
        
        # Decision Logic:
        # 1. If all 3 eligible members reviewed and 2+ approvals â†’ Forward to Dean
        # 2. If 2 members reviewed, both approved, and only 2 eligible â†’ Forward to Dean immediately
        # 3. If 2 approvals but 3rd member hasn't reviewed â†’ Wait for 3rd member
        # 4. If all eligible members reviewed and <2 approvals â†’ Reject
        
        # Check if all eligible members have reviewed
        all_reviewed = (total_reviews >= eligible_acm_count)
        
        # Timer-based logic: Track when 2nd approval happens
        if approval_count >= 2:
            if not application.acm_second_approval_at:
                # First time reaching 2 approvals - start 48-hour timer
                application.acm_second_approval_at = now_ist()
                application.acm_auto_forward_scheduled = True
                
                # Notify 3rd member about deadline
                notify_third_member_deadline(application, eligible_acm_count)
        
        # Only forward if all members reviewed OR timer expired
        if approval_count >= 2 and (all_reviewed or getattr(application, 'acm_auto_forwarded', False)):
            # All members reviewed OR auto-forwarded by timer - forward to Dean
            application.status = 'acm_approved'
            application.acm_approved_at = now_ist()  # Record when forwarded to Dean
            
            # Mark as forwarded (either by all reviews or auto-forward)
            if not all_reviewed:
                application.acm_auto_forwarded = True
            
            # Track field changes for notification
            field_changes = {}
            field_labels = {
                'journal_impact_factor': 'Impact Factor',
                'indexing_type': 'Indexing Type',
                'journal_quartile': 'Journal Quartile'
            }
            
            # Apply edited fields from all approved reviews
            for acm_review in all_reviews:
                if acm_review.decision == 'approved' and acm_review.edited_fields:
                    try:
                        edits = json.loads(acm_review.edited_fields)
                        # Track changes and apply each edited field
                        for field, change_data in edits.items():
                            if hasattr(application, field):
                                # Handle both old format (direct value) and new format (dict with old/new)
                                if isinstance(change_data, dict) and 'new' in change_data:
                                    new_value = change_data['new']
                                    old_value = change_data['old']
                                else:
                                    # Old format compatibility
                                    new_value = change_data
                                    old_value = getattr(application, field)
                                
                                if old_value != new_value:
                                    # Store the change
                                    if field not in field_changes:
                                        field_changes[field] = {
                                            'old': old_value,
                                            'new': new_value,
                                            'label': field_labels.get(field, field)
                                        }
                                setattr(application, field, new_value)
                    except:
                        pass  # Skip invalid JSON
            
            # Apply current review's edits if approved
            if decision == 'approve' and edited_fields_with_history:
                for field, change_data in edited_fields_with_history.items():
                    if hasattr(application, field):
                        new_value = change_data['new']
                        old_value = change_data['old']
                        if old_value != new_value:
                            if field not in field_changes:
                                field_changes[field] = {
                                    'old': old_value,
                                    'new': new_value,
                                    'label': field_labels.get(field, field)
                                }
                        setattr(application, field, new_value)
            
            # Build notification message with field changes
            notification_msg = f'Your application {application.tracking_id} has been approved by the Assessment Committee and forwarded to the Dean for final review.'
            if field_changes:
                notification_msg += '\n\nThe following fields were updated by the ACM panel:'
                for field, change in field_changes.items():
                    notification_msg += f"\nâ€¢ {change['label']}: {change['old']} â†’ {change['new']}"
            
            # Notify applicant
            create_notification(
                user_id=application.applicant_id,
                title='Application Approved by ACM Panel',
                message=notification_msg,
                notification_type='success'
            )
            
            # Notify Dean
            dean_user = User.query.filter_by(role='dean').first()
            if dean_user:
                create_notification(
                    user_id=dean_user.id,
                    title='New Application for Review',
                    message=f'Application {application.tracking_id} has been approved by ACM panel and requires your review.',
                    notification_type='info'
                )
        
        elif all_reviewed and approval_count < 2:
            # All eligible members reviewed but less than 2 approvals - reject
            application.status = 'rejected'
            
            # Notify applicant
            create_notification(
                user_id=application.applicant_id,
                title='Application Rejected',
                message=f'Your application {application.tracking_id} has been rejected by the Assessment Committee. Please review the ACM comments.',
                notification_type='danger'
            )
        else:
            # Still pending more reviews (waiting for 3rd member or more approvals)
            # Notify applicant of review progress
            create_notification(
                user_id=application.applicant_id,
                title='ACM Review Progress',
                message=f'Your application {application.tracking_id} has received {total_reviews} ACM review(s) ({approval_count} approved).',
                notification_type='info'
            )
        
        db.session.commit()
        
        # Create audit log
        create_audit_log(
            action='acm_review',
            target_type='application',
            target_id=application.id,
            details=f'Reviewed application {application.tracking_id}: {decision}',
            user_id=current_user.id
        )
        
        return jsonify({
            'success': True,
            'message': f'Review submitted successfully. Application {decision}d.',
            'approval_count': approval_count,
            'total_reviews': total_reviews,
            'status': application.status
        })
        
    except Exception as e:
        db.session.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f"ERROR in submit_acm_review: {error_details}")
        return jsonify({'success': False, 'error': f'{str(e)} - Check server logs for details'}), 500


# Dean Review Routes
@app.route('/dean/review-applications')
@login_required
@role_required('dean')
def dean_review_applications():
    """Dean interface to review ACM-approved applications"""
    from models import ApplicationForm
    
    # Pending applications: ACM approved, awaiting Dean review
    pending_applications = ApplicationForm.query.filter_by(
        status='acm_approved'
    ).order_by(ApplicationForm.created_at.asc()).all()
    
    # Reviewed applications: Dean has reviewed
    reviewed_applications = ApplicationForm.query.filter(
        ApplicationForm.status.in_(['dean_approved', 'dean_review', 'principal_approved', 'director_approved', 'rejected']),
        ApplicationForm.dean_reviewed_at.isnot(None)
    ).order_by(ApplicationForm.dean_reviewed_at.desc()).all()
    
    # Statistics
    approved_count = ApplicationForm.query.filter_by(dean_approved=True).count()
    rejected_count = ApplicationForm.query.filter(
        ApplicationForm.dean_approved == False,
        ApplicationForm.dean_reviewed_at.isnot(None)
    ).count()
    total_reviewed = approved_count + rejected_count
    pending_count = len(pending_applications)
    
    return render_template('dean_review_applications.html',
                         pending_applications=pending_applications,
                         reviewed_applications=reviewed_applications,
                         pending_count=pending_count,
                         approved_count=approved_count,
                         rejected_count=rejected_count,
                         total_reviewed=total_reviewed)


@app.route('/dean/get-acm-reviews/<int:app_id>')
@login_required
@role_required('dean')
def get_acm_reviews(app_id):
    """Get ACM review details for an application"""
    from models import ACMReview, ApplicationForm
    
    try:
        application = ApplicationForm.query.get_or_404(app_id)
        reviews = ACMReview.query.filter_by(application_id=app_id).all()
        
        review_data = []
        for review in reviews:
            review_data.append({
                'reviewer_name': review.reviewer.name,
                'decision': review.decision,
                'comments': review.comments,
                'reviewed_at': review.reviewed_at.strftime('%d %b %Y %I:%M %p')
            })
        
        # Include application details (especially author_type for publication incentive)
        app_details = {
            'title': application.title_of_paper,
            'application_type': application.application_type,
            'author_type': application.author_type if application.application_type == 'publication_incentive' else None,
            'journal_quartile': application.journal_quartile if application.application_type == 'publication_incentive' else None,
            'indexing_type': application.indexing_type,
            'journal_impact_factor': application.journal_impact_factor if application.application_type in ['publication_incentive', 'journal_support'] else None,
            'publication_category': application.publication_category,
            'is_national_international': application.is_national_international,
            'num_first_authors_sjec': application.num_first_authors_sjec,
            'num_corresponding_authors_sjec': application.num_corresponding_authors_sjec,
            'num_coauthors_sjec': application.num_coauthors_sjec,
            # Faculty requested amounts - using correct field names from model
            'actual_amount': application.conference_fee if application.application_type == 'conference_registration' else (application.apc_fees if application.application_type == 'journal_support' else None),
            'expected_amount': application.conference_fee if application.application_type == 'conference_registration' else (application.expected_charges if application.application_type == 'journal_support' else None),
        }
        
        return jsonify({
            'success': True, 
            'reviews': review_data,
            'application': app_details
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/dean/get-faculty-wallet/<int:app_id>')
@login_required
@role_required('dean')
def get_faculty_wallet(app_id):
    """Get faculty wallet balance and requested amount for an application"""
    from models import ApplicationForm
    
    try:
        application = ApplicationForm.query.get_or_404(app_id)
        
        # Get wallet balances
        rrf_balance = application.applicant.wallet_balance or 0
        epp_balance = application.applicant.epp_balance or 0
        epp_inr = epp_balance * 2000  # EPP conversion rate
        
        # Get requested amount based on application type
        requested_amount = 0
        if application.application_type == 'conference_registration':
            requested_amount = application.conference_fee or 0
        elif application.application_type == 'journal_support':
            requested_amount = application.expected_charges or 0
        
        return jsonify({
            'success': True,
            'wallet': {
                'rrf_balance': rrf_balance,
                'epp_balance': epp_balance,
                'epp_inr': epp_inr,
                'total_available': rrf_balance + epp_inr
            },
            'requested_amount': requested_amount
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/dean/review/<int:app_id>', methods=['POST'])
@login_required
@role_required('dean')
def submit_dean_review(app_id):
    """Submit Dean's review and recommendation"""
    from models import ApplicationForm
    
    try:
        application = ApplicationForm.query.get_or_404(app_id)
        
        # Verify application is in correct status
        if application.status != 'acm_approved':
            return jsonify({'success': False, 'error': 'Application not in correct status for Dean review'}), 400
        
        # Get review data
        data = request.get_json()
        print(f"DEBUG: Raw request data = {data}")
        print(f"DEBUG: Application type = {application.application_type}")
        
        decision = data.get('decision')
        comments = data.get('comments', '').strip()
        
        if not decision or decision not in ['approve', 'reject']:
            return jsonify({'success': False, 'error': 'Invalid decision'}), 400
        
        # Comments mandatory only for rejection, optional for approval
        if decision == 'reject' and not comments:
            return jsonify({'success': False, 'error': 'Comments are required for rejection'}), 400
        
        # Update application with Dean's review
        application.dean_comments = comments
        application.dean_reviewed_at = now_ist()
        
        # Save snapshot of dean's signature and name at approval time
        dean_user = User.query.filter_by(role='dean').first()
        if dean_user:
            # Read signature file as binary data
            signature_binary = None
            if dean_user.signature:
                try:
                    signature_path = os.path.join(get_user_folder(dean_user.id, 'signatures'), dean_user.signature)
                    if os.path.exists(signature_path):
                        with open(signature_path, 'rb') as f:
                            signature_binary = f.read()
                except Exception as e:
                    print(f"Warning: Could not read Dean signature: {e}")
            application.dean_signature_snapshot = signature_binary
            application.dean_name_snapshot = dean_user.name
        
        if decision == 'approve':
            # Handle amounts based on application type
            if application.application_type == 'publication_incentive':
                # 3-tier amounts for publication incentive
                amount_first = data.get('amount_first', 0)
                amount_corresponding = data.get('amount_corresponding', 0)
                amount_coauthors = data.get('amount_coauthors', 0)
                
                # Debug: Print what we received
                print(f"DEBUG: Received amounts - First: {amount_first}, Corr: {amount_corresponding}, Co: {amount_coauthors}")
                print(f"DEBUG: Full data received: {data}")
                
                # Validate that at least ONE amount is provided (the applicant's author type)
                total_amount = float(amount_first or 0) + float(amount_corresponding or 0) + float(amount_coauthors or 0)
                if total_amount == 0:
                    return jsonify({'success': False, 'error': 'At least one amount must be provided for the applicant\'s author type'}), 400
                
                application.dean_first_author_amount = float(amount_first or 0)
                application.dean_corresponding_author_amount = float(amount_corresponding or 0)
                application.dean_coauthor_amount = float(amount_coauthors or 0)
                
                # Debug: Print what we're setting
                print(f"DEBUG: Set amounts - First: {application.dean_first_author_amount}, Corr: {application.dean_corresponding_author_amount}, Co: {application.dean_coauthor_amount}")
            else:
                # Single amount for conference or journal support
                amount = data.get('amount', 0)
                
                if not amount:
                    return jsonify({'success': False, 'error': 'Amount is required'}), 400
                
                application.dean_recommended_amount = float(amount)
            
            application.dean_approved = True
            application.status = 'dean_approved'
            
            # Notify applicant
            create_notification(
                user_id=application.applicant_id,
                title='Application Approved by Dean',
                message=f'Your application {application.tracking_id} has been approved by the Dean R&D and forwarded to the Principal for final approval.',
                notification_type='success'
            )
            
            # Notify Principal
            principal_user = User.query.filter_by(role='principal').first()
            if principal_user:
                create_notification(
                    user_id=principal_user.id,
                    title='New Application for Approval',
                    message=f'Application {application.tracking_id} has been approved by Dean R&D and requires your approval.',
                    notification_type='info'
                )
        else:
            # Reject application
            application.dean_approved = False
            application.status = 'rejected'
            
            # Notify applicant
            create_notification(
                user_id=application.applicant_id,
                title='Application Rejected by Dean',
                message=f'Your application {application.tracking_id} has been rejected by the Dean R&D. Please review the Dean\'s comments.',
                notification_type='danger'
            )
        
        db.session.commit()
        
        # Create audit log
        create_audit_log(
            action='dean_review',
            target_type='application',
            target_id=application.id,
            details=f'Reviewed application {application.tracking_id}: {decision}',
            user_id=current_user.id
        )
        
        return jsonify({
            'success': True,
            'message': f'Review submitted successfully. Application {decision}d.',
            'status': application.status
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/dean/calculate-incentive-amounts/<int:app_id>', methods=['GET'])
@login_required
@role_required('dean')
def calculate_incentive_amounts(app_id):
    """Calculate recommended incentive amounts based on SJEC Circular 2025/091 formulas"""
    from models import ApplicationForm, IncentiveConfig
    
    try:
        application = ApplicationForm.query.get_or_404(app_id)
        config = IncentiveConfig.query.first()
        
        if not config:
            return jsonify({'success': False, 'error': 'Incentive configuration not found'}), 404
        
        if application.application_type != 'publication_incentive':
            return jsonify({'success': False, 'error': 'This route is only for publication incentive applications'}), 400
        
        # Use new SJEC circular calculation functions
        cash_award = calculate_cash_award(application, config)
        epp_points = calculate_epp_points(application, config)
        
        return jsonify({
            'success': True,
            'author_type': application.author_type,
            'amounts': {
                'first_author': cash_award['first_author'],
                'corresponding_author': cash_award['corresponding_author'],
                'coauthor': cash_award['coauthor'],
                'total': cash_award['total']
            },
            'epp_points': epp_points,
            'calculation_details': {
                'publication_category': application.publication_category,
                'is_national_international': application.is_national_international,
                'quartile': application.journal_quartile,
                'impact_factor': application.journal_impact_factor,
                'num_first_authors_sjec': application.num_first_authors_sjec,
                'num_corresponding_authors_sjec': application.num_corresponding_authors_sjec,
                'num_coauthors_sjec': application.num_coauthors_sjec
            },
            # Include application data for modal display
            'application': {
                'publication_category': application.publication_category,
                'is_national_international': application.is_national_international,
                'journal_quartile': application.journal_quartile,
                'journal_impact_factor': application.journal_impact_factor,
                'indexing_type': application.indexing_type,
                'num_first_authors_sjec': application.num_first_authors_sjec,
                'num_corresponding_authors_sjec': application.num_corresponding_authors_sjec,
                'num_coauthors_sjec': application.num_coauthors_sjec,
                'author_type': application.author_type,
                'title_of_paper': application.title_of_paper
            }
        })
        
    except Exception as e:
        print(f"ERROR in calculate_incentive_amounts: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/dean/calculate-amounts/<int:app_id>', methods=['GET'])
@login_required
@role_required('dean')
def calculate_circular_amounts(app_id):
    """Calculate cash award and EPP based on SJEC circular formulas"""
    from models import ApplicationForm, IncentiveConfig
    
    try:
        application = ApplicationForm.query.get_or_404(app_id)
        config = IncentiveConfig.query.first()
        
        if not config:
            return jsonify({'success': False, 'error': 'Incentive configuration not found'}), 404
        
        if application.application_type != 'publication_incentive':
            return jsonify({'success': False, 'error': 'Calculation only available for publication incentive applications'}), 400
        
        # Calculate amounts using circular formulas
        cash_award = calculate_cash_award(application, config)
        epp_points = calculate_epp_points(application, config)
        
        return jsonify({
            'success': True,
            'amounts': {
                'first_author': cash_award['first_author'],
                'corresponding_author': cash_award['corresponding_author'],
                'coauthor': cash_award['coauthor'],
                'total': cash_award['total']
            },
            'epp_points': epp_points,
            'calculation_details': {
                'quartile': application.quartile,
                'impact_factor': application.impact_factor,
                'is_first_author': application.is_first_author,
                'is_corresponding_author': application.is_corresponding_author,
                'num_coauthors': application.num_coauthors
            }
        })
        
    except Exception as e:
        print(f"ERROR in calculate_circular_amounts: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/dean/all-applications')
@login_required
@role_required('dean')
def dean_applications():
    """Dean view of all applications"""
    from models import ApplicationForm
    
    # Get filter parameters
    status_filter = request.args.get('status', 'all')
    app_type_filter = request.args.get('type', 'all')
    
    # Base query
    query = ApplicationForm.query
    
    # Apply filters
    if status_filter != 'all':
        if status_filter == 'principal_approved':
            # Filter for applications where principal approved
            # Could be: 1) Only principal approved (status=dean_approved) OR 2) Both approved (status=approved)
            query = query.filter(ApplicationForm.principal_approved == True)
        elif status_filter == 'director_approved':
            # Filter for applications where director approved
            # Could be: 1) Only director approved (status=dean_approved) OR 2) Both approved (status=approved)
            query = query.filter(ApplicationForm.director_approved == True)
        else:
            # Standard status filter
            query = query.filter_by(status=status_filter)
    
    if app_type_filter != 'all':
        query = query.filter_by(application_type=app_type_filter)
    
    # Get all applications ordered by creation date
    applications = query.order_by(ApplicationForm.created_at.desc()).all()
    
    # Statistics
    total_applications = ApplicationForm.query.count()
    pending_acm = ApplicationForm.query.filter(
        ApplicationForm.status.in_(['submitted', 'acm_review'])
    ).count()
    acm_approved = ApplicationForm.query.filter_by(status='acm_approved').count()
    dean_approved = ApplicationForm.query.filter_by(dean_approved=True).count()
    rejected = ApplicationForm.query.filter_by(status='rejected').count()
    
    return render_template('dean_all_applications.html',
                         applications=applications,
                         total_applications=total_applications,
                         pending_acm=pending_acm,
                         acm_approved=acm_approved,
                         dean_approved=dean_approved,
                         rejected=rejected,
                         status_filter=status_filter,
                         app_type_filter=app_type_filter)


@app.route('/dean/incentive-config', methods=['GET', 'POST'])
@login_required
@role_required('dean')
def dean_incentive_config():
    """Dean configuration panel for incentive amounts and EPP"""
    from models import IncentiveConfig
    
    # Get or create config
    config = IncentiveConfig.query.first()
    if not config:
        config = IncentiveConfig(id=1)
        db.session.add(config)
        db.session.commit()
    
    if request.method == 'POST':
        try:
            # Update Q1 values
            config.q1_base_amount = float(request.form.get('q1_base_amount', 30000))
            config.q1_if_multiplier = float(request.form.get('q1_if_multiplier', 2000))
            config.q1_max_amount = float(request.form.get('q1_max_amount', 40000))
            config.q1_epp_fa = int(request.form.get('q1_epp_fa', 5))
            config.q1_epp_ca = int(request.form.get('q1_epp_ca', 5))
            config.q1_epp_both = int(request.form.get('q1_epp_both', 10))
            
            # Update Q2 values
            config.q2_base_amount = float(request.form.get('q2_base_amount', 20000))
            config.q2_if_multiplier = float(request.form.get('q2_if_multiplier', 2000))
            config.q2_max_amount = float(request.form.get('q2_max_amount', 30000))
            config.q2_epp_fa = int(request.form.get('q2_epp_fa', 4))
            config.q2_epp_ca = int(request.form.get('q2_epp_ca', 4))
            config.q2_epp_both = int(request.form.get('q2_epp_both', 8))
            
            # Update Q3 values
            config.q3_base_amount = float(request.form.get('q3_base_amount', 10000))
            config.q3_if_multiplier = float(request.form.get('q3_if_multiplier', 2000))
            config.q3_max_amount = float(request.form.get('q3_max_amount', 20000))
            config.q3_epp_fa = int(request.form.get('q3_epp_fa', 3))
            config.q3_epp_ca = int(request.form.get('q3_epp_ca', 3))
            config.q3_epp_both = int(request.form.get('q3_epp_both', 6))
            
            # Update Q4 values
            config.q4_base_amount = float(request.form.get('q4_base_amount', 5000))
            config.q4_if_multiplier = float(request.form.get('q4_if_multiplier', 2000))
            config.q4_max_amount = float(request.form.get('q4_max_amount', 10000))
            config.q4_epp_fa = int(request.form.get('q4_epp_fa', 2))
            config.q4_epp_ca = int(request.form.get('q4_epp_ca', 2))
            config.q4_epp_both = int(request.form.get('q4_epp_both', 4))
            
            # Update Scopus/WoS values
            config.scopus_wos_amount = float(request.form.get('scopus_wos_amount', 5000))
            config.scopus_wos_epp_fa = int(request.form.get('scopus_wos_epp_fa', 1))
            config.scopus_wos_epp_ca = int(request.form.get('scopus_wos_epp_ca', 1))
            config.scopus_wos_epp_both = int(request.form.get('scopus_wos_epp_both', 2))
            
            # Update conference and other values
            config.conference_amount = float(request.form.get('conference_amount', 5000))
            config.annual_rrf_amount = float(request.form.get('annual_rrf_amount', 10000))
            config.epp_to_inr_rate = float(request.form.get('epp_to_inr_rate', 2000))
            
            config.updated_by = current_user.id
            config.updated_at = now_ist()
            
            db.session.commit()
            
            # Create audit log
            create_audit_log(
                action='update_incentive_config',
                target_type='config',
                target_id=config.id,
                details='Updated incentive configuration and EPP values',
                user_id=current_user.id
            )
            
            flash('Incentive configuration updated successfully!', 'success')
            return redirect(url_for('dean_incentive_config'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating configuration: {str(e)}', 'danger')
    
    return render_template('dean-incentive-config.html', config=config)


@app.route('/dean/annual-dispersal', methods=['GET', 'POST'])
@login_required
@role_required('dean')
def dean_annual_dispersal():
    """Dean performs annual RRF dispersal to all faculty"""
    from models import WalletTransaction
    
    # Get all faculty users
    faculties = User.query.filter_by(role='faculty').order_by(User.name).all()
    
    # Get config for default amount
    config = IncentiveConfig.query.first()
    default_amount = config.annual_rrf_amount if config else 10000.0
    
    if request.method == 'POST':
        try:
            dispersal_amount = float(request.form.get('dispersal_amount', default_amount))
            
            if dispersal_amount <= 0:
                flash('Dispersal amount must be greater than zero!', 'danger')
                return redirect(url_for('dean_annual_dispersal'))
            
            # Confirm action
            confirm = request.form.get('confirm')
            if not confirm:
                flash('Please confirm the dispersal action.', 'warning')
                return redirect(url_for('dean_annual_dispersal'))
            
            dispersed_count = 0
            total_amount = 0
            
            # Disperse to each faculty
            for faculty in faculties:
                # RESET wallet balance to new dispersal amount (not accumulate)
                old_balance = faculty.wallet_balance or 0
                faculty.wallet_balance = dispersal_amount
                
                # Create transaction record
                transaction = WalletTransaction(
                    user_id=faculty.id,
                    transaction_type='annual_dispersal',
                    amount=dispersal_amount,
                    balance_after=faculty.wallet_balance,
                    description=f'Annual RRF Dispersal - Wallet reset to Rs {dispersal_amount:,.2f} (previous balance: Rs {old_balance:,.2f})',
                    created_by=current_user.id,
                    created_at=now_ist()
                )
                db.session.add(transaction)
                
                # Notify faculty with email
                create_notification(
                    user_id=faculty.id,
                    title='ðŸ’° Annual RRF Dispersal',
                    message=f'Rs {dispersal_amount:,.2f} has been credited to your Research Reserve Fund wallet. New balance: Rs {faculty.wallet_balance:,.2f}',
                    notification_type='success',
                    send_email=True
                )
                
                dispersed_count += 1
                total_amount += dispersal_amount
            
            db.session.commit()
            
            # Create audit log
            create_audit_log(
                action='annual_rrf_dispersal',
                target_type='wallet',
                target_id=current_user.id,
                details=f'Dispersed Rs {dispersal_amount:,.2f} to {dispersed_count} faculty members. Total: Rs {total_amount:,.2f}',
                user_id=current_user.id
            )
            
            flash(f'âœ… Successfully dispersed Rs {dispersal_amount:,.2f} to {dispersed_count} faculty members!', 'success')
            return redirect(url_for('dean_annual_dispersal'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error during dispersal: {str(e)}', 'danger')
    
    # Calculate total if dispersed
    total_if_dispersed = len(faculties) * default_amount
    
    return render_template('dean_annual_dispersal.html',
                         faculties=faculties,
                         default_amount=default_amount,
                         total_if_dispersed=total_if_dispersed)


# ============================================================================
# FINAL APPROVAL ROUTES (Principal, Director, Dean Secretary)
# ============================================================================

@app.route('/final-approval/pending')
@login_required
@role_required('principal', 'director', 'dean_secretary', 'dean')
def final_approval_pending():
    """Dashboard for final approvers to see dean-approved applications"""
    from models import ApplicationForm, IncentiveConfig
    
    # Get dean-approved applications pending final approval
    applications = ApplicationForm.query.filter_by(
        status='dean_approved',
        dean_approved=True
    ).order_by(ApplicationForm.dean_reviewed_at.desc()).all()
    
    # Get incentive config for reference
    config = IncentiveConfig.query.first()
    
    # Calculate amounts dynamically for each publication incentive application
    app_amounts = {}
    for app in applications:
        if app.application_type == 'publication_incentive':
            # Calculate using new SJEC circular formulas
            award_breakdown = calculate_cash_award(app, config)
            
            # Determine which amount applies to this applicant
            if app.author_type == 'Author':
                applicant_amount = award_breakdown['first_author']
            elif app.author_type == 'Corresponding Author':
                applicant_amount = award_breakdown['corresponding_author']
            elif app.author_type == 'Co-Author':
                applicant_amount = award_breakdown['coauthor']
            else:
                applicant_amount = award_breakdown['first_author']
            
            app_amounts[app.id] = {
                'first_author': award_breakdown['first_author'],
                'corresponding_author': award_breakdown['corresponding_author'],
                'coauthor': award_breakdown['coauthor'],
                'total': award_breakdown['total'],
                'applicant_amount': applicant_amount
            }
    
    # Statistics
    total_pending = len(applications)
    principal_approved_count = ApplicationForm.query.filter(
        ApplicationForm.principal_approved_by.isnot(None)
    ).count()
    director_approved_count = ApplicationForm.query.filter(
        ApplicationForm.director_approved_by.isnot(None)
    ).count()
    wallet_debited_count = ApplicationForm.query.filter_by(wallet_debited=True).count()
    
    return render_template('final-approval-dashboard.html',
                         applications=applications,
                         total_pending=total_pending,
                         principal_approved_count=principal_approved_count,
                         director_approved_count=director_approved_count,
                         wallet_debited_count=wallet_debited_count,
                         config=config,
                         app_amounts=app_amounts)


@app.route('/final-approval/mark-approved/<int:app_id>', methods=['POST'])
@login_required
@role_required('principal', 'director', 'dean_secretary')
def mark_final_approval(app_id):
    """Mark an application as approved by Principal or Director"""
    from models import ApplicationForm, WalletTransaction, IncentiveConfig, Notification
    
    try:
        app_form = ApplicationForm.query.get_or_404(app_id)
        
        # Check if already rejected
        if app_form.principal_rejected or app_form.director_rejected:
            flash('This application has been rejected and cannot be approved.', 'warning')
            return redirect(url_for('final_approval_pending'))
        
        # ===== WALLET BALANCE VALIDATION FOR DEBIT APPLICATIONS =====
        # Only validate for conference_registration and journal_support (which will debit/require funds)
        # Publication incentive is a CREDIT so no validation needed
        if app_form.application_type in ['conference_registration', 'journal_support']:
            config = IncentiveConfig.query.first()
            applicant = app_form.applicant
            
            # Calculate available funds
            rrf_balance = applicant.wallet_balance or 0
            epp_balance = applicant.epp_balance or 0
            epp_to_inr_rate = config.epp_to_inr_rate if config else 2000
            total_available = rrf_balance + (epp_balance * epp_to_inr_rate)
            
            # Calculate required amount
            required_amount = app_form.dean_recommended_amount or 0
            
            # Check if sufficient funds
            if total_available < required_amount:
                flash(
                    f'âŒ Cannot approve: Insufficient wallet balance. '
                    f'Faculty has â‚¹{total_available:,.2f} available (RRF: â‚¹{rrf_balance:,.2f} + EPP: {epp_balance:.1f} points = â‚¹{epp_balance * epp_to_inr_rate:,.2f}), '
                    f'but â‚¹{required_amount:,.2f} is required. Shortfall: â‚¹{required_amount - total_available:,.2f}',
                    'danger'
                )
                return redirect(url_for('final_approval_pending'))
        
        # Determine which role is approving
        approver_role = current_user.role
        
        # For dean_secretary, get the role they're approving as
        approve_as_role = request.form.get('approve_as_role', approver_role)
        
        # Mark approval based on role
        if approve_as_role == 'Principal' or (approver_role == 'principal'):
            if app_form.principal_approved_by:
                flash('Principal approval has already been granted for this application.', 'warning')
                return redirect(url_for('final_approval_pending'))
            
            app_form.principal_approved_by = current_user.id
            app_form.principal_approved_at = now_ist()
            app_form.principal_approved = True
            
            # Save snapshot of principal's signature and name at approval time
            principal_user = User.query.filter_by(role='principal').first()
            if principal_user:
                # Read signature file as binary data
                signature_binary = None
                if principal_user.signature:
                    try:
                        signature_path = os.path.join(get_user_folder(principal_user.id, 'signatures'), principal_user.signature)
                        if os.path.exists(signature_path):
                            with open(signature_path, 'rb') as f:
                                signature_binary = f.read()
                    except Exception as e:
                        print(f"Warning: Could not read Principal signature: {e}")
                app_form.principal_signature_snapshot = signature_binary
                app_form.principal_name_snapshot = principal_user.name
            
            if approver_role == 'dean_secretary':
                approval_msg = f"Dean Secretary {current_user.name} approved as Principal"
            else:
                approval_msg = f"Principal {current_user.name} approved"
            
        elif approve_as_role == 'Director' or (approver_role == 'director'):
            if app_form.director_approved_by:
                flash('Director approval has already been granted for this application.', 'warning')
                return redirect(url_for('final_approval_pending'))
            
            app_form.director_approved_by = current_user.id
            app_form.director_approved_at = now_ist()
            app_form.director_approved = True
            
            # Save snapshot of director's signature and name at approval time
            director_user = User.query.filter_by(role='director').first()
            if director_user:
                # Read signature file as binary data
                signature_binary = None
                if director_user.signature:
                    try:
                        signature_path = os.path.join(get_user_folder(director_user.id, 'signatures'), director_user.signature)
                        if os.path.exists(signature_path):
                            with open(signature_path, 'rb') as f:
                                signature_binary = f.read()
                    except Exception as e:
                        print(f"Warning: Could not read Director signature: {e}")
                app_form.director_signature_snapshot = signature_binary
                app_form.director_name_snapshot = director_user.name
            
            if approver_role == 'dean_secretary':
                approval_msg = f"Dean Secretary {current_user.name} approved as Director"
            else:
                approval_msg = f"Director {current_user.name} approved"
        
        db.session.commit()
        
        # Check if BOTH Principal AND Director have approved
        if app_form.principal_approved_by and app_form.director_approved_by and not app_form.wallet_debited:
            # Process wallet deduction and EPP award
            result = process_wallet_deduction(app_form)
            
            if result['success']:
                flash(f'{approval_msg}. Wallet deduction of Rs {result["amount"]:,.2f} processed successfully. '
                      f'EPP awarded: {result["epp"]} points.', 'success')
            else:
                flash(f'{approval_msg}. However, wallet deduction failed: {result["error"]}', 'warning')
        else:
            flash(f'{approval_msg}. Awaiting approval from other authority.', 'success')
            
            # Send notification to applicant
            notification = Notification(
                user_id=app_form.applicant_id,
                title='Application Approved',
                message=f'Your application {app_form.tracking_id} has been approved by {approve_as_role}. '
                        f'Awaiting final approval from {"Director" if approve_as_role == "Principal" else "Principal"}.',
                type='info'
            )
            db.session.add(notification)
            db.session.commit()
        
        return redirect(url_for('final_approval_pending'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error marking approval: {str(e)}', 'danger')
        return redirect(url_for('final_approval_pending'))


@app.route('/final-approval/reject/<int:app_id>', methods=['POST'])
@login_required
@role_required('principal', 'director', 'dean_secretary')
def reject_final_approval(app_id):
    """Reject an application by Principal or Director"""
    from models import ApplicationForm, Notification
    
    try:
        app_form = ApplicationForm.query.get_or_404(app_id)
        rejection_reason = request.form.get('rejection_reason', '').strip()
        
        if not rejection_reason:
            flash('Rejection reason is required.', 'danger')
            return redirect(url_for('final_approval_pending'))
        
        # Determine which role is rejecting
        approver_role = current_user.role
        
        # For dean_secretary, get the role they're rejecting as
        reject_as_role = request.form.get('reject_as_role', approver_role)
        
        # Mark rejection based on role
        if reject_as_role == 'Principal' or (approver_role == 'principal'):
            if app_form.principal_rejected or app_form.principal_approved_by:
                flash('Principal has already processed this application.', 'warning')
                return redirect(url_for('final_approval_pending'))
            
            app_form.principal_rejected = True
            app_form.principal_rejection_reason = rejection_reason
            app_form.principal_rejected_by = current_user.id
            app_form.principal_rejected_at = now_ist()
            app_form.status = 'rejected'
            if approver_role == 'dean_secretary':
                rejection_msg = f"Application rejected by Dean Secretary {current_user.name} as Principal"
            else:
                rejection_msg = f"Application rejected by Principal {current_user.name}"
            
        elif reject_as_role == 'Director' or (approver_role == 'director'):
            if app_form.director_rejected or app_form.director_approved_by:
                flash('Director has already processed this application.', 'warning')
                return redirect(url_for('final_approval_pending'))
            
            app_form.director_rejected = True
            app_form.director_rejection_reason = rejection_reason
            app_form.director_rejected_by = current_user.id
            app_form.director_rejected_at = now_ist()
            app_form.status = 'rejected'
            if approver_role == 'dean_secretary':
                rejection_msg = f"Application rejected by Dean Secretary {current_user.name} as Director"
            else:
                rejection_msg = f"Application rejected by Director {current_user.name}"
        
        db.session.commit()
        
        # Send notification to applicant
        notification = Notification(
            user_id=app_form.applicant_id,
            title='Application Rejected',
            message=f'Your application {app_form.tracking_id} has been rejected by {reject_as_role}. '
                    f'Reason: {rejection_reason}',
            type='danger'
        )
        db.session.add(notification)
        db.session.commit()
        
        flash(f'{rejection_msg}. The applicant has been notified.', 'success')
        return redirect(url_for('final_approval_pending'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error rejecting application: {str(e)}', 'danger')
        return redirect(url_for('final_approval_pending'))


@app.route('/final-approval/disperse/<int:app_id>', methods=['POST'])
@login_required
@role_required('dean_secretary')
def disperse_final_approval(app_id):
    """Dean Secretary final approval - disperse amount after Principal & Director approval"""
    from models import ApplicationForm, Notification
    from datetime import datetime
    
    try:
        app_form = ApplicationForm.query.get_or_404(app_id)
        
        # Verify both Principal and Director have approved
        if not (app_form.principal_approved_by and app_form.director_approved_by):
            flash('Both Principal and Director must approve before final disbursement.', 'warning')
            return redirect(url_for('final_approval_pending'))
        
        # Verify not already dispersed
        if app_form.dean_sec_approved_by:
            flash('This application has already been processed for final disbursement.', 'info')
            return redirect(url_for('final_approval_pending'))
        
        # Get form data
        approval_date_str = request.form.get('approval_date')
        principal_sign = request.form.get('principal_sign_attached') == 'on'
        director_sign = request.form.get('director_sign_attached') == 'on'
        comments = request.form.get('comments', '').strip()
        
        if not approval_date_str:
            flash('Approval/disbursement date is required.', 'danger')
            return redirect(url_for('final_approval_pending'))
        
        # Parse date
        approval_date = datetime.strptime(approval_date_str, '%Y-%m-%d').date()
        
        # Update application
        app_form.dean_sec_approval_date = approval_date
        app_form.dean_sec_principal_sign_attached = principal_sign
        app_form.dean_sec_director_sign_attached = director_sign
        app_form.dean_sec_comments = comments
        app_form.dean_sec_approved_by = current_user.id
        app_form.dean_sec_approved_at = now_ist()
        
        # Save snapshot of dean secretary's signature and name at approval time
        # Read signature file as binary data
        signature_binary = None
        if current_user.signature:
            try:
                signature_path = os.path.join(get_user_folder(current_user.id, 'signatures'), current_user.signature)
                if os.path.exists(signature_path):
                    with open(signature_path, 'rb') as f:
                        signature_binary = f.read()
            except Exception as e:
                print(f"Warning: Could not read Dean Secretary signature: {e}")
        app_form.dean_sec_signature_snapshot = signature_binary
        app_form.dean_sec_name_snapshot = current_user.name
        
        # Process wallet transaction if not already done
        if not app_form.wallet_debited:
            result = process_wallet_deduction(app_form)
            
            if result['success']:
                flash(f'Final approval processed successfully! Amount dispersed: â‚¹{result["amount"]:,.2f}', 'success')
            else:
                flash(f'Final approval recorded, but wallet processing failed: {result["error"]}', 'warning')
        else:
            # Already processed, just record the final approval details
            db.session.commit()
            flash('Final approval details recorded successfully.', 'success')
        
        # Notify applicant
        notification = Notification(
            user_id=app_form.applicant_id,
            title='Final Approval Completed',
            message=f'Your application {app_form.tracking_id} has received final approval from the Dean Secretary office. '
                    f'The amount has been dispersed to your wallet. Date: {approval_date.strftime("%d %b %Y")}',
            type='success'
        )
        db.session.add(notification)
        db.session.commit()
        
        return redirect(url_for('final_approval_pending'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error processing final approval: {str(e)}', 'danger')
        return redirect(url_for('final_approval_pending'))


# ============================================================================
# WALLET CALCULATION FUNCTIONS (Based on SJEC Circular 2025/091)
# ============================================================================

def calculate_cash_award(app_form, config):
    """
    Calculate cash award for publication incentive based on SJEC Circular 2025/091.
    
    Implementation follows exact formulas from circular:
    - Journals: Base + (IF Ã— Multiplier), capped at Max, then distributed by author formula
    - Books/Book Chapters: Flat rates (National vs International)
    - Conference: Scopus-indexed proceedings = Rs 5,000
    
    Args:
        app_form: ApplicationForm object
        config: IncentiveConfig object
        
    Returns:
        dict: {
            'first_author': float,
            'corresponding_author': float,
            'coauthor': float,
            'total': float
        }
    """
    if app_form.application_type != 'publication_incentive':
        return {'first_author': 0, 'corresponding_author': 0, 'coauthor': 0, 'total': 0}
    
    publication_category = app_form.publication_category
    is_national = (app_form.is_national_international == 'National')
    
    # ========== BOOK / BOOK CHAPTER ==========
    # Flat rates, no division by author type
    if publication_category in ['Book', 'Book Chapter']:
        if publication_category == 'Book':
            total_amount = 10000 if is_national else 20000
        else:  # Book Chapter
            total_amount = 5000 if is_national else 10000
        
        # No author type division for books/chapters - single applicant gets full amount
        return {
            'first_author': total_amount,
            'corresponding_author': 0,
            'coauthor': 0,
            'total': total_amount
        }
    
    # ========== JOURNAL PUBLICATION ==========
    # Use complex formula from circular
    if publication_category == 'Journal':
        quartile = app_form.journal_quartile
        impact_factor = app_form.journal_impact_factor or 0
        
        # Get author counts
        N_FA = app_form.num_first_authors_sjec or 0  # Number of First Authors from SJEC
        N_CA = app_form.num_corresponding_authors_sjec or 0  # Number of Corresponding Authors from SJEC
        N_CoA = app_form.num_coauthors_sjec or 0  # Number of Co-Authors from SJEC
        N = N_CoA  # N in formula refers to co-authors (excluding FA/CA)
        
        # Calculate base award based on quartile
        if quartile == 'Q1':
            base_amount = 30000
            if_multiplier = 2000
            max_amount = 40000
        elif quartile == 'Q2':
            base_amount = 20000
            if_multiplier = 2000
            max_amount = 30000
        elif quartile == 'Q3':
            base_amount = 10000
            if_multiplier = 2000
            max_amount = 20000
        elif quartile == 'Q4':
            base_amount = 5000
            if_multiplier = 2000
            max_amount = 10000
        elif quartile in ['Scopus', 'WoS', 'Scopus/WoS', 'Non-Quartile']:
            # Non-quartile Scopus/WoS indexed = Rs 5,000
            total_amount = 5000
            # Apply distribution formula
            return distribute_journal_amount(total_amount, N_FA, N_CA, N)
        else:
            return {'first_author': 0, 'corresponding_author': 0, 'coauthor': 0, 'total': 0}
        
        # Calculate total award: Base + (IF Ã— Multiplier), capped at Max
        total_award = min(base_amount + (impact_factor * if_multiplier), max_amount)
        
        # Apply distribution formula
        return distribute_journal_amount(total_award, N_FA, N_CA, N)
    
    # Default fallback
    return {'first_author': 0, 'corresponding_author': 0, 'coauthor': 0, 'total': 0}


def distribute_journal_amount(total_amount, N_FA, N_CA, N):
    """
    Distribute journal publication amount using SJEC Circular formulas.
    
    Formulas from circular:
    - If FA and CA are same: S_CA/FA = 100/(1+0.6N)  and  S_CoA = (100-2Ã—S_CA/FA)/N
    - If FA and CA are different: S_CA/FA = 50/(1+0.3N)  and  S_CoA = (100-2Ã—S_CA/FA)/N
    
    Then divide by respective counts (N_FA, N_CA, N_CoA)
    
    Args:
        total_amount: Total amount to distribute
        N_FA: Number of First Authors from SJEC (0 or 1)
        N_CA: Number of Corresponding Authors from SJEC (0 or 1)
        N: Number of Co-Authors from SJEC (0 to 10)
    
    Returns:
        dict: Distribution of amounts
    """
    # Check if FA and CA are the same person (both are 1 and same)
    # In our model, if applicant is both, we track it differently
    # For now, assume if both N_FA=1 and N_CA=1, they might be same or different
    # The circular formula applies when BOTH roles exist
    
    if N_FA == 1 and N_CA == 1:
        # FA and CA exist, assume they are DIFFERENT people
        # Formula: S_CA/FA = 50/(1+0.3N)
        if N > 0:
            share_percent_each = 50 / (1 + 0.3 * N)
            share_coauthor_percent = (100 - 2 * share_percent_each) / N
        else:
            # No co-authors, FA and CA split 50-50
            share_percent_each = 50
            share_coauthor_percent = 0
        
        fa_amount = (share_percent_each / 100) * total_amount / N_FA if N_FA > 0 else 0
        ca_amount = (share_percent_each / 100) * total_amount / N_CA if N_CA > 0 else 0
        coauthor_amount = (share_coauthor_percent / 100) * total_amount / N if N > 0 else 0
        
    elif N_FA == 1 and N_CA == 0:
        # Only FA exists, no CA
        if N > 0:
            # FA gets larger share, co-authors share remainder
            fa_percent = 70  # Approximate based on formula spirit
            coauthor_percent = 30 / N
        else:
            fa_percent = 100
            coauthor_percent = 0
        
        fa_amount = (fa_percent / 100) * total_amount
        ca_amount = 0
        coauthor_amount = (coauthor_percent / 100) * total_amount if N > 0 else 0
        
    elif N_FA == 0 and N_CA == 1:
        # Only CA exists, no FA
        if N > 0:
            # CA gets larger share, co-authors share remainder
            ca_percent = 70
            coauthor_percent = 30 / N
        else:
            ca_percent = 100
            coauthor_percent = 0
        
        fa_amount = 0
        ca_amount = (ca_percent / 100) * total_amount
        coauthor_amount = (coauthor_percent / 100) * total_amount if N > 0 else 0
        
    else:
        # Only co-authors (N_FA=0, N_CA=0, N>0)
        if N > 0:
            coauthor_amount = total_amount / N
        else:
            coauthor_amount = 0
        fa_amount = 0
        ca_amount = 0
    
    return {
        'first_author': round(fa_amount, 2),
        'corresponding_author': round(ca_amount, 2),
        'coauthor': round(coauthor_amount, 2),
        'total': round(total_amount, 2)
    }


def calculate_epp_points(app_form, config):
    """
    Calculate EPP (Earned Publication Points) based on SJEC Circular 2025/091.
    
    EPP Points (from circular):
    - Q1: FA=5, CA=5, Both=10
    - Q2: FA=4, CA=4, Both=8
    - Q3: FA=3, CA=3, Both=6
    - Q4: FA=2, CA=2, Both=4
    - Scopus/WoS (non-quartile): FA=1, CA=1, Both=2
    - Books/Book Chapters: 0 EPP (not mentioned in circular)
    - Co-Author only: 0 EPP
    
    Args:
        app_form: ApplicationForm object
        config: IncentiveConfig object
        
    Returns:
        float: EPP points to award
    """
    if app_form.application_type != 'publication_incentive':
        return 0
    
    # Books and Book Chapters do NOT earn EPP
    if app_form.publication_category in ['Book', 'Book Chapter']:
        return 0
    
    # Only Journals earn EPP
    if app_form.publication_category != 'Journal':
        return 0
    
    quartile = app_form.journal_quartile
    
    # Get author counts
    N_FA = app_form.num_first_authors_sjec or 0
    N_CA = app_form.num_corresponding_authors_sjec or 0
    
    # Co-author only (no FA or CA role) = 0 EPP
    if N_FA == 0 and N_CA == 0:
        return 0
    
    # Determine base EPP based on quartile
    if quartile == 'Q1':
        epp_fa = config.q1_epp_fa  # 5
        epp_ca = config.q1_epp_ca  # 5
        epp_both = config.q1_epp_both  # 10
    elif quartile == 'Q2':
        epp_fa = config.q2_epp_fa  # 4
        epp_ca = config.q2_epp_ca  # 4
        epp_both = config.q2_epp_both  # 8
    elif quartile == 'Q3':
        epp_fa = config.q3_epp_fa  # 3
        epp_ca = config.q3_epp_ca  # 3
        epp_both = config.q3_epp_both  # 6
    elif quartile == 'Q4':
        epp_fa = config.q4_epp_fa  # 2
        epp_ca = config.q4_epp_ca  # 2
        epp_both = config.q4_epp_both  # 4
    elif quartile in ['Scopus', 'WoS', 'Scopus/WoS', 'Non-Quartile']:
        epp_fa = config.scopus_wos_epp_fa  # 1
        epp_ca = config.scopus_wos_epp_ca  # 1
        epp_both = config.scopus_wos_epp_both  # 2
    else:
        return 0
    
    # Determine EPP based on applicant's role
    if N_FA == 1 and N_CA == 1:
        # Both FA and CA roles (might be same person or different)
        # Circular doesn't specify if same person gets both - assume yes
        return epp_both
    elif N_FA == 1:
        # Only FA role
        return epp_fa
    elif N_CA == 1:
        # Only CA role
        return epp_ca
    
    return 0


def process_wallet_deduction(app_form):
    """
    Process wallet transactions based on application type:
    - Conference Registration: DEBIT conference fee
    - Publication Incentive: CREDIT incentive amount
    - Journal Support: NO wallet action (notification only)
    
    Returns:
        dict: {'success': bool, 'amount': float, 'epp': float, 'error': str}
    """
    from models import WalletTransaction, IncentiveConfig, Notification
    
    try:
        config = IncentiveConfig.query.first()
        if not config:
            return {
                'success': False,
                'amount': 0,
                'epp': 0,
                'error': 'Incentive configuration not found. Please contact administrator.'
            }
        
        applicant = app_form.applicant
        
        # ===== PUBLICATION INCENTIVE: CREDIT WALLET =====
        if app_form.application_type == 'publication_incentive':
            # Calculate the EXACT amount for THIS applicant based on their role
            award_breakdown = calculate_cash_award(app_form, config)
            
            # Determine which amount applies to the applicant
            author_type = app_form.author_type
            
            if author_type == 'Author':
                # Applicant is First Author
                incentive_amount = award_breakdown['first_author']
            elif author_type == 'Corresponding Author':
                # Applicant is Corresponding Author
                incentive_amount = award_breakdown['corresponding_author']
            elif author_type == 'Co-Author':
                # Applicant is Co-Author
                incentive_amount = award_breakdown['coauthor']
            else:
                # Fallback
                incentive_amount = award_breakdown['first_author'] or award_breakdown['corresponding_author'] or award_breakdown['coauthor']
            
            # Calculate EPP points based on quartile
            try:
                epp_points = calculate_epp_points(app_form, config)
            except Exception as epp_error:
                # Log error but continue with 0 EPP points
                print(f"Error calculating EPP points: {str(epp_error)}")
                epp_points = 0
            
            # CREDIT the wallet with incentive amount
            applicant.wallet_balance += incentive_amount
            
            # Award EPP points
            if epp_points > 0:
                applicant.epp_balance += epp_points
                app_form.epp_awarded = epp_points
            
            # Create CREDIT transaction record
            transaction = WalletTransaction(
                user_id=applicant.id,
                transaction_type='credit',
                amount=incentive_amount,
                balance_after=applicant.wallet_balance,
                description=f'Publication Incentive - {app_form.tracking_id} ({app_form.author_type})',
                application_id=app_form.id,
                created_by=current_user.id,
                created_at=now_ist()
            )
            db.session.add(transaction)
            
            # Mark as processed
            app_form.wallet_debited = True  # Keeping field name for consistency
            app_form.wallet_transaction_id = transaction.id
            app_form.status = 'approved'
            
            # Notification
            notification_msg = (
                f'ðŸŽ‰ Congratulations! Your publication incentive application {app_form.tracking_id} has been approved. '
                f'â‚¹{incentive_amount:,.2f} has been credited to your Research Reward Fund (RRF) wallet.'
            )
            if epp_points > 0:
                notification_msg += f' | {epp_points} EPP points awarded.'
            
            # Send notification with email
            create_notification(
                user_id=applicant.id,
                title='Publication Incentive Approved - Wallet Credited',
                message=notification_msg,
                notification_type='success',
                send_email=True
            )
            
            db.session.commit()
            
            return {
                'success': True,
                'amount': incentive_amount,
                'epp': epp_points,
                'error': None
            }
        
        # ===== JOURNAL SUPPORT: NOTIFICATION ONLY (NO WALLET ACTION) =====
        elif app_form.application_type == 'journal_support':
            recommended_amount = app_form.dean_recommended_amount or 0
            
            # Mark as approved WITHOUT wallet transaction
            app_form.wallet_debited = True  # Mark as processed
            app_form.status = 'approved'
            
            # Notification: Contact Dean R&D Office
            notification_msg = (
                f'âœ… Your journal support application {app_form.tracking_id} has been approved for â‚¹{recommended_amount:,.2f}. '
                f'This is a notification of approval. Please contact the Dean R&D office to arrange payment for the Article Processing Charge (APC). '
                f'No immediate wallet deduction has been made.'
            )
            
            notification = Notification(
                user_id=applicant.id,
                title='Journal Support Approved - Contact Dean R&D Office',
                message=notification_msg,
                type='info'
            )
            db.session.add(notification)
            db.session.commit()
            
            return {
                'success': True,
                'amount': recommended_amount,
                'epp': 0,
                'error': None
            }
        
        # ===== CONFERENCE REGISTRATION: DEBIT WALLET =====
        elif app_form.application_type == 'conference_registration':
            conference_fee = app_form.dean_recommended_amount or app_form.conference_fee or 0
            epp_points = 0  # No EPP for conference
            
            # Check if wallet balance is sufficient
            current_balance = applicant.wallet_balance
            epp_inr_value = applicant.epp_balance * (config.epp_to_inr_rate if config else 2000)
            total_available = current_balance + epp_inr_value
            
            if total_available < conference_fee:
                return {
                    'success': False,
                    'amount': conference_fee,
                    'epp': 0,
                    'error': f'Insufficient funds. Available: â‚¹{total_available:,.2f}, Required: â‚¹{conference_fee:,.2f}'
                }
            
            # DEBIT from wallet (RRF first, then EPP if needed)
            amount_remaining = conference_fee
            rrf_deducted = 0
            epp_deducted = 0
            
            if current_balance >= amount_remaining:
                # Sufficient RRF balance
                rrf_deducted = amount_remaining
                applicant.wallet_balance -= amount_remaining
                amount_remaining = 0
            else:
                # Use all RRF and some EPP
                rrf_deducted = current_balance
                applicant.wallet_balance = 0
                amount_remaining -= rrf_deducted
                
                # Convert remaining amount to EPP points needed
                epp_to_inr = config.epp_to_inr_rate if config else 2000
                epp_needed = amount_remaining / epp_to_inr
                applicant.epp_balance -= epp_needed
                epp_deducted = epp_needed
            
            # Create DEBIT transaction record
            transaction = WalletTransaction(
                user_id=applicant.id,
                transaction_type='debit',
                amount=conference_fee,
                balance_after=applicant.wallet_balance,
                description=f'Conference Registration - {app_form.tracking_id}',
                application_id=app_form.id,
                created_by=current_user.id,
                created_at=now_ist()
            )
            db.session.add(transaction)
            
            # Mark as processed
            app_form.wallet_debited = True
            app_form.wallet_transaction_id = transaction.id
            app_form.status = 'approved'
            
            # Notification
            notification_msg = (
                f'âœ… Your conference registration application {app_form.tracking_id} has been approved. '
                f'â‚¹{conference_fee:,.2f} has been debited from your wallet.'
            )
            
            if rrf_deducted > 0:
                notification_msg += f' | RRF deducted: â‚¹{rrf_deducted:,.2f}'
            
            if epp_deducted > 0:
                notification_msg += f' | EPP used: {epp_deducted:.2f} points (â‚¹{epp_deducted * (config.epp_to_inr_rate if config else 2000):,.2f})'
            
            # Send notification with email
            create_notification(
                user_id=applicant.id,
                title='Conference Registration Approved - Wallet Debited',
                message=notification_msg,
                notification_type='success',
                send_email=True
            )
            
            db.session.commit()
            
            return {
                'success': True,
                'amount': conference_fee,
                'epp': 0,
                'error': None
            }
        
        # ===== FALLBACK FOR UNKNOWN TYPE =====
        else:
            return {
                'success': False,
                'amount': 0,
                'epp': 0,
                'error': f'Unknown application type: {app_form.application_type}'
            }
        
    except Exception as e:
        db.session.rollback()
        return {
            'success': False,
            'amount': 0,
            'epp': 0,
            'error': str(e)
        }


# ============================================================================
# PRINCIPAL APPLICATION ROUTES
# ============================================================================

@app.route('/principal/view-applications')
@login_required
@role_required('principal', 'director')
def principal_view_applications():
    """Principal view for Dean-approved applications"""
    from models import ApplicationForm
    
    # Get all dean-approved applications
    applications = ApplicationForm.query.filter_by(status='dean_approved')\
        .order_by(ApplicationForm.created_at.desc()).all()
    
    # Statistics
    total_dean_approved = len(applications)
    
    return render_template('dean_all_applications.html',
                         applications=applications,
                         total_applications=total_dean_approved,
                         pending_acm=0,
                         acm_approved=0,
                         dean_approved=total_dean_approved,
                         rejected=0,
                         status_filter='dean_approved',
                         app_type_filter='all',
                         is_principal_view=True)


# ============================================================================
# CIRCULAR ROUTES
# ============================================================================

@app.route('/circulars')
def view_circulars():
    """View all circulars (accessible to all users, even before login)"""
    from models import Circular
    
    circulars = Circular.query.filter_by(is_active=True)\
        .order_by(Circular.created_at.desc()).all()
    
    return render_template('circulars.html', circulars=circulars)


@app.route('/circulars/<int:circular_id>')
def view_circular_detail(circular_id):
    """View single circular detail"""
    from models import Circular, CircularRead
    
    circular = Circular.query.get_or_404(circular_id)
    
    if not circular.is_active:
        flash('This circular is no longer active.', 'warning')
        return redirect(url_for('view_circulars'))
    
    # Mark as read if user is authenticated
    if current_user.is_authenticated:
        # Check if already read
        existing_read = CircularRead.query.filter_by(
            circular_id=circular_id,
            user_id=current_user.id
        ).first()
        
        if not existing_read:
            # Create read record
            new_read = CircularRead(
                circular_id=circular_id,
                user_id=current_user.id
            )
            db.session.add(new_read)
            try:
                db.session.commit()
            except:
                db.session.rollback()
    
    return render_template('circular_detail.html', circular=circular)


@app.route('/circulars/upload', methods=['GET', 'POST'])
@login_required
@role_required('dean')
def upload_circular():
    """Dean upload circular form"""
    from models import Circular
    import os
    from werkzeug.utils import secure_filename
    
    if request.method == 'POST':
        title = request.form.get('title', '').strip()
        subject = request.form.get('subject', '').strip()
        body = request.form.get('body', '').strip()
        
        # Validation - only title is required
        if not title:
            flash('Title is required.', 'danger')
            return redirect(url_for('upload_circular'))
        
        # Handle file upload
        file = request.files.get('file')
        file_path = None
        file_type = None
        
        if file and file.filename:
            filename = secure_filename(file.filename)
            file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            
            if file_ext not in app.config['ALLOWED_CIRCULAR_EXTENSIONS']:
                flash('Invalid file type. Only PDF, JPG, JPEG, and PNG are allowed.', 'danger')
                return redirect(url_for('upload_circular'))
            
            # Create unique filename with timestamp
            timestamp = now_ist().strftime('%Y%m%d_%H%M%S')
            unique_filename = f"circular_{timestamp}_{filename}"
            file_path = os.path.join(app.config['CIRCULARS_FOLDER'], unique_filename)
            
            # Save file
            file.save(file_path)
            file_type = file_ext
            
            # Store relative path for database
            file_path = os.path.join('uploads', 'circulars', unique_filename)
        
        # Create circular
        circular = Circular(
            title=title,
            subject=subject,
            body=body,
            file_path=file_path,
            file_type=file_type,
            uploaded_by=current_user.id
        )
        
        db.session.add(circular)
        db.session.commit()
        
        # Create audit log
        audit = AuditLog(
            user_id=current_user.id,
            action='Upload Circular',
            target_type='circular',
            target_id=circular.id,
            details=f'Uploaded circular: {title}',
            ip_address=request.remote_addr
        )
        db.session.add(audit)
        db.session.commit()
        
        # Notify all users
        all_users = User.query.filter(User.id != current_user.id).all()
        for user in all_users:
            notification = Notification(
                user_id=user.id,
                title='New Circular',
                message=f'New circular posted: {title}',
                type='info'
            )
            db.session.add(notification)
        
        db.session.commit()
        
        flash(f'Circular "{title}" uploaded successfully!', 'success')
        return redirect(url_for('view_circulars'))
    
    return render_template('upload-circular.html')


@app.route('/circulars/<int:circular_id>/delete', methods=['POST'])
@login_required
@role_required('dean')
def delete_circular(circular_id):
    """Deactivate a circular (soft delete) and remove associated file"""
    from models import Circular
    import os
    
    circular = Circular.query.get_or_404(circular_id)
    
    # Delete the physical file if it exists
    if circular.file_path:
        try:
            file_full_path = os.path.join(app.root_path, circular.file_path)
            if os.path.exists(file_full_path):
                os.remove(file_full_path)
        except Exception as e:
            # Log the error but continue with soft delete
            print(f"Error deleting file: {e}")
    
    # Soft delete by setting is_active to False
    circular.is_active = False
    db.session.commit()
    
    # Create audit log
    audit = AuditLog(
        user_id=current_user.id,
        action='Delete Circular',
        target_type='circular',
        target_id=circular.id,
        details=f'Deactivated circular: {circular.title}',
        ip_address=request.remote_addr
    )
    db.session.add(audit)
    db.session.commit()
    
    flash('Circular and its attachment have been removed.', 'success')
    return redirect(url_for('view_circulars'))


@app.route('/circulars/<int:circular_id>/download')
def download_circular(circular_id):
    """Download circular file"""
    from models import Circular
    import os
    
    circular = Circular.query.get_or_404(circular_id)
    
    if not circular.file_path:
        flash('No file attached to this circular.', 'warning')
        return redirect(url_for('view_circular_detail', circular_id=circular_id))
    
    file_path = os.path.join(app.root_path, circular.file_path)
    
    if not os.path.exists(file_path):
        flash('File not found.', 'danger')
        return redirect(url_for('view_circular_detail', circular_id=circular_id))
    
    return send_file(file_path, as_attachment=True)


# ============================================================================
# HEAD OF RESEARCH (HoR) MANAGEMENT ROUTES
# ============================================================================

@app.route('/head-of-research/manage', methods=['GET', 'POST'])
@login_required
@role_required('dean')
def manage_head_of_research():
    """Dean manages Head of Research appointments"""
    from models import Department, User, Notification
    from datetime import datetime, date
    
    if request.method == 'POST':
        action = request.form.get('action')
        department_id = request.form.get('department_id')
        
        dept = Department.query.get_or_404(department_id)
        
        if action == 'appoint' or action == 'change':
            faculty_id = request.form.get('faculty_id')
            
            faculty = User.query.get_or_404(faculty_id)
            
            # Update department
            dept.head_of_research_id = faculty_id
            dept.hor_appointed_at = now_ist()
            dept.hor_appointed_by = current_user.id
            
            db.session.commit()
            
            # Send notification to faculty
            notification = Notification(
                user_id=faculty_id,
                title='Appointed as Head of Research',
                message=f'You have been appointed as Head of Research & Development for {dept.name} department. '
                        f'You now have access to department research statistics and publications.',
                type='success'
            )
            db.session.add(notification)
            db.session.commit()
            
            action_text = 'appointed' if action == 'appoint' else 'changed to'
            flash(f'{faculty.name} has been {action_text} Head of Research for {dept.name}.', 'success')
            
        elif action == 'remove':
            if dept.head_of_research_id:
                # Send notification before removing
                notification = Notification(
                    user_id=dept.head_of_research_id,
                    title='Removed from Head of Research',
                    message=f'You have been removed from the position of Head of Research & Development for {dept.name} department.',
                    type='info'
                )
                db.session.add(notification)
                
                dept.head_of_research_id = None
                dept.hor_appointed_at = None
                dept.hor_appointed_by = None
                
                db.session.commit()
                
                flash(f'Head of Research removed from {dept.name}.', 'success')
        
        return redirect(url_for('manage_head_of_research'))
    
    # GET request - show management page
    departments = Department.query.order_by(Department.name).all()
    today = date.today()
    
    return render_template('manage_head_of_research.html',
                         departments=departments,
                         today=today)


@app.route('/api/department-faculty/<int:dept_id>')
@login_required
@role_required('dean')
def get_department_faculty(dept_id):
    """API endpoint to get faculty list for a department"""
    from models import User
    
    faculty = User.query.filter_by(
        department_id=dept_id,
        role='faculty'
    ).order_by(User.name).all()
    
    faculty_list = [{
        'id': f.id,
        'name': f.name,
        'email': f.email
    } for f in faculty]
    
    return jsonify({'faculty': faculty_list})


@app.route('/head-of-research/dashboard')
@login_required
@role_required('faculty')
def head_of_research_dashboard():
    """Dashboard for faculty who are Head of Research"""
    from models import Department, User, Publication
    from sqlalchemy import func
    
    # Check if user is Head of Research
    department = Department.query.filter_by(head_of_research_id=current_user.id).first()
    
    if not department:
        flash('You are not assigned as Head of Research for any department.', 'warning')
        return redirect(url_for('dashboard'))
    
    # Get all faculty in the department
    faculty_members = User.query.filter_by(
        department_id=department.id,
        role='faculty'
    ).order_by(User.name).all()
    
    # Calculate stats for each faculty
    for faculty in faculty_members:
        pubs = Publication.query.filter_by(user_id=faculty.id).all()
        faculty.publication_count = len(pubs)
        faculty.scopus_count = len([p for p in pubs if p.indexing_status == 'Scopus'])
        faculty.citation_count = sum([p.citation_count or 0 for p in pubs])
    
    # Get all publications in the department
    publications = Publication.query.filter_by(department_id=department.id)\
        .order_by(Publication.year.desc()).all()
    
    # Department-level statistics
    total_publications = len(publications)
    scopus_count = len([p for p in publications if p.indexing_status == 'Scopus'])
    total_citations = sum([p.citation_count or 0 for p in publications])
    
    # Year-wise statistics
    from sqlalchemy import case
    year_wise = db.session.query(
        Publication.year,
        func.count(Publication.id).label('count'),
        func.sum(case((Publication.indexing_status == 'Scopus', 1), else_=0)).label('scopus_count')
    ).filter_by(department_id=department.id)\
     .group_by(Publication.year)\
     .order_by(Publication.year.desc()).all()
    
    year_wise_stats = [
        {'year': y.year, 'count': y.count, 'scopus_count': y.scopus_count or 0}
        for y in year_wise
    ]
    
    # Indexing breakdown
    indexing_counts = db.session.query(
        Publication.indexing_status,
        func.count(Publication.id).label('count')
    ).filter_by(department_id=department.id)\
     .group_by(Publication.indexing_status).all()
    
    indexing_stats = {idx.indexing_status: idx.count for idx in indexing_counts}
    
    # Top contributors (faculty sorted by publications)
    top_contributors = sorted(faculty_members, key=lambda f: f.publication_count, reverse=True)[:10]
    
    return render_template('head_of_research_dashboard.html',
                         department=department,
                         faculty_members=faculty_members,
                         faculty_count=len(faculty_members),
                         publications=publications,
                         total_publications=total_publications,
                         scopus_count=scopus_count,
                         total_citations=total_citations,
                         year_wise_stats=year_wise_stats,
                         indexing_stats=indexing_stats,
                         top_contributors=top_contributors,
                         hor_appointed_at=department.hor_appointed_at,
                         hor_term_end=department.hor_term_end_date)


# Initialize Background Scheduler for ACM Auto-Forward Timer
scheduler = BackgroundScheduler()
scheduler.start()

# Add job to check every 30 minutes for applications that need auto-forwarding
scheduler.add_job(
    func=check_and_auto_forward_applications,
    trigger=IntervalTrigger(minutes=30),
    id='acm_auto_forward_job',
    name='Check ACM applications for auto-forward',
    replace_existing=True
)

# Shut down the scheduler when exiting the app
atexit.register(lambda: scheduler.shutdown())

print("âœ… ACM Auto-Forward Scheduler started - checking every 30 minutes")


if __name__ == '__main__':
    with app.app_context():
        try:
            db.create_all()
        except Exception as e:
            print(f"Database tables already exist or error occurred: {e}")
    # host='0.0.0.0' makes the app accessible from other computers on the network
    app.run(host='0.0.0.0', port=5000, debug=True)

