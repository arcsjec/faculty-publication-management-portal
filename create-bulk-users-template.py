import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Users'

# Define headers
headers = [
    'name',
    'email',
    'password',
    'role',
    'department'
]

# Add headers
ws.append(headers)

# Style headers (matching publications template - blue header)
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF', size=11)
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Set column widths
ws.column_dimensions['A'].width = 25  # name
ws.column_dimensions['B'].width = 30  # email
ws.column_dimensions['C'].width = 15  # password
ws.column_dimensions['D'].width = 20  # role
ws.column_dimensions['E'].width = 30  # department

# Add data validation for role column
role_validation = DataValidation(
    type='list',
    formula1='"faculty,hod,dean,principal,vice_principal,admin,director,dean_secretary"',
    allow_blank=False
)
role_validation.error = 'Please select a valid role'
role_validation.errorTitle = 'Invalid Role'
ws.add_data_validation(role_validation)
role_validation.add('D2:D1000')  # Apply to role column

# Add example rows with proper styling
example_rows = [
    ['Dr. John Smith', 'john.smith@sjec.ac.in', 'password123', 'faculty', 'Computer Science'],
    ['Dr. Jane Doe', 'jane.doe@sjec.ac.in', 'password123', 'hod', 'Electronics'],
    ['Dr. Admin User', 'admin.user@sjec.ac.in', 'admin123', 'admin', ''],
    ['Dr. Dean User', 'dean.user@sjec.ac.in', 'dean123', 'dean', '']
]

for row_data in example_rows:
    ws.append(row_data)

# Style example rows
for row in ws.iter_rows(min_row=2, max_row=5):
    for cell in row:
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

# Freeze header row
ws.freeze_panes = 'A2'

# Create Instructions sheet
ws_instructions = wb.create_sheet('Instructions')
ws_instructions.column_dimensions['A'].width = 20
ws_instructions.column_dimensions['B'].width = 60

# Instructions title
ws_instructions['A1'] = 'Bulk User Import Instructions'
ws_instructions['A1'].font = Font(bold=True, size=14, color='FFFFFF')
ws_instructions['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ws_instructions.merge_cells('A1:B1')
ws_instructions['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_instructions.row_dimensions[1].height = 25

# Add instructions
instructions = [
    ('', ''),
    ('Field', 'Description'),
    ('name', 'Full name of the user (Required)'),
    ('email', 'Email address - must be @sjec.ac.in (Required, Unique)'),
    ('password', 'Initial password for the user (Required)'),
    ('role', 'User role (Required) - Select from: faculty, hod, dean, principal, vice_principal, admin, director, dean_secretary'),
    ('department', 'Department name (Required for faculty and hod roles, Leave empty for admin roles)'),
    ('', ''),
    ('Important Notes:', ''),
    ('1.', 'Mandatory fields: name, email, password, role'),
    ('2.', 'Department is REQUIRED for faculty and hod roles'),
    ('3.', 'Department must be LEFT EMPTY for leadership roles (dean, principal, vice_principal, admin, director, dean_secretary)'),
    ('4.', 'Email must be unique and end with @sjec.ac.in'),
    ('5.', 'Department name must match exactly with existing departments in the system'),
    ('6.', 'Users will receive welcome emails at their registered email addresses'),
    ('7.', 'All example rows in the Users sheet can be deleted before uploading'),
    ('', ''),
    ('Valid Roles:', 'faculty, hod, dean, principal, vice_principal, admin, director, dean_secretary'),
]

row_num = 2
for instruction in instructions:
    ws_instructions[f'A{row_num}'] = instruction[0]
    ws_instructions[f'B{row_num}'] = instruction[1]
    
    if row_num == 3:  # Header row
        ws_instructions[f'A{row_num}'].font = Font(bold=True, color='FFFFFF')
        ws_instructions[f'B{row_num}'].font = Font(bold=True, color='FFFFFF')
        ws_instructions[f'A{row_num}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        ws_instructions[f'B{row_num}'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    elif instruction[0] in ['Important Notes:', 'Valid Roles:']:
        ws_instructions[f'A{row_num}'].font = Font(bold=True, size=11)
        ws_instructions.merge_cells(f'A{row_num}:B{row_num}')
    
    ws_instructions[f'A{row_num}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws_instructions[f'B{row_num}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    row_num += 1

# Adjust row heights for instructions
for row in range(2, row_num):
    ws_instructions.row_dimensions[row].height = 30 if row > 10 else 20

# Save workbook
wb.save('static/templates/bulk_users_template.xlsx')
print('✅ Bulk Users Template created successfully!')
print('   Location: static/templates/bulk_users_template.xlsx')
print('   - Users sheet with 4 example rows')
print('   - Instructions sheet with detailed field descriptions')
print('   - Role dropdown validation')
print('   - Professional blue header styling')
