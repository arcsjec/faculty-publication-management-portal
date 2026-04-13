import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Publications'

# Define headers with new columns (no asterisks in Excel headers)
headers = [
    'email',
    'publication_type',  # NEW
    'title',
    'authors_names',
    'journal_conference_name',
    'publisher_name',
    'volume',
    'issue',
    'pages',
    'indexing_status',
    'quartile',
    'impact_factor',
    'isbn',  # NEW
    'edition',  # NEW
    'doi',
    'year',
    'month',
    'citation_count',
    'abstract',
    'bibtex_entry'
]

# Add headers
ws.append(headers)

# Style headers
for cell in ws[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Add example rows
ws.append([
    'faculty@example.com',
    'Journal',
    'Sample Journal Article',
    'John Doe, Jane Smith',
    'International Journal of Computer Science',
    'IEEE',
    '10',
    '2',
    '145-158',
    'SCI',
    'Q1',
    '3.5',
    '',  # isbn (empty for journal)
    '',  # edition (empty for journal)
    '10.1234/ijcs.2024',
    '2024',
    '3',
    '15',
    'This is a sample abstract for a journal article...',
    '@article{doe2024sample, title={Sample Journal Article}, author={Doe, John and Smith, Jane}, journal={International Journal of Computer Science}, year={2024}}'
])

ws.append([
    'faculty@example.com',
    'Book',
    'Advanced Machine Learning Techniques',
    'John Doe',
    '',  # journal_conference_name (empty for book)
    'Springer',
    '',  # volume (empty for book)
    '',  # issue (empty for book)
    '450',  # total pages
    '',  # indexing_status (empty for book)
    '',  # quartile (empty for book)
    '',  # impact_factor (empty for book)
    '978-3-16-148410-0',  # isbn
    '2nd Edition',  # edition
    '',
    '2024',
    '6',
    '',  # citation_count (empty for book)
    'This is a comprehensive guide to advanced machine learning techniques...',
    '@book{doe2024advanced, title={Advanced Machine Learning Techniques}, author={Doe, John}, publisher={Springer}, year={2024}, edition={2nd}}'
])

ws.append([
    'faculty@example.com',
    'Book Chapter',
    'Neural Networks in Natural Language Processing',
    'Jane Smith',
    'Advanced Topics in Artificial Intelligence',  # Book title
    'Elsevier',
    '',
    '',
    '25-45',  # chapter page range
    '',  # indexing_status (empty for book chapter)
    '',  # quartile (empty for book chapter)
    '',  # impact_factor (empty for book chapter)
    '978-3-16-148410-1',  # isbn of the book
    '',  # edition
    '',
    '2024',
    '9',
    '',  # citation_count (empty for book chapter)
    'This chapter explores the application of neural networks in NLP...',
    '@inbook{smith2024neural, title={Neural Networks in Natural Language Processing}, author={Smith, Jane}, booktitle={Advanced Topics in Artificial Intelligence}, publisher={Elsevier}, year={2024}}'
])

# Add data validation for publication_type column
dv = DataValidation(
    type='list',
    formula1='"Journal,Book,Book Chapter"',
    allow_blank=False
)
dv.error = 'Invalid publication type. Must be: Journal, Book, or Book Chapter'
dv.errorTitle = 'Invalid Input'
ws.add_data_validation(dv)
dv.add('B2:B1000')  # Apply to publication_type column

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column].width = min(max_length + 2, 50)

# Add instructions sheet
ws_instructions = wb.create_sheet('Instructions')
ws_instructions.append(['Bulk Import Instructions'])
ws_instructions.append([''])
ws_instructions.append(['MANDATORY Fields (must be filled):'])
ws_instructions.append(['  - email, publication_type, title, authors_names, year'])
ws_instructions.append([''])
ws_instructions.append(['Column Descriptions:'])
ws_instructions.append(['email', 'Faculty email address (must exist in system) - REQUIRED'])
ws_instructions.append(['publication_type', 'Type: Journal, Book, or Book Chapter - REQUIRED'])
ws_instructions.append(['title', 'Publication title - REQUIRED'])
ws_instructions.append(['authors_names', 'Comma-separated author names - REQUIRED'])
ws_instructions.append(['journal_conference_name', 'Journal name (for Journal) or Book title (for Book Chapter)'])
ws_instructions.append(['publisher_name', 'Publisher name'])
ws_instructions.append(['volume', 'Volume number (for Journal only)'])
ws_instructions.append(['issue', 'Issue number (for Journal only)'])
ws_instructions.append(['pages', 'Page range or total pages'])
ws_instructions.append(['indexing_status', 'SCI/SCIE/Scopus/Web of Science/etc. (for Journal only)'])
ws_instructions.append(['quartile', 'Q1/Q2/Q3/Q4 (for Journal only)'])
ws_instructions.append(['impact_factor', 'Impact factor number (for Journal only)'])
ws_instructions.append(['isbn', 'ISBN number (for Book and Book Chapter)'])
ws_instructions.append(['edition', 'Edition (e.g., "2nd Edition") (for Book only)'])
ws_instructions.append(['doi', 'Digital Object Identifier'])
ws_instructions.append(['year', 'Publication year - REQUIRED'])
ws_instructions.append(['month', 'Publication month (1-12)'])
ws_instructions.append(['citation_count', 'Citation count (optional - auto-updates quarterly)'])
ws_instructions.append(['abstract', 'Publication abstract'])
ws_instructions.append(['bibtex_entry', 'BibTeX citation entry'])
ws_instructions.append([''])
ws_instructions.append(['Important Notes:'])
ws_instructions.append(['• Mandatory fields: email, publication_type, title, authors_names, year'])
ws_instructions.append(['• publication_type determines which fields are relevant'])
ws_instructions.append(['• For Journal: Fill indexing_status, quartile, impact_factor'])
ws_instructions.append(['• For Book: Fill isbn, edition'])
ws_instructions.append(['• For Book Chapter: Fill isbn, journal_conference_name (book title)'])
ws_instructions.append(['• Citation count auto-updates quarterly via Google Scholar/Crossref'])
ws_instructions.append(['• All publications will be auto-confirmed upon import'])

# Style instructions sheet
ws_instructions['A1'].font = Font(bold=True, size=14)
ws_instructions['A3'].font = Font(bold=True, color='FF0000')  # Red for mandatory note
ws_instructions['A6'].font = Font(bold=True)
ws_instructions['A27'].font = Font(bold=True)

# Save workbook
wb.save('static/templates/bulk_import_template.xlsx')
print('[SUCCESS] Bulk import template updated successfully')
print('[SUCCESS] Added columns: publication_type, isbn, edition')
print('[SUCCESS] Added data validation for publication_type')
print('[SUCCESS] Added 3 example rows (Journal, Book, Book Chapter)')
print('[SUCCESS] Added instructions sheet')
